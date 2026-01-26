from __future__ import annotations

import os
import re
from dataclasses import dataclass
from functools import lru_cache

from openpyxl import load_workbook


# Token like "ВССТ.КМНК.1" / "СЗИ.2FA.1" / "abc.def_1"
_TOKEN_RE = re.compile(r"[\w-]+(?:\.[\w-]+)+", re.UNICODE)
_DATE_RE = re.compile(r"^\d{1,2}\.\d{1,2}\.\d{2,4}$")


def _looks_like_short_name(token: str) -> bool:
    if not token or "." not in token:
        return False
    if _DATE_RE.match(token):
        return False
    # must contain at least one letter (to avoid 31.12.2024 / 0.3 etc.)
    return any(ch.isalpha() for ch in token)


@dataclass(frozen=True)
class CellToken:
    sheet: str
    coordinate: str
    token: str
    text: str


@dataclass(frozen=True)
class IndexKbTemplate:
    path: str
    mtime_ns: int
    sheet_names: list[str]
    tokens_by_sheet: dict[str, list[CellToken]]


@lru_cache(maxsize=8)
def load_index_kb_template(path: str, mtime_ns: int) -> IndexKbTemplate:
    wb = load_workbook(path, read_only=True, data_only=False)
    tokens_by_sheet: dict[str, list[CellToken]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        hits: list[CellToken] = []
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                text = v.replace("\n", " ").strip()
                if not text:
                    continue
                for tok in _TOKEN_RE.findall(text):
                    if _looks_like_short_name(tok):
                        hits.append(CellToken(sheet=sheet_name, coordinate=cell.coordinate, token=tok, text=text))
        tokens_by_sheet[sheet_name] = hits

    return IndexKbTemplate(path=path, mtime_ns=mtime_ns, sheet_names=list(wb.sheetnames), tokens_by_sheet=tokens_by_sheet)


def get_index_kb_template(path: str) -> IndexKbTemplate:
    st = os.stat(path)
    return load_index_kb_template(path, st.st_mtime_ns)

