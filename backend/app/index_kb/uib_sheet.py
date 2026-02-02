from __future__ import annotations

import os
import re
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from functools import lru_cache

from openpyxl import load_workbook
from sqlalchemy import and_, case, func
from sqlalchemy.orm import Session, joinedload

from app.db.models import (
    Artifact,
    ArtifactLevel,
    ArtifactLevelItem,
    FileVersion,
    IndexKbManualValue,
    IndexKbTemplateRow,
    OrgArtifact,
    OrgArtifactReviewStatus,
    OrgArtifactStatus,
    Organization,
    User,
)


UIB_SHEET_NAME = "Управление ИБ"

# Token like "СУИБ.УР.1"
_TOKEN_RE = re.compile(r"^[^\d\s][\w-]+(?:\.[\w-]+)+$", re.UNICODE)


@dataclass(frozen=True)
class UibRow:
    kind: str  # group|item
    row_key: str
    title: str
    short_name: str
    group_code: str  # short code of current group section (e.g. "СУИБ", "УР")


@dataclass(frozen=True)
class UibRowView:
    row: UibRow
    is_auto: bool
    kb3: float | None
    kb2: float | None
    kb1: float | None
    source: str  # auto|manual|нет артефакта
    updated_at: object | None
    updated_by: str


@dataclass(frozen=True)
class UibTemplate:
    path: str
    mtime_ns: int
    header_row: int
    col_title: int  # B
    col_short: int  # C
    rows: list[UibRow]


def _find_uib_sheet_name(wb) -> str:
    # Prefer exact match, fallback by substring.
    if UIB_SHEET_NAME in wb.sheetnames:
        return UIB_SHEET_NAME
    for s in wb.sheetnames:
        if "Управление" in s:
            return s
    return wb.sheetnames[0]


def _find_header_row(ws, max_scan_rows: int = 120, max_scan_cols: int = 40) -> int | None:
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        texts = []
        for c in range(1, min(ws.max_column, max_scan_cols) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                texts.append(v)
        joined = " | ".join(texts)
        if "Сокращ" in joined and ("КБ1" in joined or "КБ 1" in joined or "КБ2" in joined or "КБ3" in joined):
            return r
    return None


@lru_cache(maxsize=8)
def _load_uib_template_cached(path: str, mtime_ns: int) -> UibTemplate:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[_find_uib_sheet_name(wb)]
    hdr = _find_header_row(ws)
    if not hdr:
        raise RuntimeError("Не найдена строка заголовка (Сокращенное/КБ1/КБ2/КБ3) на листе 'Управление ИБ'")

    # From analysis: B=title, C=short, row 41 is header. We'll still detect dynamically.
    col_title = 2
    col_short = 3

    rows: list[UibRow] = []
    current_group_title = ""
    current_group_code = ""

    empty_streak = 0
    for r in range(hdr + 1, ws.max_row + 1):
        title_v = ws.cell(r, col_title).value
        short_v = ws.cell(r, col_short).value

        title = title_v.strip() if isinstance(title_v, str) else ""
        short = short_v.strip() if isinstance(short_v, str) else ""
        if short.startswith("=") or title.startswith("="):
            # skip formula blocks above the main table
            continue

        if not title and not short:
            empty_streak += 1
            if empty_streak >= 8:
                break
            continue
        empty_streak = 0

        # Group header rows look like: title "Система управления ИБ", short "СУИБ" (no dot)
        if title and short and "." not in short:
            current_group_title = title
            current_group_code = short
            rows.append(UibRow(kind="group", row_key=f"group:{short}", title=title, short_name=short, group_code=short))
            continue

        # Item rows: have short_name token with dots
        if short and _TOKEN_RE.match(short):
            # stable key: token itself
            rows.append(
                UibRow(
                    kind="item",
                    row_key=short.upper(),
                    title=(title or current_group_title),
                    short_name=short,
                    group_code=current_group_code,
                )
            )

    return UibTemplate(path=path, mtime_ns=mtime_ns, header_row=hdr, col_title=col_title, col_short=col_short, rows=rows)


def get_uib_template(path: str) -> UibTemplate:
    st = os.stat(path)
    return _load_uib_template_cached(path, st.st_mtime_ns)


def _get_template_rev(db: Session, sheet_name: str) -> int:
    rev = db.query(func.max(IndexKbTemplateRow.id)).filter(IndexKbTemplateRow.sheet_name == sheet_name).scalar()
    return int(rev or 0)


def _load_template_rows_from_db(db: Session, sheet_name: str) -> list[UibRow]:
    rows = (
        db.query(IndexKbTemplateRow)
        .filter(IndexKbTemplateRow.sheet_name == sheet_name)
        .order_by(IndexKbTemplateRow.sort_order.asc(), IndexKbTemplateRow.id.asc())
        .all()
    )
    out: list[UibRow] = []
    for r in rows:
        out.append(
            UibRow(
                kind=(r.kind or "item"),
                row_key=(r.row_key or ""),
                title=(r.title or ""),
                short_name=(r.short_name or ""),
                group_code=(r.group_code or ""),
            )
        )
    return out


def get_uib_template_from_db(db: Session) -> UibTemplate | None:
    rows = _load_template_rows_from_db(db, UIB_SHEET_NAME)
    if not rows:
        return None
    rev = _get_template_rev(db, UIB_SHEET_NAME)
    return UibTemplate(path="", mtime_ns=rev, header_row=0, col_title=2, col_short=3, rows=rows)


def get_uib_template_rev(db: Session) -> int:
    return _get_template_rev(db, UIB_SHEET_NAME)


def ensure_uib_template_loaded(db: Session, *, template_path: str | None, force: bool = False) -> int:
    """
    Загружает структуру листа "Управление ИБ" в БД (один раз).
    Возвращает количество строк в шаблоне.
    """
    existing_cnt = db.query(func.count(IndexKbTemplateRow.id)).filter(IndexKbTemplateRow.sheet_name == UIB_SHEET_NAME).scalar()
    if int(existing_cnt or 0) > 0 and not force:
        return int(existing_cnt or 0)

    if not template_path or not os.path.exists(template_path):
        raise RuntimeError("Шаблон УИБ не загружен в БД и Excel-эталон не найден. Нужна загрузка шаблона в БД.")

    tpl = get_uib_template(template_path)
    db.query(IndexKbTemplateRow).filter(IndexKbTemplateRow.sheet_name == UIB_SHEET_NAME).delete(synchronize_session=False)
    for i, r in enumerate(tpl.rows, start=1):
        db.add(
            IndexKbTemplateRow(
                sheet_name=UIB_SHEET_NAME,
                sort_order=i,
                kind=r.kind,
                row_key=r.row_key,
                title=r.title,
                short_name=r.short_name,
                group_code=r.group_code,
            )
        )
    db.commit()
    return len(tpl.rows)


def _allowed_artifact_ids_subquery(db: Session, *, org: Organization):
    """
    Эффективный набор артефактов для уровня организации:
    union(items) по уровням sort_order <= выбранного уровня.
    """
    level_id = getattr(org, "artifact_level_id", None)
    if not level_id:
        return None
    lvl = db.get(ArtifactLevel, int(level_id))
    if not lvl:
        return None
    return (
        db.query(ArtifactLevelItem.artifact_id)
        .join(ArtifactLevel, ArtifactLevel.id == ArtifactLevelItem.level_id)
        .filter(ArtifactLevel.sort_order <= lvl.sort_order)
        .subquery()
    )


def _allowed_artifact_ids_subquery_for_level_code(db: Session, *, level_code: str):
    code = (level_code or "").strip().upper()
    lvl = db.query(ArtifactLevel).filter(ArtifactLevel.code == code).one_or_none()
    if not lvl:
        return None
    return (
        db.query(ArtifactLevelItem.artifact_id)
        .join(ArtifactLevel, ArtifactLevel.id == ArtifactLevelItem.level_id)
        .filter(ArtifactLevel.sort_order <= lvl.sort_order)
        .subquery()
    )


def compute_auto_scores(
    db: Session,
    *,
    org: Organization,
    short_names: list[str],
    range_start: datetime | None,
    range_end: datetime | None,
) -> tuple[dict[str, dict[str, float | None]], set[str]]:
    """
    Auto-score rule (requested):
    - KB columns correspond to levels:
      - КБ1 => L1 effective set
      - КБ2 => L2 effective set
      - КБ3 => L3 effective set
    - if date range is set:
      - take latest uploaded file version within range for each org_artifact
      - count as 5 only if that version was audited (approved) within the same range
    - if no date range:
      - use current audited==current rule

    Returns:
    - dict[SHORT_NAME] -> {"kb1": v|None, "kb2": v|None, "kb3": v|None}
      None means "нет артефакта на этом уровне" (n/a for the column).
    - set of tokens present in template but missing in artifacts reference ("нет артефакта").
    """
    if not short_names:
        return {}, set()
    sns = sorted({s.upper() for s in short_names if s})

    # Determine which short_names exist in artifact справочнике.
    existing_all = {
        (sn or "").upper()
        for (sn,) in db.query(Artifact.short_name)
        .filter(Artifact.short_name != "", func.upper(Artifact.short_name).in_(sns))
        .distinct()
        .all()
        if sn
    }
    if not existing_all:
        return {}, set(sns)

    missing_in_artifacts = set(sns) - set(existing_all)

    def _compute_for_level(level_code: str) -> tuple[dict[str, float | None], set[str]]:
        """
        Возвращает:
        - value per sn (None если на уровне нет артефактов с таким сокращением)
        - set существующих на этом уровне sn
        """
        allowed_sub = _allowed_artifact_ids_subquery_for_level_code(db, level_code=level_code)
        if allowed_sub is None:
            return {}, set()

        base_filters = [OrgArtifact.org_id == int(org.id), OrgArtifact.artifact_id.in_(allowed_sub)]

        p_start, p_end = range_start, range_end
        if p_start or p_end:
            fv_conds = []
            if p_start:
                fv_conds.append(FileVersion.created_at >= p_start)
            if p_end:
                fv_conds.append(FileVersion.created_at < p_end)
            sub = (
                db.query(
                    FileVersion.org_artifact_id.label("oa_id"),
                    func.max(FileVersion.id).label("fv_id"),
                )
                .filter(*fv_conds)
                .group_by(FileVersion.org_artifact_id)
                .subquery()
            )
            aud_conds = []
            if p_start:
                aud_conds.append(OrgArtifact.audited_at >= p_start)
            if p_end:
                aud_conds.append(OrgArtifact.audited_at < p_end)
            audited_flag = case(
                (
                    and_(
                        sub.c.fv_id.isnot(None),
                        OrgArtifact.audited_file_version_id.isnot(None),
                        OrgArtifact.audited_file_version_id == sub.c.fv_id,
                        OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                        OrgArtifact.audited_at.isnot(None),
                        *aud_conds,
                    ),
                    1,
                ),
                else_=0,
            )
            rows2 = (
                db.query(
                    func.upper(Artifact.short_name).label("sn"),
                    func.count(Artifact.id).label("cnt"),
                    func.min(audited_flag).label("all_audited"),
                )
                .join(OrgArtifact, OrgArtifact.artifact_id == Artifact.id)
                .outerjoin(sub, sub.c.oa_id == OrgArtifact.id)
                .filter(*base_filters)
                .filter(Artifact.short_name != "", func.upper(Artifact.short_name).in_(sorted(existing_all)))
                .group_by(func.upper(Artifact.short_name))
                .all()
            )
        else:
            audited_flag = case(
                (
                    and_(
                        OrgArtifact.current_file_version_id.isnot(None),
                        OrgArtifact.audited_file_version_id.isnot(None),
                        OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                        OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                    ),
                    1,
                ),
                else_=0,
            )
            rows2 = (
                db.query(
                    func.upper(Artifact.short_name).label("sn"),
                    func.count(Artifact.id).label("cnt"),
                    func.min(audited_flag).label("all_audited"),
                )
                .join(OrgArtifact, OrgArtifact.artifact_id == Artifact.id)
                .filter(*base_filters)
                .filter(Artifact.short_name != "", func.upper(Artifact.short_name).in_(sorted(existing_all)))
                .group_by(func.upper(Artifact.short_name))
                .all()
            )

        existing_lvl: set[str] = set()
        out_lvl: dict[str, float | None] = {}
        for sn_u, cnt, all_audited in rows2:
            sn2 = str(sn_u).upper()
            if int(cnt or 0) <= 0:
                continue
            existing_lvl.add(sn2)
            out_lvl[sn2] = 5.0 if int(all_audited or 0) == 1 else 0.0
        # For tokens present in artifacts but not present on this level => None (n/a).
        for sn2 in existing_all:
            out_lvl.setdefault(sn2, None if sn2 not in existing_lvl else 0.0)
        return out_lvl, existing_lvl

    v1, _ = _compute_for_level("L1")
    v2, _ = _compute_for_level("L2")
    v3, _ = _compute_for_level("L3")

    out: dict[str, dict[str, float | None]] = {}
    for sn2 in existing_all:
        out[sn2] = {"kb1": v1.get(sn2), "kb2": v2.get(sn2), "kb3": v3.get(sn2)}
    return out, missing_in_artifacts


def load_manual_values(db: Session, org_id: int, sheet_name: str) -> dict[str, IndexKbManualValue]:
    rows = (
        db.query(IndexKbManualValue)
        .options(joinedload(IndexKbManualValue.updated_by))
        .filter(IndexKbManualValue.org_id == org_id, IndexKbManualValue.sheet_name == sheet_name)
        .all()
    )
    return {r.row_key: r for r in rows}


@dataclass(frozen=True)
class UibSummaryRow:
    title: str
    short_name: str
    kb3: float | None
    kb2: float | None
    kb1: float | None
    calc_2025: float | None
    current: float | None


def _mean(vals: list[float]) -> float | None:
    vals2 = [float(v) for v in vals if v is not None]
    if not vals2:
        return None
    return sum(vals2) / len(vals2)


def compute_uib_summary(rows: list[UibRowView]) -> list[UibSummaryRow]:
    """
    Итоговая таблица в Excel (скрин): по каждой категории/группе считается AVERAGE по строкам ниже.
    В нашей упрощённой форме КБ1/КБ2/КБ3 и "расчетный/текущий" пока берём одинаково (из value).
    """
    out: list[UibSummaryRow] = []
    cur_group_title = ""
    cur_group_code = ""
    acc1: list[float] = []
    acc2: list[float] = []
    acc3: list[float] = []

    def flush() -> None:
        nonlocal acc1, acc2, acc3, cur_group_title, cur_group_code
        if not cur_group_code:
            return
        m1 = _mean(acc1)
        m2 = _mean(acc2)
        m3 = _mean(acc3)
        out.append(
            UibSummaryRow(
                title=cur_group_title,
                short_name=cur_group_code,
                kb3=m3,
                kb2=m2,
                kb1=m1,
                calc_2025=m3,
                current=m3,
            )
        )
        acc1 = []
        acc2 = []
        acc3 = []

    for rv in rows:
        if rv.row.kind == "group":
            flush()
            cur_group_title = rv.row.title
            cur_group_code = rv.row.short_name
            continue
        if rv.kb1 is not None:
            acc1.append(float(rv.kb1))
        if rv.kb2 is not None:
            acc2.append(float(rv.kb2))
        if rv.kb3 is not None:
            acc3.append(float(rv.kb3))
    flush()
    return out


def build_uib_view(
    db: Session,
    *,
    org_id: int,
    actor: User,
    range_start: datetime | None = None,
    range_end: datetime | None = None,
) -> tuple[Organization, UibTemplate, list[UibRowView]]:
    org = db.get(Organization, org_id)
    if not org:
        raise RuntimeError("Организация не найдена")

    tpl = get_uib_template_from_db(db)
    if not tpl:
        # Important: do NOT parse Excel on first request. Seed must be done via Alembic migration.
        raise RuntimeError("Шаблон УИБ не загружен в БД (нужна seed-миграция)")
    short_names = [r.short_name for r in tpl.rows if r.kind == "item" and r.short_name]

    # Small TTL cache for auto scores to avoid repeated DB load on refresh.
    global _UIB_AUTO_CACHE
    try:
        _UIB_AUTO_CACHE
    except NameError:
        _UIB_AUTO_CACHE = {}  # type: ignore[var-annotated]
    cache_key = (int(org_id), int(tpl.mtime_ns), (range_start.isoformat() if range_start else ""), (range_end.isoformat() if range_end else ""))
    now = time.time()
    cached = _UIB_AUTO_CACHE.get(cache_key)  # type: ignore[name-defined]
    if cached and (now - float(cached[0])) < 15.0:
        auto, missing_in_artifacts = cached[1], cached[2]
    else:
        auto, missing_in_artifacts = compute_auto_scores(db, org=org, short_names=short_names, range_start=range_start, range_end=range_end)
        _UIB_AUTO_CACHE[cache_key] = (now, auto, missing_in_artifacts)  # type: ignore[name-defined]
    manual = load_manual_values(db, org_id, UIB_SHEET_NAME)

    out: list[UibRowView] = []
    for r in tpl.rows:
        if r.kind == "group":
            out.append(UibRowView(row=r, is_auto=True, kb3=None, kb2=None, kb1=None, source="group", updated_at=None, updated_by=""))
            continue

        key = r.row_key
        sn_u = r.short_name.upper()
        mv = manual.get(key)
        if mv:
            out.append(
                UibRowView(
                    row=r,
                    is_auto=False,
                    kb3=float(mv.value),
                    kb2=float(mv.value),
                    kb1=float(mv.value),
                    source="manual",
                    updated_at=mv.updated_at,
                    updated_by=mv.updated_by.login if mv.updated_by else "",
                )
            )
            continue

        if sn_u in missing_in_artifacts:
            # Нет такого сокращения в справочнике артефактов — можно вводить вручную.
            out.append(UibRowView(row=r, is_auto=False, kb3=None, kb2=None, kb1=None, source="нет артефакта", updated_at=None, updated_by=""))
            continue

        v = auto.get(sn_u)
        if isinstance(v, dict):
            out.append(
                UibRowView(
                    row=r,
                    is_auto=True,
                    kb3=v.get("kb3"),
                    kb2=v.get("kb2"),
                    kb1=v.get("kb1"),
                    source="auto",
                    updated_at=None,
                    updated_by="",
                )
            )
        else:
            out.append(UibRowView(row=r, is_auto=False, kb3=None, kb2=None, kb1=None, source="empty", updated_at=None, updated_by=""))

    return org, tpl, out


def upsert_manual_value(
    db: Session,
    *,
    org_id: int,
    sheet_name: str,
    row_key: str,
    value: float,
    actor: User,
) -> None:
    row_key = (row_key or "").strip()
    if not row_key:
        return
    value = float(value)
    if value < 0:
        value = 0.0
    if value > 5:
        value = 5.0

    exists = (
        db.query(IndexKbManualValue)
        .filter(IndexKbManualValue.org_id == org_id, IndexKbManualValue.sheet_name == sheet_name, IndexKbManualValue.row_key == row_key)
        .one_or_none()
    )
    if exists:
        exists.value = value
        exists.updated_by_user_id = actor.id
        # Let ORM store UTC timestamp; display is converted in UI.
        from datetime import datetime

        exists.updated_at = datetime.utcnow()
    else:
        db.add(
            IndexKbManualValue(
                org_id=org_id,
                sheet_name=sheet_name,
                row_key=row_key,
                value=value,
                updated_by_user_id=actor.id,
            )
        )

