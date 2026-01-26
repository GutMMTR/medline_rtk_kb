from __future__ import annotations

import re
from dataclasses import dataclass

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.db.models import Artifact, FileVersion, OrgArtifact, OrgArtifactStatus, User
from app.index_kb.template_loader import get_index_kb_template


_TOKEN_CELL_RE = re.compile(r"^[^\d\s][\w-]+(?:\.[\w-]+)+$", re.UNICODE)


@dataclass(frozen=True)
class ShortNameRollup:
    short_name: str
    total_points: int
    uploaded_points: int
    state: str  # uploaded|partial|missing|unknown
    score_target: float
    score_value: float
    status_text: str
    last_filename: str
    last_uploaded_at: object
    last_uploaded_by: str


def _as_float(v: object) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def build_rollups_for_sheet(db: Session, org_id: int, short_names: list[str]) -> dict[str, ShortNameRollup]:
    """Rollup per short_name across all points (achievement items)."""
    if not short_names:
        return {}

    CreatedBy = User
    rows_raw = (
        db.query(OrgArtifact, Artifact, FileVersion, CreatedBy)
        .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
        .outerjoin(FileVersion, FileVersion.id == OrgArtifact.current_file_version_id)
        .outerjoin(CreatedBy, CreatedBy.id == FileVersion.created_by_user_id)
        .filter(OrgArtifact.org_id == org_id, Artifact.short_name.in_(short_names))
        .all()
    )

    acc: dict[str, dict] = {}
    for (oa, a, fv, created_by) in rows_raw:
        key = (a.short_name or "").upper()
        it = acc.get(key)
        if not it:
            it = {"total": 0, "uploaded": 0, "last_at": None, "last_fn": "", "last_by": ""}
            acc[key] = it
        it["total"] += 1
        if oa.status == OrgArtifactStatus.uploaded:
            it["uploaded"] += 1
        if fv and fv.created_at:
            if it["last_at"] is None or fv.created_at > it["last_at"]:
                it["last_at"] = fv.created_at
                it["last_fn"] = fv.original_filename or ""
                it["last_by"] = created_by.login if created_by else ""

    out: dict[str, ShortNameRollup] = {}
    for sn in short_names:
        key = sn.upper()
        it = acc.get(key) or {"total": 0, "uploaded": 0, "last_at": None, "last_fn": "", "last_by": ""}
        total_points = int(it["total"])
        uploaded_points = int(it["uploaded"])
        if total_points and uploaded_points == total_points:
            state = "uploaded"
        elif uploaded_points > 0:
            state = "partial"
        else:
            state = "missing"

        # score mapping: uploaded => full (target), missing => 0, partial => proportion
        score_target = 0.0
        score_value = 0.0
        status_text = "нет"
        if state == "uploaded":
            status_text = "да"
        elif state == "partial":
            status_text = "частично"

        out[key] = ShortNameRollup(
            short_name=sn,
            total_points=total_points,
            uploaded_points=uploaded_points,
            state=state,
            score_target=score_target,
            score_value=score_value,
            status_text=status_text,
            last_filename=it["last_fn"],
            last_uploaded_at=it["last_at"],
            last_uploaded_by=it["last_by"],
        )
    return out


def fill_workbook_for_org(template_path: str, org_id: int, db: Session) -> tuple[object, dict[str, dict[str, ShortNameRollup]]]:
    """
    Returns openpyxl workbook with filled status inputs (column detected from formula LEFT(...)),
    and rollups per sheet for rendering.
    """
    tpl = get_index_kb_template(template_path)
    wb = load_workbook(template_path, read_only=False, data_only=False)

    # Build unique short_names per sheet (case-sensitive as in DB, but match by upper)
    rollups_by_sheet: dict[str, dict[str, ShortNameRollup]] = {}

    for sheet_name in tpl.sheet_names:
        hits = tpl.tokens_by_sheet.get(sheet_name, [])
        if not hits:
            continue
        # normalize to existing Artifact.short_name where possible
        token_keys = sorted({h.token.upper() for h in hits})
        # Map to actual DB short_name (preserve original case from DB)
        short_name_map: dict[str, str] = {}
        for (sn,) in (
            db.query(Artifact.short_name)
            .filter(Artifact.short_name != "", func.upper(Artifact.short_name).in_(token_keys))
            .distinct()
            .all()
        ):
            if sn:
                short_name_map[sn.upper()] = sn
        short_names = sorted(short_name_map.values())
        if not short_names:
            continue

        # Create rollups and also compute score_target/score_value later per-row
        base_rollups = build_rollups_for_sheet(db, org_id, short_names)
        ws = wb[sheet_name]

        # Determine input col by scanning one formula like VALUE(LEFT(Nxx,3))
        input_col_letter: str | None = None
        for h in hits[:30]:
            rr = ws[h.coordinate].row
            # scan row for formula containing LEFT(
            for cc in range(1, min(ws.max_column, 40) + 1):
                v = ws.cell(row=rr, column=cc).value
                if isinstance(v, str) and "LEFT(" in v and str(rr) in v:
                    # Example: =VALUE(LEFT(N61,3))  -> input col = N
                    try:
                        m = re.search(rf"LEFT\(\s*([A-Z]{{1,3}})\s*{rr}\s*,", v, flags=re.IGNORECASE)
                    except re.error:
                        m = None
                    if m:
                        input_col_letter = m.group(1).upper()
                        break
            if input_col_letter:
                break

        if not input_col_letter:
            # cannot fill; still keep rollups for highlighting
            rollups_by_sheet[sheet_name] = base_rollups
            continue

        input_col = ws[input_col_letter + "1"].column  # type: ignore[attr-defined]

        # For each token hit, set input value on that row based on rollup state.
        for h in hits:
            tok = h.token.upper()
            sn = short_name_map.get(tok)
            if not sn:
                continue
            rr = ws[h.coordinate].row
            ru = base_rollups.get(sn.upper())
            if not ru:
                continue

            # score_target heuristic: K column if numeric, else existing prefix of input cell, else 0
            k_val = _as_float(ws.cell(row=rr, column=11).value)  # K
            score_target = k_val

            if ru.state == "uploaded":
                score_value = score_target
                status_text = "да"
            elif ru.state == "partial" and ru.total_points:
                score_value = round(score_target * (ru.uploaded_points / ru.total_points), 2)
                status_text = "частично"
            else:
                score_value = 0.0
                status_text = "нет"

            # update rollup with scores
            base_rollups[sn.upper()] = ShortNameRollup(
                short_name=ru.short_name,
                total_points=ru.total_points,
                uploaded_points=ru.uploaded_points,
                state=ru.state,
                score_target=score_target,
                score_value=score_value,
                status_text=status_text,
                last_filename=ru.last_filename,
                last_uploaded_at=ru.last_uploaded_at,
                last_uploaded_by=ru.last_uploaded_by,
            )

            # format as in template: "5     -да" / "1,6   - Не ..."
            score_str = str(score_value).replace(".", ",") if (score_value % 1) else str(int(score_value))
            ws.cell(row=rr, column=input_col).value = f"{score_str}     -{status_text}"

        rollups_by_sheet[sheet_name] = base_rollups

    return wb, rollups_by_sheet

