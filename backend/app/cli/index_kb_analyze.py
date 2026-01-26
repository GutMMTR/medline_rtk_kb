from __future__ import annotations

import os
import re
from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook


TOKEN_RE = re.compile(r"[\w-]+(?:\.[\w-]+)+", re.UNICODE)


@dataclass(frozen=True)
class TokenHit:
    sheet: str
    cell: str
    token: str
    snippet: str


def _iter_hits(path: str, limit: int = 100) -> Iterable[TokenHit]:
    wb = load_workbook(path, read_only=True, data_only=False)
    count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                v_norm = v.replace("\n", " ").strip()
                for tok in TOKEN_RE.findall(v_norm):
                    yield TokenHit(sheet=sheet_name, cell=cell.coordinate, token=tok, snippet=v_norm[:120])
                    count += 1
                    if count >= limit:
                        return


def main() -> None:
    path = os.environ.get("INDEX_KB_TEMPLATE_PATH", "/mnt/docs/Эталон_ИндексКБ_шаб.xlsx")
    print("PATH:", path)
    wb = load_workbook(path, read_only=True, data_only=False)
    print("SHEETS:", wb.sheetnames)
    print()
    print("SAMPLE TOKENS:")
    for h in _iter_hits(path, limit=60):
        print(f"{h.sheet}\t{h.cell}\t{h.token}\t{h.snippet}")


if __name__ == "__main__":
    main()

