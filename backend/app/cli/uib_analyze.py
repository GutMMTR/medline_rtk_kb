from __future__ import annotations

import os
import re

from openpyxl import load_workbook


def main() -> None:
    path = os.environ.get("INDEX_KB_TEMPLATE_PATH", "/mnt/docs/Эталон_ИндексКБ_шаб.xlsx")
    wb = load_workbook(path, read_only=True, data_only=False)

    sheet_name = next((s for s in wb.sheetnames if "Управление" in s), wb.sheetnames[0])
    ws = wb[sheet_name]
    print("PATH:", path)
    print("SHEET:", ws.title)
    print("MAX:", ws.max_row, ws.max_column)

    maxr = min(ws.max_row, 600)
    maxc = min(ws.max_column, 80)

    hdr = None
    for r in range(1, maxr + 1):
        cells = [ws.cell(r, c).value for c in range(1, maxc + 1)]
        txt = " | ".join([v.strip() for v in cells if isinstance(v, str) and v.strip()])
        if "Сокращ" in txt and ("КБ1" in txt or "КБ 1" in txt or "КБ2" in txt or "КБ3" in txt):
            hdr = r
            break

    print("HEADER_ROW:", hdr)
    if not hdr:
        return

    def find_col(substr: str) -> int | None:
        for c in range(1, maxc + 1):
            v = ws.cell(hdr, c).value
            if isinstance(v, str) and substr in v:
                return c
        return None

    col_short = find_col("Сокращ")
    col_kb3 = find_col("КБ3")
    col_kb2 = find_col("КБ2")
    col_kb1 = find_col("КБ1") or find_col("КБ 1")
    print("COLS:", {"short": col_short, "kb3": col_kb3, "kb2": col_kb2, "kb1": col_kb1})

    print()
    print("SAMPLE ROWS (A..F):")
    for r in range(hdr + 1, min(hdr + 20, ws.max_row) + 1):
        vals = []
        for c in range(1, 7):
            v = ws.cell(r, c).value
            if v is None:
                vals.append("")
            else:
                vals.append(str(v).strip().replace("\n", " ")[:80])
        print(r, vals)


if __name__ == "__main__":
    main()

