from __future__ import annotations

import os
from dataclasses import dataclass

from openpyxl import load_workbook


TEMPLATE_PATH = os.environ.get("INDEX_KB_TEMPLATE_PATH") or "/mnt/docs/Эталон_ИндексКБ_шаб.xlsx"
SHEET_NAME = "Управление ИБ"


@dataclass(frozen=True)
class Anchor:
    cell: str
    value: str


def _find_anchors(ws, needles: list[str], max_row: int = 200, max_col: int = 120) -> dict[str, list[Anchor]]:
    found: dict[str, list[Anchor]] = {k: [] for k in needles}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row), min_col=1, max_col=min(ws.max_column, max_col)):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            vv = v.strip()
            if not vv:
                continue
            for k in needles:
                if k in vv:
                    found[k].append(Anchor(cell=cell.coordinate, value=vv))
    return found


def _dump_table(ws, top_left: str, width: int = 10, max_rows: int = 40) -> None:
    start = ws[top_left]
    r0, c0 = start.row, start.column
    print(f"\n== DUMP from {top_left} (r={r0}, c={c0}) width={width} rows={max_rows} ==")
    for r in range(r0, r0 + max_rows):
        row_out = []
        for c in range(c0, c0 + width):
            cell = ws.cell(r, c)
            v = cell.value
            if v is None:
                row_out.append("")
            else:
                s = str(v)
                if len(s) > 34:
                    s = s[:31] + "..."
                row_out.append(s)
        if all(not x for x in row_out):
            break
        print(f"{r:>4} | " + " | ".join(f"{x:<34}" for x in row_out))


def main() -> None:
    p = TEMPLATE_PATH
    print("template:", p)
    print("exists:", os.path.exists(p))
    wb = load_workbook(p, data_only=False)
    print("sheets:", wb.sheetnames)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    needles = ["Категория", "Исторические данные", "Расчетный показатель", "Текущий показатель"]
    found = _find_anchors(ws, needles)
    for k in needles:
        print(k, [f"{a.cell}:{a.value}" for a in found[k][:10]])

    # Try to find header "Категория" and dump area around it.
    if found["Категория"]:
        cat = found["Категория"][0].cell
        _dump_table(ws, cat, width=12, max_rows=30)

    # Dump the detail table area referenced by the first summary row (B60..)
    _dump_table(ws, "B60", width=12, max_rows=30)
    # Dump far-right columns around N that feed current indicator (L = VALUE(LEFT(N..,3)))
    _dump_table(ws, "L60", width=8, max_rows=20)

    # Try to locate "Текущий показатель" and dump area around it (a bit to the left).
    if found["Текущий показатель"]:
        cur = ws[found["Текущий показатель"][0].cell]
        # shift left by 6 columns to include KB columns
        left_col = max(1, cur.column - 6)
        top_left = ws.cell(cur.row, left_col).coordinate
        _dump_table(ws, top_left, width=12, max_rows=30)


if __name__ == "__main__":
    main()

