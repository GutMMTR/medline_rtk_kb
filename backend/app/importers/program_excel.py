from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import load_workbook


@dataclass(frozen=True)
class ProgramArtifactRow:
    topic: str
    domain: str
    indicator_name: str
    short_name: str
    kb_level: str
    achievement_text: str
    achievement_item_no: int | None
    achievement_item_text: str
    row_no: int


def _norm_header(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = s.replace("\n", "").replace("\r", "")
    s = re.sub(r"\s+", "", s)
    # убираем пунктуацию/небуквенные символы, чтобы переживать эксельные артефакты
    s = re.sub(r"[^0-9a-zа-яё]+", "", s)
    return s


def _detect_key(norm: str) -> str | None:
    if not norm:
        return None
    # Частичные совпадения — так переживаем переносы строк, лишние префиксы, объединённые ячейки.
    if "тематика" in norm:
        return "topic"
    if "домен" in norm:
        return "domain"
    if ("наименование" in norm and "показател" in norm) or "наименованиепоказателя" in norm:
        return "indicator_name"
    if ("сокращ" in norm and "наимен" in norm) or "сокр" in norm:
        return "short_name"
    if ("уровен" in norm and "кб" in norm) or "уровенькб" in norm:
        return "kb_level"
    if ("артефакт" in norm and "достиж" in norm) or "артефактыдостижения" in norm:
        return "achievement_text"
    return None


_ITEMS_RE = re.compile(r"(?ms)^\s*(\d+)\.\s*(.+?)(?=^\s*\d+\.|\Z)")


def _split_items(text: str) -> list[tuple[int | None, str]]:
    text = (text or "").strip()
    if not text:
        return []
    items = [(int(m.group(1)), (m.group(2) or "").strip()) for m in _ITEMS_RE.finditer(text) if (m.group(2) or "").strip()]
    if items:
        items.sort(key=lambda x: x[0] if x[0] is not None else 0)
        return items
    return [(None, text)]


def parse_program_xlsx(content: bytes, sheet_name_hint: str | None = None) -> list[ProgramArtifactRow]:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = None
    candidates = list(wb.sheetnames)
    if sheet_name_hint:
        for name in candidates:
            if name.strip().lower() == sheet_name_hint.strip().lower():
                sheet = wb[name]
                break
    if sheet is None:
        for name in candidates:
            n = name.strip().lower()
            if "программа" in n or "programma" in n or "program" in n:
                sheet = wb[name]
                break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    # На реальных Excel заголовок часто "многострочный" (объединённые ячейки),
    # поэтому ищем ключевые заголовки в верхнем блоке листа, а не в одной строке.
    required = {"topic", "domain", "short_name", "achievement_text"}
    found: dict[str, list[tuple[int, int]]] = {}

    for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=300, values_only=True), start=1):
        for c_idx, cell in enumerate(row, start=1):
            key = _detect_key(_norm_header(cell))
            if not key:
                continue
            found.setdefault(key, []).append((r_idx, c_idx))

        # быстрый выход: если нашли все required хотя бы раз
        if required.issubset(found.keys()):
            # но всё равно дойдём ещё чуть-чуть ниже, чтобы поймать "нижнюю" строку заголовка,
            # если она разнесена на 2 строки (частая история)
            if r_idx >= 25:
                break

    missing = sorted(required - set(found.keys()))
    if missing:
        raise ValueError(
            "Не нашли строку заголовков на листе. Ожидаем колонки: Тематика, Домен, Сокращенное наименование, Артефакты достижения."
        )

    # Для каждого ключа берём самое "нижнее" вхождение — так мы устойчивы к многострочным шапкам.
    col_map: dict[str, int] = {k: max(found[k], key=lambda rc: rc[0])[1] for k in found.keys()}
    header_row_idx = max(max(found[k], key=lambda rc: rc[0])[0] for k in required)

    def get_cell(row_values: tuple[object, ...], key: str) -> str:
        idx = col_map.get(key)
        if not idx:
            return ""
        v = row_values[idx - 1] if idx - 1 < len(row_values) else None
        return ("" if v is None else str(v)).strip()

    out: list[ProgramArtifactRow] = []
    for r_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        topic = get_cell(row, "topic")
        domain = get_cell(row, "domain")
        indicator_name = get_cell(row, "indicator_name")
        short_name = get_cell(row, "short_name")
        kb_level = get_cell(row, "kb_level")
        achievement_text = get_cell(row, "achievement_text")

        if not any([topic, domain, indicator_name, short_name, kb_level, achievement_text]):
            continue
        if not short_name:
            raise ValueError(f"Пустое 'Сокращенное наименование' в строке {r_idx}.")
        if not achievement_text:
            raise ValueError(f"Пустое 'Артефакты достижения' в строке {r_idx} (short_name={short_name}).")

        for item_no, item_text in _split_items(achievement_text):
            out.append(
                ProgramArtifactRow(
                    topic=topic,
                    domain=domain,
                    indicator_name=indicator_name,
                    short_name=short_name,
                    kb_level=kb_level,
                    achievement_text=achievement_text,
                    achievement_item_no=item_no,
                    achievement_item_text=item_text,
                    row_no=r_idx,
                )
            )
    return out

