from __future__ import annotations

import hashlib
import io
import os
import time
import subprocess
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Dict, Tuple
from urllib.parse import quote, urlencode
from datetime import timedelta

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile, status
from fastapi.responses import HTMLResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.exc import IntegrityError
from sqlalchemy import and_, case, func, insert, or_, select
import sqlalchemy as sa
from sqlalchemy.orm import Session, aliased, joinedload

from app.audit.service import write_audit_log
from app.auth.dependencies import get_current_user, get_user_role_for_org, require_admin
from app.auth.security import JWT_COOKIE_NAME, create_access_token, hash_password, verify_password
from app.core.config import settings
import json

from app.db.models import (
    Artifact,
    ArtifactLevel,
    ArtifactLevelItem,
    ArtifactNode,
    AuditLog,
    FilePreview,
    FileVersion,
    IndexKbTemplateRow,
    OrgIndexKbTarget,
    NextcloudIntegrationSettings,
    NextcloudRemoteFileState,
    OrgArtifact,
    OrgArtifactComment,
    OrgArtifactReviewStatus,
    OrgArtifactStatus,
    Organization,
    Role,
    StoredFile,
    ChatMessage,
    ChatThread,
    ChatThreadRead,
    User,
    UserOrgMembership,
)
from app.db.session import get_db
from app.importers.program_excel import parse_program_xlsx
from app.index_kb.excel_fill import fill_workbook_for_org
from app.index_kb.formula_eval import build_evaluator_from_openpyxl_workbook
from app.index_kb.sheet_render import iter_render_rows
from app.index_kb.template_loader import get_index_kb_template
from app.index_kb.uib_sheet import (
    UIB_SHEET_NAME,
    build_uib_view,
    get_uib_template_from_db,
    get_uib_template_rev,
    upsert_manual_value,
)
from app.index_kb.szi_sheet import SZI_SHEET_NAME, build_szi_view, get_szi_template_from_db, get_szi_template_rev
from app.integrations.nextcloud_dav import NextcloudDavClient, build_webdav_base_url
from app.integrations.nextcloud_sync import sync_from_nextcloud


router = APIRouter()
templates = Jinja2Templates(directory="app/templates")

DEFAULT_ORG_NAME = "Default"


def _require_chat_access(db: Session, user: User, org_id: int, *, allow_customer: bool) -> Role:
    """
    Чат доступен:
    - admin: всегда
    - auditor: как "глобальный аудитор" (см. get_user_role_for_org)
    - customer: только для своих организаций (если allow_customer=True)
    """
    role = get_user_role_for_org(db, user, org_id)
    if role == Role.admin:
        return role
    if role == Role.auditor:
        return role
    if allow_customer and role == Role.customer:
        return role
    raise HTTPException(status_code=403, detail="Нет доступа к чату этой организации")


def _get_or_create_chat_thread(db: Session, *, org_id: int, org_artifact_id: int | None, actor: User) -> ChatThread:
    q = db.query(ChatThread).filter(ChatThread.org_id == int(org_id))
    if org_artifact_id is None:
        q = q.filter(ChatThread.org_artifact_id.is_(None))
    else:
        q = q.filter(ChatThread.org_artifact_id == int(org_artifact_id))
    t = q.one_or_none()
    if t:
        return t
    t = ChatThread(org_id=int(org_id), org_artifact_id=int(org_artifact_id) if org_artifact_id else None, created_by_user_id=actor.id)
    db.add(t)
    db.flush()
    return t


def _chat_message_to_dict(m: ChatMessage) -> dict:
    return {
        "id": int(m.id),
        "thread_id": int(m.thread_id),
        "author_user_id": int(m.author_user_id) if m.author_user_id else None,
        "author_login": (m.author.login if getattr(m, "author", None) else None),
        "body": m.body or "",
        # ISO for JS Date()
        "created_at": (m.created_at.isoformat() if m.created_at else ""),
    }


def _get_chat_unread_for_org(
    db: Session,
    *,
    user: User,
    org_id: int,
    only_org_artifact_ids: list[int] | None = None,
) -> tuple[int, dict[int, int], dict[int, int]]:
    """
    Возвращает:
    - total_unread: суммарно непрочитанных сообщений по организации
    - unread_by_thread_id: {thread_id: unread_cnt}
    - unread_by_org_artifact_id: {org_artifact_id: unread_cnt} (только для тредов по артефактам)
    """
    org_id = int(org_id)
    if org_id <= 0:
        return 0, {}, {}

    t_ids_sub = db.query(ChatThread.id).filter(ChatThread.org_id == org_id)
    if only_org_artifact_ids is not None:
        oa_ids = [int(x) for x in only_org_artifact_ids if int(x) > 0]
        # include "org chat" always if requested list is limited
        t_ids_sub = t_ids_sub.filter(or_(ChatThread.org_artifact_id.is_(None), ChatThread.org_artifact_id.in_(oa_ids)))

    # unread = messages with id > last_read_message_id (per user/thread) and not authored by this user
    unread_rows = (
        db.query(
            ChatMessage.thread_id.label("thread_id"),
            func.count(ChatMessage.id).label("unread_cnt"),
        )
        .join(ChatThread, ChatThread.id == ChatMessage.thread_id)
        .outerjoin(
            ChatThreadRead,
            and_(
                ChatThreadRead.thread_id == ChatThread.id,
                ChatThreadRead.user_id == int(user.id),
            ),
        )
        .filter(ChatMessage.thread_id.in_(t_ids_sub))
        .filter(ChatMessage.author_user_id.is_(None) | (ChatMessage.author_user_id != int(user.id)))
        .filter(ChatMessage.id > func.coalesce(ChatThreadRead.last_read_message_id, 0))
        .group_by(ChatMessage.thread_id)
        .all()
    )
    unread_by_thread_id = {int(tid): int(cnt) for (tid, cnt) in unread_rows}
    total_unread = sum(unread_by_thread_id.values())

    # Map thread->org_artifact_id for artifact threads
    t_pairs = (
        db.query(ChatThread.id, ChatThread.org_artifact_id)
        .filter(ChatThread.id.in_(list(unread_by_thread_id.keys()) or [-1]))
        .all()
    )
    unread_by_org_artifact_id: dict[int, int] = {}
    for tid, oa_id in t_pairs:
        if oa_id:
            unread_by_org_artifact_id[int(oa_id)] = unread_by_thread_id.get(int(tid), 0)

    return total_unread, unread_by_thread_id, unread_by_org_artifact_id

# Индекс КБ: список листов/плиток как в эталонном Excel.
# Реализованы интерактивно только "Управление ИБ" и "СЗИ", остальные помечаем как "Скоро".
INDEX_KB_SHEET_TILES: list[str] = [
    "Управление ИБ",
    "Описание",
    "Итог",
    "Расчет",
    "Программа",
    "Нормативное соответствие",
    "250",
    "СЗИ",
    "Мониторинг",
    "Реагирование",
    "Восстановление",
    "Светофор",
    "Условные сокращения",
    "Приоритет",
    "Экспресс",
]


def _filter_out_default_orgs(orgs: list[Organization]) -> list[Organization]:
    return [o for o in (orgs or []) if (o.name or "").strip() != DEFAULT_ORG_NAME]

def _parse_date_range_bounds_utc(date_from: str | None, date_to: str | None) -> tuple[date | None, date | None, datetime | None, datetime | None, str | None]:
    """
    Парсим диапазон дат из query/form.

    Возвращаем:
    - (df, dt) как date (для отображения)
    - (start, end) как UTC-границы для фильтрации created_at: [start, end)
    - err: текст ошибки (если некорректный ввод)
    """
    df_raw = (date_from or "").strip()
    dt_raw = (date_to or "").strip()
    df = None
    dt = None
    try:
        if df_raw:
            df = date.fromisoformat(df_raw)
    except Exception:
        return None, None, None, None, "Некорректная дата 'с' (ожидается YYYY-MM-DD)"
    try:
        if dt_raw:
            dt = date.fromisoformat(dt_raw)
    except Exception:
        return None, None, None, None, "Некорректная дата 'по' (ожидается YYYY-MM-DD)"

    if df and dt and df > dt:
        return df, dt, None, None, "date_from не может быть больше date_to"

    start = datetime.combine(df, datetime.min.time()) if df else None
    end = (datetime.combine(dt, datetime.min.time()) + timedelta(days=1)) if dt else None
    return df, dt, start, end, None


def _get_active_artifact_levels(db: Session) -> list[ArtifactLevel]:
    return db.query(ArtifactLevel).filter(ArtifactLevel.is_active.is_(True)).order_by(ArtifactLevel.sort_order.asc(), ArtifactLevel.id.asc()).all()


def _get_effective_artifact_ids_for_level(db: Session, *, level_id: int | None) -> list[int]:
    """
    Эффективный набор артефактов для уровня:
    берём все items из уровней с sort_order <= выбранного уровня.
    """
    if not level_id:
        return []
    level = db.get(ArtifactLevel, int(level_id))
    if not level:
        return []
    rows = (
        db.query(ArtifactLevelItem.artifact_id)
        .join(ArtifactLevel, ArtifactLevel.id == ArtifactLevelItem.level_id)
        .filter(ArtifactLevel.sort_order <= level.sort_order)
        .distinct()
        .all()
    )
    return [int(aid) for (aid,) in rows if aid]


def _get_effective_artifact_ids_for_level_code(db: Session, *, level_code: str | None) -> list[int]:
    code = (level_code or "").strip().upper()
    if not code:
        return []
    lvl = db.query(ArtifactLevel).filter(ArtifactLevel.code == code).one_or_none()
    if not lvl:
        return []
    return _get_effective_artifact_ids_for_level(db, level_id=lvl.id)


def _get_default_artifact_level_id(db: Session) -> int | None:
    lvl = db.query(ArtifactLevel).filter(ArtifactLevel.code == "L3").one_or_none()
    if lvl:
        return int(lvl.id)
    lvl2 = db.query(ArtifactLevel).filter(ArtifactLevel.is_active.is_(True)).order_by(ArtifactLevel.sort_order.asc(), ArtifactLevel.id.asc()).first()
    return int(lvl2.id) if lvl2 else None


def _fmt_dt(value: object) -> str:
    """
    Форматируем datetime для UI: делаем стабильный UTC+3 (MSK), т.к. в контейнере/браузере
    на практике бывают проблемы с tzdata/кешем. Для MVP считаем, что все даты в БД в UTC.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        dt = value
        # Treat naive as UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        dt = dt.astimezone(timezone.utc).replace(microsecond=0, tzinfo=None)
        dt = dt + timedelta(hours=3)  # MSK
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


templates.env.filters["dt"] = _fmt_dt


def _fmt_date(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, datetime):
        try:
            return value.date().isoformat()
        except Exception:
            return str(value)
    return str(value)


templates.env.filters["d"] = _fmt_date


def _validate_password(password: str) -> str | None:
    """Минимальная политика пароля для админки."""
    p = password or ""
    if len(p) < 8:
        return "Пароль должен быть не короче 8 символов"
    if any(ch.isspace() for ch in p):
        return "Пароль не должен содержать пробелы"
    has_lower = any("a" <= ch.lower() <= "z" and ch.islower() for ch in p)
    has_upper = any("a" <= ch.lower() <= "z" and ch.isupper() for ch in p)
    has_digit = any(ch.isdigit() for ch in p)
    has_special = any(not ch.isalnum() for ch in p)
    if not (has_lower and has_upper and has_digit and has_special):
        return "Пароль должен содержать: строчные и заглавные буквы, цифры и спецсимвол"
    return None


def _redirect(url: str) -> RedirectResponse:
    return RedirectResponse(url=url, status_code=status.HTTP_303_SEE_OTHER)


@router.get("/login", response_class=HTMLResponse)
def login_page(request: Request) -> HTMLResponse:
    resp = templates.TemplateResponse("login.html", {"request": request, "error": None, "container_class": "container-wide"})
    # Важно: страница логина часто кешируется браузером (особенно при back/forward).
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    return resp


@router.post("/login")
def login_action(
    request: Request,
    login: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
) -> Response:
    user = db.query(User).filter(User.login == login).one_or_none()
    if user and not user.is_active:
        resp = templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Пользователь заблокирован", "container_class": "container-wide"},
            status_code=403,
        )
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        return resp

    if not user or not verify_password(password, user.password_hash):
        resp = templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Неверный логин или пароль", "container_class": "container-wide"},
            status_code=400,
        )
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        return resp

    token = create_access_token(user.id)
    resp = _redirect("/")
    resp.set_cookie(
        key=JWT_COOKIE_NAME,
        value=token,
        httponly=True,
        samesite="lax",
        secure=False,  # в проде включим за reverse-proxy с HTTPS
    )
    return resp


@router.post("/logout")
def logout_action() -> Response:
    resp = _redirect("/login")
    resp.delete_cookie(JWT_COOKIE_NAME)
    return resp


@router.get("/", response_class=HTMLResponse)
def index(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
) -> Response:
    # Admin UI: админку открываем сразу (это основной кабинет администратора).
    if user.is_admin:
        return _redirect("/admin")

    # Доступные организации
    # MVP-допущение: аудитор глобальный, если есть хотя бы один membership auditor
    is_global_auditor = (
        db.query(UserOrgMembership)
        .filter(UserOrgMembership.user_id == user.id, UserOrgMembership.role == Role.auditor)
        .first()
        is not None
    )
    if user.is_admin or is_global_auditor:
        orgs = db.query(Organization).order_by(Organization.name.asc()).all()
    else:
        orgs = (
            db.query(Organization)
            .join(UserOrgMembership, UserOrgMembership.org_id == Organization.id)
            .filter(UserOrgMembership.user_id == user.id)
            .order_by(Organization.name.asc())
            .all()
        )
    if not orgs:
        return templates.TemplateResponse(
            "empty.html",
            {"request": request, "user": user, "message": "Нет доступных организаций. Обратитесь к администратору."},
        )

    # Для аудитора/админа не подставляем организацию автоматически — сначала выбор организации.
    # (Иначе первая страница аудитора сразу открывает таблицу по "первой" организации.)
    if is_global_auditor:
        return _redirect("/auditor/artifacts")

    # Do not auto-select the system organization
    visible_orgs = _filter_out_default_orgs(orgs)
    selected_org_id = org_id or (visible_orgs[0].id if visible_orgs else orgs[0].id)
    role = get_user_role_for_org(db, user, selected_org_id)
    if not role:
        selected_org_id = (visible_orgs[0].id if visible_orgs else orgs[0].id)
        role = get_user_role_for_org(db, user, selected_org_id)

    # Customer UI: по умолчанию ведём пользователя в таблицу артефактов по организации.
    if role == Role.customer:
        return _redirect(f"/my/artifacts?org_id={selected_org_id}")

    # Auditor UI: отдельный экран (сначала выбор организации).
    if role == Role.auditor and not user.is_admin:
        return _redirect("/auditor/artifacts")

    files = db.query(StoredFile).filter(StoredFile.org_id == selected_org_id).order_by(StoredFile.created_at.desc()).all()
    return templates.TemplateResponse(
        "files.html",
        {
            "request": request,
            "user": user,
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "role": role.value if role else None,
            "files": files,
            "max_upload_mb": settings.max_upload_mb,
        },
    )


def _ensure_org_artifacts_materialized_simple(db: Session, org_id: int) -> None:
    artifact_ids = [a_id for (a_id,) in db.query(Artifact.id).all()]
    if not artifact_ids:
        return
    existing_ids = set(
        a_id for (a_id,) in db.query(OrgArtifact.artifact_id).filter(OrgArtifact.org_id == org_id).all()
    )
    now = datetime.utcnow()
    for a_id in artifact_ids:
        if a_id in existing_ids:
            continue
        db.add(OrgArtifact(org_id=org_id, artifact_id=a_id, status=OrgArtifactStatus.missing, created_at=now, updated_at=now))
    db.flush()

def _ensure_org_artifacts_materialized(db: Session, org_id: int) -> None:
    # Быстрое заполнение org_artifacts для организации (если импортировали новые artifacts).
    now = datetime.utcnow()
    stmt = insert(OrgArtifact).from_select(
        ["org_id", "artifact_id", "status", "created_at", "updated_at"],
        select(
            func.cast(org_id, OrgArtifact.org_id.type),
            Artifact.id,
            func.cast(OrgArtifactStatus.missing.value, OrgArtifact.status.type),
            func.cast(now, OrgArtifact.created_at.type),
            func.cast(now, OrgArtifact.updated_at.type),
        ).where(
            ~select(1)
            .where(and_(OrgArtifact.org_id == org_id, OrgArtifact.artifact_id == Artifact.id))
            .exists()
        ),
    )
    db.execute(stmt)


@router.get("/my/artifacts", response_class=HTMLResponse)
def my_artifacts_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    topic: str | None = None,
    domain: str | None = None,
    kb_level: str | None = None,
    short_name: str | None = None,
    q: str | None = None,
    status: str | None = None,
    audit: str | None = None,
    in_period: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    page: int = 1,
    page_size: int = 50,
) -> HTMLResponse:
    # Customer UI: показываем только организации, где у пользователя роль customer.
    orgs = (
        db.query(Organization)
        .join(UserOrgMembership, UserOrgMembership.org_id == Organization.id)
        .filter(UserOrgMembership.user_id == user.id, UserOrgMembership.role == Role.customer)
        .order_by(Organization.name.asc())
        .all()
    )
    orgs = _filter_out_default_orgs(orgs)
    if not orgs:
        raise HTTPException(status_code=403, detail="Страница доступна только роли customer")

    # Support multi-org customers via org_id selector
    selected_org_id = org_id or orgs[0].id
    if selected_org_id not in {o.id for o in orgs}:
        selected_org_id = orgs[0].id
    selected_org = next((o for o in orgs if o.id == selected_org_id), orgs[0])
    in_period_flag = (in_period or "").strip().lower() in ("1", "true", "yes", "on")
    df, dt, p_start, p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)

    page = max(int(page or 1), 1)
    page_size = int(page_size or 50)
    if page_size < 10:
        page_size = 10
    if page_size > 200:
        page_size = 200

    _ensure_org_artifacts_materialized(db, selected_org_id)
    db.commit()

    # базовые фильтры (без статуса) — используем и для табличного вывода, и для расчёта процента заполненности
    filters = [OrgArtifact.org_id == selected_org_id]
    # Ограничение по уровню (customer видит только свой effective-набор уровня организации).
    if getattr(selected_org, "artifact_level_id", None):
        allowed_artifact_ids = _get_effective_artifact_ids_for_level(db, level_id=int(selected_org.artifact_level_id))
        filters.append(OrgArtifact.artifact_id.in_(allowed_artifact_ids))
    # Диапазон дат: опциональный фильтр "только загруженные в диапазоне".
    period_active = bool((p_start or p_end) and in_period_flag and not range_err)
    if topic:
        filters.append(Artifact.topic == topic)
    if domain:
        filters.append(Artifact.domain == domain)
    if kb_level:
        filters.append(Artifact.kb_level == kb_level)
    if short_name:
        filters.append(Artifact.short_name == short_name)
    if q:
        like = f"%{q.strip()}%"
        filters.append(
            (Artifact.indicator_name.ilike(like))
            | (Artifact.title.ilike(like))
            | (Artifact.achievement_text.ilike(like))
            | (Artifact.achievement_item_text.ilike(like))
        )

    base_count_q = db.query(OrgArtifact.id).join(Artifact, Artifact.id == OrgArtifact.artifact_id).filter(*filters)
    PeriodFv = aliased(FileVersion)
    period_sub = None
    if (p_start or p_end) and not range_err:
        conds = []
        if p_start:
            conds.append(PeriodFv.created_at >= p_start)
        if p_end:
            conds.append(PeriodFv.created_at < p_end)
        period_sub = (
            db.query(
                PeriodFv.org_artifact_id.label("oa_id"),
                func.max(PeriodFv.id).label("fv_id"),
            )
            .filter(*conds)
            .group_by(PeriodFv.org_artifact_id)
            .subquery()
        )
    if period_active and period_sub is not None:
        base_count_q = base_count_q.join(period_sub, period_sub.c.oa_id == OrgArtifact.id).filter(period_sub.c.fv_id.isnot(None))

    completion_total = base_count_q.count()
    completion_uploaded = base_count_q.filter(OrgArtifact.status == OrgArtifactStatus.uploaded).count()
    completion_pct = int(round((completion_uploaded * 100.0 / completion_total), 0)) if completion_total else 0

    status_filter = (status or "").strip().lower() or ""
    if status_filter not in ("uploaded", "missing", ""):
        status_filter = ""

    list_count_q = base_count_q
    if status_filter:
        list_count_q = list_count_q.filter(OrgArtifact.status == OrgArtifactStatus(status_filter))
    total = list_count_q.count()

    audit_filter = (audit or "").strip().lower() or ""
    if audit_filter not in ("needs", "audited", "changed", "needs_correction", ""):
        audit_filter = ""
    if audit_filter:
        if audit_filter == "needs":
            list_count_q = list_count_q.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.is_(None),
                OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
            )
        elif audit_filter == "needs_correction":
            list_count_q = list_count_q.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction,
            )
        elif audit_filter == "audited":
            list_count_q = list_count_q.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
            )
        elif audit_filter == "changed":
            list_count_q = list_count_q.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id != OrgArtifact.current_file_version_id,
            )
        total = list_count_q.count()

    CommentBy = aliased(User)
    AuditedBy = aliased(User)
    # latest comment per org_artifact (auditor comment visible to customer)
    sub = (
        db.query(
            OrgArtifactComment.org_artifact_id.label("oa_id"),
            func.max(OrgArtifactComment.created_at).label("max_created_at"),
        )
        .filter(OrgArtifactComment.org_id == selected_org_id)
        .group_by(OrgArtifactComment.org_artifact_id)
        .subquery()
    )
    latest_comment = (
        db.query(OrgArtifactComment)
        .join(sub, and_(OrgArtifactComment.org_artifact_id == sub.c.oa_id, OrgArtifactComment.created_at == sub.c.max_created_at))
        .subquery()
    )

    PeriodFvJoin = aliased(FileVersion)
    if period_sub is not None:
        query = (
            db.query(
                OrgArtifact,
                Artifact,
                FileVersion,
                PeriodFvJoin,
                latest_comment.c.comment_text,
                latest_comment.c.created_at,
                CommentBy,
                AuditedBy,
            )
            .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
            .outerjoin(FileVersion, FileVersion.id == OrgArtifact.current_file_version_id)
            .outerjoin(period_sub, period_sub.c.oa_id == OrgArtifact.id)
            .outerjoin(PeriodFvJoin, PeriodFvJoin.id == period_sub.c.fv_id)
            .outerjoin(latest_comment, latest_comment.c.org_artifact_id == OrgArtifact.id)
            .outerjoin(CommentBy, CommentBy.id == latest_comment.c.author_user_id)
            .outerjoin(AuditedBy, AuditedBy.id == OrgArtifact.audited_by_user_id)
            .filter(*filters)
            .order_by(Artifact.topic.asc(), Artifact.domain.asc(), Artifact.short_name.asc(), Artifact.achievement_item_no.asc().nullsfirst())
        )
        if period_active:
            query = query.filter(period_sub.c.fv_id.isnot(None))
    else:
        query = (
            db.query(
                OrgArtifact,
                Artifact,
                FileVersion,
                sa.literal(None).label("fv_period"),
                latest_comment.c.comment_text,
                latest_comment.c.created_at,
                CommentBy,
                AuditedBy,
            )
            .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
            .outerjoin(FileVersion, FileVersion.id == OrgArtifact.current_file_version_id)
            .outerjoin(latest_comment, latest_comment.c.org_artifact_id == OrgArtifact.id)
            .outerjoin(CommentBy, CommentBy.id == latest_comment.c.author_user_id)
            .outerjoin(AuditedBy, AuditedBy.id == OrgArtifact.audited_by_user_id)
            .filter(*filters)
            .order_by(Artifact.topic.asc(), Artifact.domain.asc(), Artifact.short_name.asc(), Artifact.achievement_item_no.asc().nullsfirst())
        )
    if status_filter:
        query = query.filter(OrgArtifact.status == OrgArtifactStatus(status_filter))
    if audit_filter:
        if audit_filter == "needs":
            query = query.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.is_(None),
                OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
            )
        elif audit_filter == "needs_correction":
            query = query.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction,
            )
        elif audit_filter == "audited":
            query = query.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
            )
        elif audit_filter == "changed":
            query = query.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id != OrgArtifact.current_file_version_id,
            )

    offset = (page - 1) * page_size
    rows = []
    for (oa, a, fv, fv_period, c_text, c_at, c_by, audited_by) in query.offset(offset).limit(page_size).all():
        # UI-friendly audit badge (prevents template drift)
        review = (getattr(oa, "review_status", None).value if getattr(oa, "review_status", None) else "")
        if not oa.current_file_version_id:
            audit_label = "—"
            audit_class = "badge badge-neutral"
        elif review == OrgArtifactReviewStatus.needs_correction.value:
            audit_label = "Требует корректировки"
            audit_class = "badge badge-danger"
        elif (
            review == OrgArtifactReviewStatus.approved.value
            and oa.audited_file_version_id
            and oa.audited_file_version_id == oa.current_file_version_id
        ):
            audit_label = "Проаудировано"
            audit_class = "badge badge-info"
        elif oa.audited_file_version_id:
            audit_label = "Изменён"
            audit_class = "badge badge-warn"
        else:
            audit_label = "Требует аудита"
            audit_class = "badge badge-warn"

        rows.append(
            {
                "oa": oa,
                "a": a,
                "fv": fv,
                "fv_period": fv_period,
                "comment_text": c_text or "",
                "comment_at": c_at,
                "comment_by": c_by.login if c_by else "",
                "audited_by": audited_by.login if audited_by else "",
                "audit_label": audit_label,
                "audit_class": audit_class,
            }
        )

    topics = [t for (t,) in db.query(Artifact.topic).filter(Artifact.topic != "").distinct().order_by(Artifact.topic.asc()).all()]
    domains = [d for (d,) in db.query(Artifact.domain).filter(Artifact.domain != "").distinct().order_by(Artifact.domain.asc()).all()]
    kb_levels = [k for (k,) in db.query(Artifact.kb_level).filter(Artifact.kb_level != "").distinct().order_by(Artifact.kb_level.asc()).all()]

    total_pages = max((total + page_size - 1) // page_size, 1)
    if page > total_pages:
        page = total_pages
        offset = (page - 1) * page_size

    # Базовый querystring для пагинации (фильтры + page_size), без org_id.
    base_query = urlencode(
        {
            "org_id": str(selected_org_id),
            "topic": topic or "",
            "domain": domain or "",
            "kb_level": kb_level or "",
            "short_name": short_name or "",
            "q": q or "",
            "status": status_filter,
            "audit": audit_filter,
            "in_period": ("1" if in_period_flag else ""),
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "page_size": str(page_size),
        }
    )

    # Список страниц вокруг текущей (для кликабельных номеров).
    window = 3
    start = max(1, page - window)
    end = min(total_pages, page + window)
    page_links = list(range(start, end + 1))

    page_oa_ids = [int(r["oa"].id) for r in rows if r.get("oa") is not None]
    chat_unread_total, _unread_by_thread_id, chat_unread_by_oa_id = _get_chat_unread_for_org(
        db, user=user, org_id=int(selected_org_id), only_org_artifact_ids=page_oa_ids
    )

    resp = templates.TemplateResponse(
        "customer_artifacts.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "org_name": selected_org.name,
            "selected_org_id": selected_org_id,
            "rows": rows,
            "max_upload_mb": settings.max_upload_mb,
            "topic": topic,
            "domain": domain,
            "kb_level": kb_level,
            "short_name": short_name,
            "q": q,
            "status": status_filter,
            "audit": audit_filter,
            "in_period": in_period_flag,
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "date_range_error": range_err or "",
            "org_level": selected_org.artifact_level,
            "topics": topics,
            "domains": domains,
            "kb_levels": kb_levels,
            "completion_total": completion_total,
            "completion_uploaded": completion_uploaded,
            "completion_pct": completion_pct,
            "page": page,
            "page_size": page_size,
            "total": total,
            "chat_unread_total": chat_unread_total,
            "chat_unread_by_oa_id": chat_unread_by_oa_id,
            "total_pages": total_pages,
            "has_prev": page > 1,
            "has_next": offset + page_size < total,
            "page_links": page_links,
            "base_query": base_query,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


def _get_customer_orgs(db: Session, user: User) -> list[Organization]:
    orgs = (
        db.query(Organization)
        .join(UserOrgMembership, UserOrgMembership.org_id == Organization.id)
        .filter(UserOrgMembership.user_id == user.id, UserOrgMembership.role == Role.customer)
        .order_by(Organization.name.asc())
        .all()
    )
    return _filter_out_default_orgs(orgs)


def _get_customer_selected_org(db: Session, user: User, org_id: int | None) -> tuple[list[Organization], Organization]:
    orgs = _get_customer_orgs(db, user)
    if not orgs:
        raise HTTPException(status_code=403, detail="Недостаточно прав")
    selected_id = org_id or orgs[0].id
    selected = next((o for o in orgs if o.id == selected_id), None) or orgs[0]
    return orgs, selected


def _split_short_name(sn: str) -> list[str]:
    return [p.strip() for p in (sn or "").split(".") if p.strip()]


@router.get("/my/files", response_class=HTMLResponse)
def my_files_explorer(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    path: str | None = None,
) -> HTMLResponse:
    orgs, org = _get_customer_selected_org(db, user, org_id)
    _ensure_org_artifacts_materialized(db, org.id)
    db.commit()

    # Берём все артефакты этой организации (MVP: до нескольких тысяч) и строим дерево путей по short_name.
    CreatedBy = aliased(User)
    UpdatedBy = aliased(User)
    CommentBy = aliased(User)
    AuditedBy = aliased(User)

    sub = (
        db.query(
            OrgArtifactComment.org_artifact_id.label("oa_id"),
            func.max(OrgArtifactComment.created_at).label("max_created_at"),
        )
        .filter(OrgArtifactComment.org_id == org.id)
        .group_by(OrgArtifactComment.org_artifact_id)
        .subquery()
    )
    latest_comment = (
        db.query(OrgArtifactComment)
        .join(sub, and_(OrgArtifactComment.org_artifact_id == sub.c.oa_id, OrgArtifactComment.created_at == sub.c.max_created_at))
        .subquery()
    )
    allowed_artifact_ids: list[int] | None = None
    if getattr(org, "artifact_level_id", None):
        allowed_artifact_ids = _get_effective_artifact_ids_for_level(db, level_id=int(org.artifact_level_id))

    q_rows = (
        db.query(
            OrgArtifact,
            Artifact,
            FileVersion,
            CreatedBy,
            UpdatedBy,
            latest_comment.c.comment_text,
            latest_comment.c.created_at,
            CommentBy,
            AuditedBy,
        )
        .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
        .outerjoin(FileVersion, FileVersion.id == OrgArtifact.current_file_version_id)
        .outerjoin(CreatedBy, CreatedBy.id == FileVersion.created_by_user_id)
        .outerjoin(UpdatedBy, UpdatedBy.id == OrgArtifact.updated_by_user_id)
        .outerjoin(latest_comment, latest_comment.c.org_artifact_id == OrgArtifact.id)
        .outerjoin(CommentBy, CommentBy.id == latest_comment.c.author_user_id)
        .outerjoin(AuditedBy, AuditedBy.id == OrgArtifact.audited_by_user_id)
        .filter(OrgArtifact.org_id == org.id)
    )
    if allowed_artifact_ids is not None:
        q_rows = q_rows.filter(OrgArtifact.artifact_id.in_(allowed_artifact_ids))
    rows = q_rows.all()

    # Нормализуем path: "ВССТ/КМНК/1" -> ["ВССТ","КМНК","1"]
    path = (path or "").strip().strip("/")
    cur_segments = [p for p in path.split("/") if p] if path else []

    # Вычисляем "детей" текущей директории.
    subfolders: dict[str, int] = {}
    leaf_items: list[dict] = []
    for (oa, a, fv, created_by, updated_by, c_text, c_at, c_by, audited_by) in rows:
        segs = _split_short_name(a.short_name)
        if not segs:
            continue
        if segs[: len(cur_segments)] != cur_segments:
            continue
        if len(segs) > len(cur_segments):
            nxt = segs[len(cur_segments)]
            subfolders[nxt] = subfolders.get(nxt, 0) + 1
        else:
            review = (getattr(oa, "review_status", None).value if getattr(oa, "review_status", None) else "")
            if not oa.current_file_version_id:
                audit_label = "—"
                audit_class = "badge badge-neutral"
            elif review == OrgArtifactReviewStatus.needs_correction.value:
                audit_label = "Требует корректировки"
                audit_class = "badge badge-danger"
            elif (
                review == OrgArtifactReviewStatus.approved.value
                and oa.audited_file_version_id
                and oa.audited_file_version_id == oa.current_file_version_id
            ):
                audit_label = "Проаудировано"
                audit_class = "badge badge-info"
            elif oa.audited_file_version_id:
                audit_label = "Изменён"
                audit_class = "badge badge-warn"
            else:
                audit_label = "Требует аудита"
                audit_class = "badge badge-warn"

            leaf_items.append(
                {
                    "oa": oa,
                    "a": a,
                    "fv": fv,
                    "uploaded_by": created_by.login if created_by else "",
                    "updated_by": updated_by.login if updated_by else "",
                    "comment_text": c_text or "",
                    "comment_at": c_at,
                    "comment_by": c_by.login if c_by else "",
                    "audited_by": audited_by.login if audited_by else "",
                    "audit_label": audit_label,
                    "audit_class": audit_class,
                }
            )

    # Сортировка папок/листов
    folders_sorted = sorted(subfolders.items(), key=lambda x: x[0])
    leaf_items.sort(key=lambda r: (r["a"].short_name, r["a"].achievement_item_no or 0))

    # Breadcrumbs
    crumbs = []
    acc = []
    for s in cur_segments:
        acc.append(s)
        crumbs.append({"name": s, "path": "/".join(acc)})

    chat_unread_total, _unread_by_thread_id, _unread_by_oa_id = _get_chat_unread_for_org(
        db, user=user, org_id=int(org.id), only_org_artifact_ids=None
    )

    resp = templates.TemplateResponse(
        "customer_files.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "org_name": org.name,
            "selected_org_id": org.id,
            "path": path,
            "crumbs": crumbs,
            "folders": folders_sorted,
            "leaf_items": leaf_items,
            "max_upload_mb": settings.max_upload_mb,
            "chat_unread_total": chat_unread_total,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/my/index-kb", response_class=HTMLResponse)
def my_index_kb_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    sheet: str | None = None,
    q: str | None = None,
) -> HTMLResponse:
    orgs, org = _get_customer_selected_org(db, user, org_id)
    # If customer has multiple orgs and didn't choose — force explicit choice (no auto)
    if len(orgs) > 1 and not org_id:
        resp = templates.TemplateResponse(
            "select_org.html",
            {
                "request": request,
                "user": user,
                "container_class": "container-wide",
                "title": "Индекс КБ",
                "subtitle": "Выберите организацию, чтобы открыть Индекс КБ (только для ваших организаций).",
                "action_path": "/my/index-kb",
                "orgs": orgs,
            },
        )
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp

    selected_org_id = org.id
    df, dt, _p_start, _p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)
    available_sheets: list[str] = []
    if get_uib_template_from_db(db):
        available_sheets.append(UIB_SHEET_NAME)
    if get_szi_template_from_db(db):
        available_sheets.append(SZI_SHEET_NAME)
    err = None if available_sheets else "Шаблоны Индекса КБ не загружены в БД (нужны seed‑миграции)."
    resp = templates.TemplateResponse(
        "auditor_index_kb.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "date_range_error": range_err or "",
            "sheet_names": INDEX_KB_SHEET_TILES,
            "available_sheets": available_sheets,
            "error": err,
            "base_prefix": "/my",
            "files_base": "/my/files",
            "artifacts_base": "/my/artifacts",
            "show_org_selector": len(orgs) > 1,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/my/index-kb/uib", response_class=HTMLResponse)
def my_index_kb_uib_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> HTMLResponse:
    orgs, org = _get_customer_selected_org(db, user, org_id)
    if len(orgs) > 1 and not org_id:
        resp = templates.TemplateResponse(
            "select_org.html",
            {
                "request": request,
                "user": user,
                "container_class": "container-wide",
                "title": "Управление ИБ (Индекс КБ)",
                "subtitle": "Выберите организацию, чтобы открыть форму (только для ваших организаций).",
                "action_path": "/my/index-kb/uib",
                "orgs": orgs,
            },
        )
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp

    selected_org_id = org.id
    df, dt, p_start, p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)
    if not get_uib_template_from_db(db):
        resp = templates.TemplateResponse(
            "auditor_index_kb_uib.html",
            {
                "request": request,
                "user": user,
                "container_class": "container-wide",
                "orgs": orgs,
                "selected_org_id": selected_org_id,
                "error": "Шаблон УИБ не загружен в БД.",
                "rows": [],
                "summary_rows": [],
                "sheet_name": UIB_SHEET_NAME,
                "org": org,
                "base_prefix": "/my",
                "files_base": "/my/files",
                "artifacts_base": "/my/artifacts",
                "show_org_selector": len(orgs) > 1,
                "readonly": True,
                "date_from": (df.isoformat() if df else ""),
                "date_to": (dt.isoformat() if dt else ""),
                "date_range_error": range_err or "",
            },
            status_code=200,
        )
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp

    org_obj, tpl, rows = build_uib_view(db, org_id=selected_org_id, actor=user, range_start=p_start, range_end=p_end)
    from app.index_kb.uib_sheet import compute_uib_summary

    summary_rows = compute_uib_summary(rows)
    resp = templates.TemplateResponse(
        "auditor_index_kb_uib.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "org": org_obj,
            "sheet_name": UIB_SHEET_NAME,
            "rows": rows,
            "summary_rows": summary_rows,
            "error": None,
            "base_prefix": "/my",
            "files_base": "/my/files",
            "artifacts_base": "/my/artifacts",
            "show_org_selector": len(orgs) > 1,
            "readonly": True,
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "date_range_error": range_err or "",
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


def _get_accessible_orgs_for_auditor(db: Session, user: User) -> list[Organization]:
    if user.is_admin:
        return _filter_out_default_orgs(db.query(Organization).order_by(Organization.name.asc()).all())
    # По текущему MVP правилу auditor считается глобальным, если есть хотя бы один membership auditor.
    is_global_auditor = (
        db.query(UserOrgMembership)
        .filter(UserOrgMembership.user_id == user.id, UserOrgMembership.role == Role.auditor)
        .first()
        is not None
    )
    if is_global_auditor:
        return _filter_out_default_orgs(db.query(Organization).order_by(Organization.name.asc()).all())
    # fallback: только свои (на случай, если правила изменятся)
    return _filter_out_default_orgs(
        db.query(Organization)
        .join(UserOrgMembership, UserOrgMembership.org_id == Organization.id)
        .filter(UserOrgMembership.user_id == user.id)
        .order_by(Organization.name.asc())
        .all()
    )


@router.get("/auditor/artifacts", response_class=HTMLResponse)
def auditor_artifacts_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    topic: str | None = None,
    domain: str | None = None,
    kb_level: str | None = None,
    short_name: str | None = None,
    q: str | None = None,
    status: str | None = None,
    level: str | None = None,
    in_period: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    page: int = 1,
    page_size: int = 50,
) -> HTMLResponse:
    # только auditor/admin
    orgs = _get_accessible_orgs_for_auditor(db, user)
    levels = _get_active_artifact_levels(db)
    if not orgs:
        return templates.TemplateResponse(
            "empty.html",
            {"request": request, "user": user, "message": "Нет доступных организаций. Обратитесь к администратору."},
        )
    allowed_ids = {o.id for o in orgs}
    if not org_id or org_id not in allowed_ids:
        # Не показываем таблицу, пока организация не выбрана — но оставляем фильтры и селект.
        # Это улучшает UX и избегает "автовыбора" организации.
        resp = templates.TemplateResponse(
            "auditor_artifacts.html",
            {
                "request": request,
                "user": user,
                "container_class": "container-wide",
                "orgs": orgs,
                "selected_org_id": None,
                "org_required": True,
                # filters (keep user input visible)
                "topic": topic,
                "domain": domain,
                "kb_level": kb_level,
                "short_name": short_name,
                "q": q,
                "status": status,
                "level": level,
                "levels": levels,
                "date_from": (date_from or ""),
                "date_to": (date_to or ""),
                "date_range_error": "",
                "topics": [],
                "domains": [],
                "kb_levels": [],
                "page": 1,
                "page_size": int(page_size or 50),
                "rows": [],
                "total": 0,
                "total_pages": 1,
                "has_prev": False,
                "has_next": False,
                "page_links": [1],
                "base_query": "",
                "export_query": "",
                "completion_total": 0,
                "completion_uploaded": 0,
                "completion_pct": 0,
                "current_url": request.url.path,
                "can_delete_files": False,
                "org_level": None,
            },
        )
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp
    selected_org_id = org_id
    org = db.get(Organization, selected_org_id)
    role = get_user_role_for_org(db, user, selected_org_id)
    if role not in (Role.auditor, Role.admin):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")
    can_delete_files = role == Role.admin
    current_url = request.url.path + (f"?{request.url.query}" if request.url.query else "")

    page = max(int(page or 1), 1)
    page_size = int(page_size or 50)
    if page_size < 10:
        page_size = 10
    if page_size > 200:
        page_size = 200

    _ensure_org_artifacts_materialized(db, selected_org_id)
    db.commit()

    CreatedBy = aliased(User)
    UpdatedBy = aliased(User)
    CommentBy = aliased(User)

    # latest comment per org_artifact
    sub = (
        db.query(
            OrgArtifactComment.org_artifact_id.label("oa_id"),
            func.max(OrgArtifactComment.created_at).label("max_created_at"),
        )
        .filter(OrgArtifactComment.org_id == selected_org_id)
        .group_by(OrgArtifactComment.org_artifact_id)
        .subquery()
    )
    latest_comment = (
        db.query(OrgArtifactComment)
        .join(sub, and_(OrgArtifactComment.org_artifact_id == sub.c.oa_id, OrgArtifactComment.created_at == sub.c.max_created_at))
        .subquery()
    )

    filters = [OrgArtifact.org_id == selected_org_id]
    in_period_flag = (in_period or "").strip().lower() in ("1", "true", "yes", "on")
    df, dt, p_start, p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)
    # Диапазон дат: опционально фильтруем "только загруженные в диапазоне".
    period_active = bool((p_start or p_end) and in_period_flag and not range_err)
    # Фильтр уровня (для аудитора): показываем effective-набор выбранного уровня.
    level_filter = (level or "").strip().upper()
    if level_filter:
        allowed_artifact_ids = _get_effective_artifact_ids_for_level_code(db, level_code=level_filter)
        if allowed_artifact_ids:
            filters.append(OrgArtifact.artifact_id.in_(allowed_artifact_ids))
        else:
            # неизвестный уровень -> показываем пусто
            filters.append(OrgArtifact.id == -1)
    if topic:
        filters.append(Artifact.topic == topic)
    if domain:
        filters.append(Artifact.domain == domain)
    if kb_level:
        filters.append(Artifact.kb_level == kb_level)
    if short_name:
        filters.append(Artifact.short_name == short_name)
    if q:
        like = f"%{q.strip()}%"
        filters.append(
            (Artifact.indicator_name.ilike(like))
            | (Artifact.title.ilike(like))
            | (Artifact.achievement_text.ilike(like))
            | (Artifact.achievement_item_text.ilike(like))
        )

    base_count_q = db.query(OrgArtifact.id).join(Artifact, Artifact.id == OrgArtifact.artifact_id).filter(*filters)
    if period_active:
        conds = []
        if p_start:
            conds.append(FileVersion.created_at >= p_start)
        if p_end:
            conds.append(FileVersion.created_at < p_end)
        base_count_q = base_count_q.filter(OrgArtifact.current_file_version_id.isnot(None)).join(
            FileVersion, FileVersion.id == OrgArtifact.current_file_version_id
        ).filter(*conds)
    completion_total = base_count_q.count()
    completion_uploaded = base_count_q.filter(OrgArtifact.status == OrgArtifactStatus.uploaded).count()
    completion_pct = int(round((completion_uploaded * 100.0 / completion_total), 0)) if completion_total else 0

    status_filter = (status or "").strip().lower() or ""
    if status_filter not in ("uploaded", "missing", "changed", "audited", "needs_correction", ""):
        status_filter = ""

    list_count_q = base_count_q
    if status_filter:
        if status_filter in ("missing", "uploaded"):
            if status_filter == "missing":
                list_count_q = list_count_q.filter(OrgArtifact.status == OrgArtifactStatus.missing)
            else:
                # "uploaded" in UI == "требует аудита" (файл есть, но ещё не проаудирован)
                list_count_q = list_count_q.filter(
                    OrgArtifact.status == OrgArtifactStatus.uploaded,
                    OrgArtifact.current_file_version_id.isnot(None),
                    OrgArtifact.audited_file_version_id.is_(None),
                    OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                )
        elif status_filter == "needs_correction":
            list_count_q = list_count_q.filter(
                OrgArtifact.status == OrgArtifactStatus.uploaded,
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction,
            )
        elif status_filter == "audited":
            list_count_q = list_count_q.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
            )
        elif status_filter == "changed":
            list_count_q = list_count_q.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id != OrgArtifact.current_file_version_id,
            )
    total = list_count_q.count()

    query = (
        db.query(OrgArtifact, Artifact, FileVersion, CreatedBy, UpdatedBy, latest_comment.c.comment_text, latest_comment.c.created_at, CommentBy)
        .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
        .outerjoin(FileVersion, FileVersion.id == OrgArtifact.current_file_version_id)
        .outerjoin(CreatedBy, CreatedBy.id == FileVersion.created_by_user_id)
        .outerjoin(UpdatedBy, UpdatedBy.id == OrgArtifact.updated_by_user_id)
        .outerjoin(latest_comment, latest_comment.c.org_artifact_id == OrgArtifact.id)
        .outerjoin(CommentBy, CommentBy.id == latest_comment.c.author_user_id)
        .filter(*filters)
        .order_by(Artifact.topic.asc(), Artifact.domain.asc(), Artifact.short_name.asc(), Artifact.achievement_item_no.asc().nullsfirst())
    )
    if period_active:
        conds = []
        if p_start:
            conds.append(FileVersion.created_at >= p_start)
        if p_end:
            conds.append(FileVersion.created_at < p_end)
        query = query.filter(
            OrgArtifact.current_file_version_id.isnot(None),
            *conds,
        )
    if status_filter:
        if status_filter in ("missing", "uploaded"):
            if status_filter == "missing":
                query = query.filter(OrgArtifact.status == OrgArtifactStatus.missing)
            else:
                query = query.filter(
                    OrgArtifact.status == OrgArtifactStatus.uploaded,
                    OrgArtifact.current_file_version_id.isnot(None),
                    OrgArtifact.audited_file_version_id.is_(None),
                    OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                )
        elif status_filter == "needs_correction":
            query = query.filter(
                OrgArtifact.status == OrgArtifactStatus.uploaded,
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction,
            )
        elif status_filter == "audited":
            query = query.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
            )
        elif status_filter == "changed":
            query = query.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id != OrgArtifact.current_file_version_id,
            )

    offset = (page - 1) * page_size
    rows = []
    for (oa, a, fv, created_by, updated_by, c_text, c_at, c_by) in query.offset(offset).limit(page_size).all():
        rows.append(
            {
                "oa": oa,
                "a": a,
                "fv": fv,
                "uploaded_by": created_by.login if created_by else "",
                "updated_by": updated_by.login if updated_by else "",
                "comment_text": c_text or "",
                "comment_at": c_at,
                "comment_by": c_by.login if c_by else "",
            }
        )

    topics = [t for (t,) in db.query(Artifact.topic).filter(Artifact.topic != "").distinct().order_by(Artifact.topic.asc()).all()]
    domains = [d for (d,) in db.query(Artifact.domain).filter(Artifact.domain != "").distinct().order_by(Artifact.domain.asc()).all()]
    kb_levels = [k for (k,) in db.query(Artifact.kb_level).filter(Artifact.kb_level != "").distinct().order_by(Artifact.kb_level.asc()).all()]

    total_pages = max((total + page_size - 1) // page_size, 1)
    base_query = urlencode(
        {
            "org_id": str(selected_org_id),
            "topic": topic or "",
            "domain": domain or "",
            "kb_level": kb_level or "",
            "short_name": short_name or "",
            "q": q or "",
            "status": status_filter,
            "level": level_filter,
            "in_period": ("1" if in_period_flag else ""),
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "page_size": str(page_size),
        }
    )
    export_query = urlencode(
        {
            "org_id": str(selected_org_id),
            "topic": topic or "",
            "domain": domain or "",
            "kb_level": kb_level or "",
            "short_name": short_name or "",
            "q": q or "",
            "status": status_filter,
            "level": level_filter,
            "in_period": ("1" if in_period_flag else ""),
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
        }
    )
    window = 3
    start = max(1, page - window)
    end = min(total_pages, page + window)
    page_links = list(range(start, end + 1))

    page_oa_ids = [int(r["oa"].id) for r in rows if r.get("oa") is not None]
    chat_unread_total, _unread_by_thread_id, chat_unread_by_oa_id = _get_chat_unread_for_org(
        db, user=user, org_id=int(selected_org_id), only_org_artifact_ids=page_oa_ids
    )

    resp = templates.TemplateResponse(
        "auditor_artifacts.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "can_delete_files": can_delete_files,
            "current_url": current_url,
            "rows": rows,
            "topic": topic,
            "domain": domain,
            "kb_level": kb_level,
            "short_name": short_name,
            "q": q,
            "status": status_filter,
            "level": level_filter,
            "in_period": in_period_flag,
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "date_range_error": range_err or "",
            "levels": levels,
            "topics": topics,
            "domains": domains,
            "kb_levels": kb_levels,
            "completion_total": completion_total,
            "completion_uploaded": completion_uploaded,
            "completion_pct": completion_pct,
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "has_prev": page > 1,
            "has_next": offset + page_size < total,
            "base_query": base_query,
            "page_links": page_links,
            "export_query": export_query,
            "org_level": org.artifact_level if org else None,
            "chat_unread_total": chat_unread_total,
            "chat_unread_by_oa_id": chat_unread_by_oa_id,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.post("/auditor/org_artifacts/{org_artifact_id}/comment")
def auditor_add_comment(
    org_artifact_id: int,
    request: Request,
    org_id: int = Form(...),
    comment: str = Form(""),
    back: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    role = get_user_role_for_org(db, user, org_id)
    if role not in (Role.auditor, Role.admin):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa or oa.org_id != org_id:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")

    txt = (comment or "").strip()
    if txt:
        db.add(OrgArtifactComment(org_id=org_id, org_artifact_id=oa.id, author_user_id=user.id, comment_text=txt))
        write_audit_log(
            db,
            actor=user,
            org_id=org_id,
            action="comment",
            entity_type="org_artifact",
            entity_id=str(oa.id),
            after={"comment": txt},
            request=request,
        )
        db.commit()

    # Prefer explicit back (sent by form), fallback to referer.
    ref = (back or "").strip() or (request.headers.get("referer") or f"/auditor/artifacts?org_id={org_id}")
    # безопасный редирект только на относительный путь
    if not ref or "://" in ref or not ref.startswith("/"):
        ref = f"/auditor/artifacts?org_id={org_id}"
    return _redirect(ref)


@router.post("/auditor/org_artifacts/{org_artifact_id}/audit")
def auditor_mark_audited(
    org_artifact_id: int,
    request: Request,
    back: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role not in (Role.auditor, Role.admin):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")
    if not oa.current_file_version_id:
        raise HTTPException(status_code=400, detail="Нет файла для аудита")

    now = datetime.utcnow()
    before = {
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at.isoformat() if oa.audited_at else None,
        "audited_by_user_id": oa.audited_by_user_id,
        "current_file_version_id": oa.current_file_version_id,
    }

    oa.audited_file_version_id = oa.current_file_version_id
    oa.audited_at = now
    oa.audited_by_user_id = user.id
    oa.review_status = OrgArtifactReviewStatus.approved
    oa.updated_at = now
    oa.updated_by_user_id = user.id

    after = {
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at.isoformat() if oa.audited_at else None,
        "audited_by_user_id": oa.audited_by_user_id,
        "current_file_version_id": oa.current_file_version_id,
        "review_status": oa.review_status.value,
    }

    write_audit_log(
        db,
        actor=user,
        org_id=oa.org_id,
        action="audit",
        entity_type="org_artifact",
        entity_id=str(oa.id),
        before=before,
        after=after,
        request=request,
    )
    db.commit()

    ref = (back or "").strip() or (request.headers.get("referer") or "")
    if not ref or "://" in ref or not ref.startswith("/"):
        ref = f"/auditor/artifacts?org_id={oa.org_id}"
    return _redirect(ref)


@router.post("/auditor/org_artifacts/{org_artifact_id}/needs_correction")
def auditor_mark_needs_correction(
    org_artifact_id: int,
    request: Request,
    comment: str = Form(""),
    back: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role not in (Role.auditor, Role.admin):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")
    if not oa.current_file_version_id:
        raise HTTPException(status_code=400, detail="Нет файла для возврата на корректировку")

    txt = (comment or "").strip()
    if not txt:
        raise HTTPException(status_code=400, detail="Комментарий обязателен")

    now = datetime.utcnow()
    before = {
        "review_status": getattr(oa, "review_status", None).value if getattr(oa, "review_status", None) else None,
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at.isoformat() if oa.audited_at else None,
        "audited_by_user_id": oa.audited_by_user_id,
        "current_file_version_id": oa.current_file_version_id,
    }

    # 1) сохранить комментарий аудитора
    db.add(OrgArtifactComment(org_id=oa.org_id, org_artifact_id=oa.id, author_user_id=user.id, comment_text=txt))

    # 2) поставить статус "Требует корректировки" + сбросить audit поля (чтобы не считалось проаудированным)
    oa.review_status = OrgArtifactReviewStatus.needs_correction
    oa.audited_file_version_id = None
    oa.audited_at = None
    oa.audited_by_user_id = None
    oa.updated_at = now
    oa.updated_by_user_id = user.id

    after = {
        "review_status": oa.review_status.value,
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at,
        "audited_by_user_id": oa.audited_by_user_id,
        "current_file_version_id": oa.current_file_version_id,
        "comment": txt,
    }

    write_audit_log(
        db,
        actor=user,
        org_id=oa.org_id,
        action="audit_needs_correction",
        entity_type="org_artifact",
        entity_id=str(oa.id),
        before=before,
        after=after,
        request=request,
    )
    db.commit()

    ref = (back or "").strip() or (request.headers.get("referer") or "")
    if not ref or "://" in ref or not ref.startswith("/"):
        ref = f"/auditor/artifacts?org_id={oa.org_id}"
    return _redirect(ref)


@router.get("/auditor/artifacts/export.xlsx")
def auditor_artifacts_export_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int = 0,
    topic: str | None = None,
    domain: str | None = None,
    indicator: str | None = None,
    kb_level: str | None = None,
    short_name: str | None = None,
    q: str | None = None,
    status: str | None = None,
    audit_status: str | None = None,
    level: str | None = None,
    in_period: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> Response:
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id обязателен")
    role = get_user_role_for_org(db, user, org_id)
    if role not in (Role.auditor, Role.admin):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")

    _ensure_org_artifacts_materialized(db, org_id)
    db.commit()

    org = db.get(Organization, org_id)
    in_period_flag = (in_period or "").strip().lower() in ("1", "true", "yes", "on")
    df, dt, p_start, p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)
    period_active = bool((p_start or p_end) and in_period_flag and not range_err)

    CreatedBy = aliased(User)
    UpdatedBy = aliased(User)
    CommentBy = aliased(User)
    AuditedBy = aliased(User)
    AuditedFv = aliased(FileVersion)

    sub = (
        db.query(
            OrgArtifactComment.org_artifact_id.label("oa_id"),
            func.max(OrgArtifactComment.created_at).label("max_created_at"),
        )
        .filter(OrgArtifactComment.org_id == org_id)
        .group_by(OrgArtifactComment.org_artifact_id)
        .subquery()
    )
    latest_comment = (
        db.query(OrgArtifactComment)
        .join(sub, and_(OrgArtifactComment.org_artifact_id == sub.c.oa_id, OrgArtifactComment.created_at == sub.c.max_created_at))
        .subquery()
    )

    query = (
        db.query(
            OrgArtifact,
            Artifact,
            FileVersion,
            CreatedBy,
            UpdatedBy,
            AuditedFv,
            AuditedBy,
            latest_comment.c.comment_text,
            latest_comment.c.created_at,
            CommentBy,
        )
        .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
        .outerjoin(FileVersion, FileVersion.id == OrgArtifact.current_file_version_id)
        .outerjoin(CreatedBy, CreatedBy.id == FileVersion.created_by_user_id)
        .outerjoin(UpdatedBy, UpdatedBy.id == OrgArtifact.updated_by_user_id)
        .outerjoin(AuditedFv, AuditedFv.id == OrgArtifact.audited_file_version_id)
        .outerjoin(AuditedBy, AuditedBy.id == OrgArtifact.audited_by_user_id)
        .outerjoin(latest_comment, latest_comment.c.org_artifact_id == OrgArtifact.id)
        .outerjoin(CommentBy, CommentBy.id == latest_comment.c.author_user_id)
        .filter(OrgArtifact.org_id == org_id)
        .order_by(Artifact.topic.asc(), Artifact.domain.asc(), Artifact.short_name.asc(), Artifact.achievement_item_no.asc().nullsfirst())
    )
    if period_active:
        conds = []
        if p_start:
            conds.append(FileVersion.created_at >= p_start)
        if p_end:
            conds.append(FileVersion.created_at < p_end)
        query = query.filter(
            OrgArtifact.current_file_version_id.isnot(None),
            *conds,
        )
    level_filter = (level or "").strip().upper()
    if level_filter:
        allowed_artifact_ids = _get_effective_artifact_ids_for_level_code(db, level_code=level_filter)
        if allowed_artifact_ids:
            query = query.filter(OrgArtifact.artifact_id.in_(allowed_artifact_ids))
        else:
            query = query.filter(OrgArtifact.id == -1)
    if topic:
        query = query.filter(Artifact.topic == topic)
    if domain:
        query = query.filter(Artifact.domain == domain)
    if kb_level:
        query = query.filter(Artifact.kb_level == kb_level)
    if short_name:
        query = query.filter(Artifact.short_name == short_name)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            (Artifact.indicator_name.ilike(like))
            | (Artifact.title.ilike(like))
            | (Artifact.achievement_text.ilike(like))
            | (Artifact.achievement_item_text.ilike(like))
        )

    if indicator:
        like_i = f"%{indicator.strip()}%"
        query = query.filter(Artifact.indicator_name.ilike(like_i))

    # Фильтр по статусу аудирования (как на /auditor/artifacts):
    # missing / uploaded(==needs audit) / needs_correction / changed / audited
    audit_filter = (audit_status or status or "").strip().lower() or ""
    if audit_filter not in ("uploaded", "missing", "changed", "audited", "needs_correction", ""):
        audit_filter = ""
    if audit_filter:
        if audit_filter in ("missing", "uploaded"):
            if audit_filter == "missing":
                query = query.filter(OrgArtifact.status == OrgArtifactStatus.missing)
            else:
                query = query.filter(
                    OrgArtifact.status == OrgArtifactStatus.uploaded,
                    OrgArtifact.current_file_version_id.isnot(None),
                    OrgArtifact.audited_file_version_id.is_(None),
                    OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                )
        elif audit_filter == "needs_correction":
            query = query.filter(
                OrgArtifact.status == OrgArtifactStatus.uploaded,
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction,
            )
        elif audit_filter == "audited":
            query = query.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
            )
        elif audit_filter == "changed":
            query = query.filter(
                OrgArtifact.current_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id.isnot(None),
                OrgArtifact.audited_file_version_id != OrgArtifact.current_file_version_id,
            )

    org_name = org.name if org else f"org_{org_id}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Artifacts"
    headers = [
        "Тематика",
        "Домен",
        "Показатель",
        "Сокращенное",
        "Пункт",
        "КБ",
        "Артефакт",
        "Статус",
        "Статус аудирования",
        "Версия текущая",
        "Файл текущий",
        "Дата загрузки",
        "Кто загрузил",
        "Дата изменения",
        "Кто изменил",
        "Версия проаудированная",
        "Проаудировано (когда)",
        "Проаудировано (кто)",
        "Комментарий",
        "Комментарий (кто)",
        "Комментарий (когда)",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"

    def fmt_dt(dt: datetime | None) -> str:
        return dt.isoformat(sep=" ", timespec="seconds") if dt else ""

    def audit_status_label(oa: OrgArtifact) -> str:
        if not oa.current_file_version_id:
            return "Нет файла"
        if getattr(oa, "review_status", None) == OrgArtifactReviewStatus.needs_correction:
            return "Требует корректировки"
        if (
            getattr(oa, "review_status", None) == OrgArtifactReviewStatus.approved
            and oa.audited_file_version_id
            and oa.audited_file_version_id == oa.current_file_version_id
        ):
            return "Проаудировано"
        if oa.audited_file_version_id:
            return "Изменён"
        return "Требует аудита"

    def upload_status_label(oa: OrgArtifact) -> str:
        # Используем понятные русские статусы, а не сырые enum-значения.
        if oa.status == OrgArtifactStatus.uploaded:
            return "Загружен"
        return "Не загружен"

    for (oa, a, fv, created_by, updated_by, aud_fv, aud_by, c_text, c_at, c_by) in query.all():
        ws.append(
            [
                a.topic,
                a.domain,
                a.indicator_name,
                a.short_name,
                a.achievement_item_no or "",
                a.kb_level,
                a.title,
                upload_status_label(oa),
                audit_status_label(oa),
                fv.version_no if fv else "",
                fv.original_filename if fv else "",
                fmt_dt(fv.created_at if fv else None),
                created_by.login if created_by else "",
                fmt_dt(oa.updated_at),
                updated_by.login if updated_by else "",
                aud_fv.version_no if aud_fv else "",
                fmt_dt(oa.audited_at),
                aud_by.login if aud_by else "",
                (c_text or ""),
                (c_by.login if c_by else ""),
                fmt_dt(c_at),
            ]
        )

    # Включаем Excel AutoFilter на шапке (выпадающие фильтры по колонкам)
    last_col = ws.cell(row=1, column=len(headers)).column_letter
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Простая автоподгонка ширин
    for idx, _ in enumerate(headers, start=1):
        col = ws.column_dimensions[ws.cell(row=1, column=idx).column_letter]
        col.width = min(max(12, len(str(headers[idx - 1])) + 2), 40)

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    # Starlette кодирует заголовки как latin-1, поэтому filename должен быть ASCII.
    # Для удобства добавляем RFC5987 filename* (UTF-8, percent-encoded).
    date_str = datetime.utcnow().date().isoformat()
    filename_ascii = f"audit_org{org_id}_{date_str}.xlsx"
    filename_utf8 = f"audit_{org_name}_{date_str}.xlsx"
    write_audit_log(
        db,
        actor=user,
        org_id=org_id,
        action="org_artifacts_export_xlsx",
        entity_type="org_artifacts",
        entity_id=str(org_id),
        after={
            "filename": filename_utf8,
            "filename_ascii": filename_ascii,
            "format": "xlsx",
            "endpoint": "/auditor/artifacts/export.xlsx",
            "filters": {
                "topic": topic,
                "domain": domain,
                "indicator": indicator,
                "kb_level": kb_level,
                "short_name": short_name,
                "q": q,
                "status": status,
                "audit_status": audit_status,
                "level": level_filter or (level or ""),
            },
        },
        request=request,
    )
    db.commit()

    cd = f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{quote(filename_utf8)}"
    headers_resp = {"Content-Disposition": cd}
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


def _download_content_disposition(filename_utf8: str, *, fallback_prefix: str = "download") -> str:
    """
    Build RFC5987-compatible Content-Disposition header.
    Starlette encodes headers as latin-1, so `filename=` must be ASCII.
    We provide UTF-8 via `filename*=` for browsers that support it.
    """
    name = (filename_utf8 or "").strip() or fallback_prefix
    ext = Path(name).suffix
    if ext and len(ext) <= 10:
        ascii_name = f"{fallback_prefix}{ext}"
    else:
        ascii_name = fallback_prefix
    return f'attachment; filename="{ascii_name}"; filename*=UTF-8\'\'{quote(name)}'


def _inline_content_disposition(filename_utf8: str, *, fallback_prefix: str = "inline") -> str:
    """
    Like _download_content_disposition, but for inline viewing (browser preview).
    """
    cd = _download_content_disposition(filename_utf8, fallback_prefix=fallback_prefix)
    if cd.lower().startswith("attachment;"):
        return "inline;" + cd[len("attachment;") :]
    return cd.replace("attachment", "inline", 1)


_OFFICE_EXTS = {".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx"}


def _is_office_file(filename: str, content_type: str) -> bool:
    ext = Path((filename or "").strip()).suffix.lower()
    if ext in _OFFICE_EXTS:
        return True
    ct = (content_type or "").lower()
    if ct.startswith("application/vnd.openxmlformats-officedocument."):
        return True
    if ct in ("application/msword", "application/vnd.ms-excel", "application/vnd.ms-powerpoint"):
        return True
    return False


def _get_or_build_pdf_preview(db: Session, fv: FileVersion) -> tuple[bytes, str]:
    """
    Return (pdf_bytes, error_message). error_message == "" when ok.
    Caches result in file_previews.
    """
    if not fv or not fv.blob:
        return b"", "Файл не найден"

    prev = db.query(FilePreview).filter(FilePreview.file_version_id == fv.id).one_or_none()
    if prev and prev.preview_blob and not prev.last_error:
        return prev.preview_blob, ""

    # Avoid tight failure loops (e.g. broken file) — backoff 2 minutes
    if prev and prev.last_error and prev.last_error_at:
        age_s = (datetime.utcnow() - prev.last_error_at.replace(tzinfo=None)).total_seconds()
        if age_s < 120:
            return b"", f"Превью временно недоступно (повторите позже): {prev.last_error}"

    in_name = (fv.original_filename or "").strip() or "file"
    ext = Path(in_name).suffix.lower() or ".bin"
    if ext not in _OFFICE_EXTS:
        # Ensure LO gets a sane extension
        ext = ".docx"

    try:
        with tempfile.TemporaryDirectory(prefix="preview_") as td:
            td_path = Path(td)
            in_path = td_path / f"input{ext}"
            in_path.write_bytes(fv.blob)

            cmd = [
                "soffice",
                "--headless",
                "--nologo",
                "--nolockcheck",
                "--norestore",
                "--nodefault",
                "--convert-to",
                "pdf",
                "--outdir",
                str(td_path),
                str(in_path),
            ]
            env = os.environ.copy()
            env.setdefault("HOME", "/tmp")
            res = subprocess.run(cmd, capture_output=True, text=True, timeout=60, env=env)
            if res.returncode != 0:
                msg = (res.stderr or res.stdout or "LibreOffice conversion failed").strip()
                msg = " ".join(msg.split())[:400]
                raise RuntimeError(msg)

            out_path = td_path / (in_path.stem + ".pdf")
            if not out_path.exists():
                pdfs = list(td_path.glob("*.pdf"))
                out_path = pdfs[0] if pdfs else out_path
            if not out_path.exists():
                raise RuntimeError("Не удалось получить PDF после конвертации")

            pdf_bytes = out_path.read_bytes()
            if not pdf_bytes:
                raise RuntimeError("Пустой PDF после конвертации")

            sha = hashlib.sha256(pdf_bytes).hexdigest()
            if not prev:
                prev = FilePreview(file_version_id=fv.id)
                db.add(prev)
            prev.preview_mime = "application/pdf"
            prev.preview_blob = pdf_bytes
            prev.preview_size_bytes = len(pdf_bytes)
            prev.preview_sha256 = sha
            prev.last_error = ""
            prev.last_error_at = None
            prev.created_at = datetime.utcnow()
            db.flush()
            return pdf_bytes, ""
    except Exception as e:
        msg = str(e).strip()
        msg = " ".join(msg.split())[:400]
        if not prev:
            prev = FilePreview(file_version_id=fv.id)
            db.add(prev)
        prev.preview_blob = None
        prev.preview_size_bytes = 0
        prev.preview_sha256 = ""
        prev.last_error = msg or "Ошибка конвертации"
        prev.last_error_at = datetime.utcnow()
        db.flush()
        return b"", prev.last_error


def _require_auditor_or_admin_for_org(db: Session, user: User, org_id: int) -> None:
    # Global admin (user.is_admin) must have access to all orgs regardless of per-org membership.
    # Otherwise admin sees org list but gets 403 on open/export.
    if getattr(user, "is_admin", False):
        return
    role = get_user_role_for_org(db, user, org_id)
    if role not in (Role.auditor, Role.admin):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")


@router.get("/auditor/files", response_class=HTMLResponse)
def auditor_files_explorer(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    path: str | None = None,
) -> HTMLResponse:
    orgs = _get_accessible_orgs_for_auditor(db, user)
    if not orgs:
        return templates.TemplateResponse(
            "empty.html",
            {"request": request, "user": user, "message": "Нет доступных организаций. Обратитесь к администратору."},
        )

    allowed_ids = {o.id for o in orgs}
    if not org_id or org_id not in allowed_ids:
        resp = templates.TemplateResponse(
            "select_org.html",
            {
                "request": request,
                "user": user,
                "container_class": "container-wide",
                "title": "Файлы",
                "subtitle": "Выберите организацию, чтобы открыть файловый список.",
                "action_path": "/auditor/files",
                "orgs": orgs,
            },
        )
        resp.headers["Cache-Control"] = "no-store, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp
    selected_org_id = org_id
    _require_auditor_or_admin_for_org(db, user, selected_org_id)
    role = get_user_role_for_org(db, user, selected_org_id)
    can_delete_files = role == Role.admin
    can_view_history = role in (Role.admin, Role.auditor)
    current_url = request.url.path + (f"?{request.url.query}" if request.url.query else "")

    _ensure_org_artifacts_materialized(db, selected_org_id)
    db.commit()

    CreatedBy = aliased(User)
    UpdatedBy = aliased(User)
    q_rows = (
        db.query(OrgArtifact, Artifact, FileVersion, CreatedBy, UpdatedBy)
        .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
        .outerjoin(FileVersion, FileVersion.id == OrgArtifact.current_file_version_id)
        .outerjoin(CreatedBy, CreatedBy.id == FileVersion.created_by_user_id)
        .outerjoin(UpdatedBy, UpdatedBy.id == OrgArtifact.updated_by_user_id)
        .filter(OrgArtifact.org_id == selected_org_id)
    )
    rows = q_rows.all()

    path = (path or "").strip().strip("/")
    cur_segments = [p for p in path.split("/") if p] if path else []

    subfolders: dict[str, int] = {}
    leaf_items: list[dict] = []
    for (oa, a, fv, created_by, updated_by) in rows:
        segs = _split_short_name(a.short_name)
        if not segs:
            continue
        if segs[: len(cur_segments)] != cur_segments:
            continue
        if len(segs) > len(cur_segments):
            nxt = segs[len(cur_segments)]
            subfolders[nxt] = subfolders.get(nxt, 0) + 1
        else:
            leaf_items.append(
                {
                    "oa": oa,
                    "a": a,
                    "fv": fv,
                    "uploaded_by": created_by.login if created_by else "",
                    "updated_by": updated_by.login if updated_by else "",
                }
            )

    folders_sorted = sorted(subfolders.items(), key=lambda x: x[0])
    leaf_items.sort(key=lambda r: (r["a"].short_name, r["a"].achievement_item_no or 0))

    crumbs = []
    acc = []
    for s in cur_segments:
        acc.append(s)
        crumbs.append({"name": s, "path": "/".join(acc)})

    chat_unread_total, _unread_by_thread_id, _unread_by_oa_id = _get_chat_unread_for_org(
        db, user=user, org_id=int(selected_org_id), only_org_artifact_ids=None
    )

    resp = templates.TemplateResponse(
        "auditor_files.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "can_delete_files": can_delete_files,
            "can_view_history": can_view_history,
            "current_url": current_url,
            "path": path,
            "crumbs": crumbs,
            "folders": folders_sorted,
            "leaf_items": leaf_items,
            "chat_unread_total": chat_unread_total,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/auditor/index-kb", response_class=HTMLResponse)
def auditor_index_kb_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    sheet: str | None = None,
    q: str | None = None,
) -> HTMLResponse:
    orgs_all = _get_accessible_orgs_for_auditor(db, user)
    orgs = orgs_all
    if not orgs:
        return templates.TemplateResponse(
            "empty.html",
            {"request": request, "user": user, "message": "Нет доступных организаций. Обратитесь к администратору."},
        )

    allowed_ids = {o.id for o in orgs}
    selected_org_id: int | None = None
    org_picker_error = False
    if org_id and org_id in allowed_ids:
        selected_org_id = org_id
        _require_auditor_or_admin_for_org(db, user, selected_org_id)
    else:
        # No org selected (or invalid org_id) — show tiles, highlight picker.
        org_picker_error = True
    df, dt, _p_start, _p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)

    available_sheets: list[str] = []
    if get_uib_template_from_db(db):
        available_sheets.append(UIB_SHEET_NAME)
    if get_szi_template_from_db(db):
        available_sheets.append(SZI_SHEET_NAME)
    err = None if available_sheets else "Шаблоны Индекса КБ не загружены в БД (нужны seed‑миграции)."

    chat_unread_total = 0
    if selected_org_id:
        chat_unread_total, _unread_by_thread_id, _unread_by_oa_id = _get_chat_unread_for_org(
            db, user=user, org_id=int(selected_org_id), only_org_artifact_ids=None
        )

    resp = templates.TemplateResponse(
        "auditor_index_kb.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "date_range_error": range_err or "",
            "sheet_names": INDEX_KB_SHEET_TILES,
            "available_sheets": available_sheets,
            "error": err,
            "base_prefix": "/auditor",
            "files_base": "/auditor/files",
            "artifacts_base": "/auditor/artifacts",
            "show_org_selector": True,
            "org_picker_error": org_picker_error,
            "chat_unread_total": chat_unread_total,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/auditor/index-kb/uib", response_class=HTMLResponse)
def auditor_index_kb_uib_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> HTMLResponse:
    orgs_all = _get_accessible_orgs_for_auditor(db, user)
    orgs = orgs_all
    if not orgs:
        return templates.TemplateResponse(
            "empty.html",
            {"request": request, "user": user, "message": "Нет доступных организаций. Обратитесь к администратору."},
        )
    allowed_ids = {o.id for o in orgs}
    selected_org_id: int | None = None
    org_picker_error = False
    if org_id and org_id in allowed_ids:
        selected_org_id = org_id
        _require_auditor_or_admin_for_org(db, user, selected_org_id)
    else:
        org_picker_error = True
    df, dt, p_start, p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)

    # Cache full view for snappy reloads. Keyed by (org_id, template_rev).
    global _UIB_VIEW_CACHE
    try:
        _UIB_VIEW_CACHE
    except NameError:
        _UIB_VIEW_CACHE = {}  # type: ignore[var-annotated]

    rows: list[object] = []
    summary_rows: list[object] = []
    org = None
    chat_unread_total = 0
    if selected_org_id:
        chat_unread_total, _unread_by_thread_id, _unread_by_oa_id = _get_chat_unread_for_org(
            db, user=user, org_id=int(selected_org_id), only_org_artifact_ids=None
        )
        if not get_uib_template_from_db(db):
            resp = templates.TemplateResponse(
                "auditor_index_kb_uib.html",
                {
                    "request": request,
                    "user": user,
                    "container_class": "container-wide",
                    "orgs": orgs,
                    "selected_org_id": selected_org_id,
                    "error": "Шаблон УИБ не загружен в БД. Проверьте, что применены миграции Alembic (в т.ч. seed-миграция).",
                    "rows": [],
                    "summary_rows": [],
                    "sheet_name": UIB_SHEET_NAME,
                    "org": None,
                    "base_prefix": "/auditor",
                    "files_base": "/auditor/files",
                    "artifacts_base": "/auditor/artifacts",
                    "show_org_selector": True,
                    "readonly": False,
                    "org_picker_error": org_picker_error,
                    "date_from": (df.isoformat() if df else ""),
                    "date_to": (dt.isoformat() if dt else ""),
                    "date_range_error": range_err or "",
                    "chat_unread_total": chat_unread_total,
                },
                status_code=200,
            )
            resp.headers["Cache-Control"] = "no-store, max-age=0"
            resp.headers["Pragma"] = "no-cache"
            return resp

        tpl_rev = get_uib_template_rev(db)
        cache_key = (int(selected_org_id), int(tpl_rev), (p_start.isoformat() if p_start else ""), (p_end.isoformat() if p_end else ""))
        now = time.time()
        cached = _UIB_VIEW_CACHE.get(cache_key)  # type: ignore[name-defined]
        if cached and (now - float(cached[0])) < 20.0:
            org, rows, summary_rows = cached[1], cached[2], cached[3]
        else:
            org, tpl, rows = build_uib_view(db, org_id=selected_org_id, actor=user, range_start=p_start, range_end=p_end)
            from app.index_kb.uib_sheet import compute_uib_summary

            summary_rows = compute_uib_summary(rows)
            _UIB_VIEW_CACHE[cache_key] = (now, org, rows, summary_rows)  # type: ignore[name-defined]
    resp = templates.TemplateResponse(
        "auditor_index_kb_uib.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "org": org,
            "sheet_name": UIB_SHEET_NAME,
            "rows": rows,
            "summary_rows": summary_rows,
            "error": None,
            "base_prefix": "/auditor",
            "files_base": "/auditor/files",
            "artifacts_base": "/auditor/artifacts",
            "show_org_selector": True,
            "readonly": False,
            "org_picker_error": org_picker_error,
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "date_range_error": range_err or "",
            "chat_unread_total": chat_unread_total,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


def _build_uib_export_xlsx(
    *,
    org_name: str,
    sheet_title: str,
    summary_rows: list[object],
    rows: list[object],
) -> bytes:
    """
    Export UIB view to a single-sheet XLSX:
    - top: summary table
    - below: main table (group rows + items)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title or "Лист"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    fill_hdr = PatternFill("solid", fgColor="F2F4F7")

    def _set_row(values: list[object], *, bold_row: bool = False, fill: PatternFill | None = None) -> None:
        ws.append(values)
        r = ws.max_row
        if bold_row or fill:
            for c in range(1, len(values) + 1):
                cell = ws.cell(row=r, column=c)
                if bold_row:
                    cell.font = bold
                if fill:
                    cell.fill = fill
                cell.alignment = center if bold_row else wrap

    # Title
    _set_row([f"Индекс КБ · {sheet_title}", "", "", "", "", "", ""], bold_row=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    _set_row([f"Организация: {org_name}", "", "", "", "", "", ""], bold_row=False)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws["A2"].alignment = Alignment(vertical="center")
    ws.append([])

    # Summary table
    _set_row(["Итоговая таблица"], bold_row=True, fill=fill_hdr)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
    _set_row(
        [
            "Категория",
            "Сокращение",
            "КБ3",
            "КБ2",
            "КБ1",
            "Расчетный показатель 2025",
            "Текущий показатель",
        ],
        bold_row=True,
        fill=fill_hdr,
    )
    for s in summary_rows:
        ws.append(
            [
                getattr(s, "title", ""),
                getattr(s, "short_name", ""),
                getattr(s, "kb3", None),
                getattr(s, "kb2", None),
                getattr(s, "kb1", None),
                getattr(s, "calc_2025", None),
                getattr(s, "current", None),
            ]
        )
    # number formats for summary numeric cols
    for r in range(6, ws.max_row + 1):
        for c in range(3, 8):
            ws.cell(row=r, column=c).number_format = "0.00"

    ws.append([])
    ws.append([])

    # Main table
    start_main = ws.max_row + 1
    _set_row(["Таблица требований"], bold_row=True, fill=fill_hdr)
    ws.merge_cells(start_row=start_main, start_column=1, end_row=start_main, end_column=6)
    ws.cell(row=start_main, column=1).alignment = Alignment(horizontal="left", vertical="center")
    _set_row(["Требование", "Сокращение", "КБ3", "КБ2", "КБ1", "Источник"], bold_row=True, fill=fill_hdr)

    # rows are UibRowView
    for rv in rows:
        row = getattr(rv, "row", None)
        kind = getattr(row, "kind", "") if row else ""
        if kind == "group":
            ws.append([getattr(row, "title", ""), getattr(row, "short_name", ""), "", "", "", ""])
            ws.cell(row=ws.max_row, column=1).font = bold
            ws.cell(row=ws.max_row, column=1).fill = fill_hdr
            for c in range(1, 7):
                ws.cell(row=ws.max_row, column=c).alignment = wrap
            continue

        src = getattr(rv, "source", "")
        v3 = getattr(rv, "kb3", None)
        v2 = getattr(rv, "kb2", None)
        v1 = getattr(rv, "kb1", None)
        ws.append([getattr(row, "title", ""), getattr(row, "short_name", ""), v3, v2, v1, src])
        ws.cell(row=ws.max_row, column=1).alignment = wrap
        ws.cell(row=ws.max_row, column=2).alignment = wrap
        for c in (3, 4, 5):
            ws.cell(row=ws.max_row, column=c).number_format = "0.0"
            ws.cell(row=ws.max_row, column=c).alignment = center

    # Freeze header for main table (keep UX for long lists)
    ws.freeze_panes = ws.cell(row=start_main + 2, column=1)  # after main header row

    # Simple column widths
    widths = {1: 70, 2: 18, 3: 10, 4: 10, 5: 10, 6: 12, 7: 22}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/auditor/index-kb/uib/export.xlsx")
def auditor_index_kb_uib_export_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int = 0,
    date_from: str | None = None,
    date_to: str | None = None,
) -> Response:
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id обязателен")
    _require_auditor_or_admin_for_org(db, user, org_id)
    if not get_uib_template_from_db(db):
        raise HTTPException(status_code=400, detail="Шаблон УИБ не загружен в БД (нужна seed-миграция).")

    org_obj = db.get(Organization, org_id)
    df, dt, p_start, p_end, _range_err = _parse_date_range_bounds_utc(date_from, date_to)
    org, tpl, rows = build_uib_view(db, org_id=org_id, actor=user, range_start=p_start, range_end=p_end)
    from app.index_kb.uib_sheet import compute_uib_summary

    summary_rows = compute_uib_summary(rows)
    content = _build_uib_export_xlsx(org_name=org.name, sheet_title=UIB_SHEET_NAME, summary_rows=summary_rows, rows=rows)

    date_str = datetime.utcnow().date().isoformat()
    filename_utf8 = f"uib_{org.name}_{date_str}.xlsx"
    cd = _download_content_disposition(filename_utf8, fallback_prefix=f"uib_org{org_id}_{date_str}")

    write_audit_log(
        db,
        actor=user,
        org_id=int(org_id),
        action="index_kb_export_uib_xlsx",
        entity_type="index_kb",
        entity_id=f"uib:{int(org_id)}",
        after={
            "sheet": UIB_SHEET_NAME,
            "format": "xlsx",
            "filename": filename_utf8,
            "endpoint": "/auditor/index-kb/uib/export.xlsx",
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
        },
        request=request,
    )
    db.commit()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )


@router.get("/my/index-kb/uib/export.xlsx")
def my_index_kb_uib_export_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int = 0,
    date_from: str | None = None,
    date_to: str | None = None,
) -> Response:
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id обязателен")
    orgs, org = _get_customer_selected_org(db, user, org_id)
    if not get_uib_template_from_db(db):
        raise HTTPException(status_code=400, detail="Шаблон УИБ не загружен в БД (нужна seed-миграция).")

    df, dt, p_start, p_end, _range_err = _parse_date_range_bounds_utc(date_from, date_to)
    org_obj, tpl, rows = build_uib_view(db, org_id=org.id, actor=user, range_start=p_start, range_end=p_end)
    from app.index_kb.uib_sheet import compute_uib_summary

    summary_rows = compute_uib_summary(rows)
    content = _build_uib_export_xlsx(org_name=org_obj.name, sheet_title=UIB_SHEET_NAME, summary_rows=summary_rows, rows=rows)

    date_str = datetime.utcnow().date().isoformat()
    filename_utf8 = f"uib_{org_obj.name}_{date_str}.xlsx"
    cd = _download_content_disposition(filename_utf8, fallback_prefix=f"uib_org{org_id}_{date_str}")

    write_audit_log(
        db,
        actor=user,
        org_id=int(org.id),
        action="index_kb_export_uib_xlsx",
        entity_type="index_kb",
        entity_id=f"uib:{int(org.id)}",
        after={
            "sheet": UIB_SHEET_NAME,
            "format": "xlsx",
            "filename": filename_utf8,
            "endpoint": "/my/index-kb/uib/export.xlsx",
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
        },
        request=request,
    )
    db.commit()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )


@router.get("/auditor/index-kb/szi", response_class=HTMLResponse)
def auditor_index_kb_szi_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> HTMLResponse:
    t0 = time.perf_counter()
    orgs_all = _get_accessible_orgs_for_auditor(db, user)
    orgs = orgs_all
    if not orgs:
        return templates.TemplateResponse(
            "empty.html",
            {"request": request, "user": user, "message": "Нет доступных организаций. Обратитесь к администратору."},
        )

    allowed_ids = {o.id for o in orgs}
    selected_org_id: int | None = None
    org_picker_error = False
    if org_id and org_id in allowed_ids:
        selected_org_id = org_id
        _require_auditor_or_admin_for_org(db, user, selected_org_id)
    else:
        org_picker_error = True
    df, dt, p_start, p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)

    # Cache full view for snappy reloads. Keyed by (org_id, template_mtime_ns).
    global _SZI_VIEW_CACHE
    try:
        _SZI_VIEW_CACHE
    except NameError:
        _SZI_VIEW_CACHE = {}  # type: ignore[var-annotated]

    rows: list[object] = []
    summary_rows: list[object] = []
    org = None
    items_count = 0
    groups_count = 0
    chat_unread_total = 0
    if selected_org_id:
        chat_unread_total, _unread_by_thread_id, _unread_by_oa_id = _get_chat_unread_for_org(
            db, user=user, org_id=int(selected_org_id), only_org_artifact_ids=None
        )
        if not get_szi_template_from_db(db):
            resp = templates.TemplateResponse(
                "auditor_index_kb_szi.html",
                {
                    "request": request,
                    "user": user,
                    "container_class": "container-wide",
                    "orgs": orgs,
                    "selected_org_id": selected_org_id,
                    "error": "Шаблон СЗИ не загружен в БД. Проверьте, что применены миграции Alembic (в т.ч. seed-миграция).",
                    "rows": [],
                    "summary_rows": [],
                    "sheet_name": SZI_SHEET_NAME,
                    "org": None,
                    "base_prefix": "/auditor",
                    "files_base": "/auditor/files",
                    "artifacts_base": "/auditor/artifacts",
                    "show_org_selector": True,
                    "readonly": False,
                    "org_picker_error": org_picker_error,
                    "date_from": (df.isoformat() if df else ""),
                    "date_to": (dt.isoformat() if dt else ""),
                    "date_range_error": range_err or "",
                    "chat_unread_total": chat_unread_total,
                },
                status_code=200,
            )
            resp.headers["Cache-Control"] = "no-store, max-age=0"
            resp.headers["Pragma"] = "no-cache"
            return resp
        t1 = time.perf_counter()
        tpl_rev = get_szi_template_rev(db)
        cache_key = (int(selected_org_id), int(tpl_rev), (p_start.isoformat() if p_start else ""), (p_end.isoformat() if p_end else ""))
        now = time.time()
        cached = _SZI_VIEW_CACHE.get(cache_key)  # type: ignore[name-defined]
        if cached and (now - float(cached[0])) < 20.0:
            org, rows, summary_rows = cached[1], cached[2], cached[3]
            t_build = 0.0
            t_sum = 0.0
        else:
            org, tpl, rows = build_szi_view(db, org_id=selected_org_id, actor=user, range_start=p_start, range_end=p_end)
            t2 = time.perf_counter()
            from app.index_kb.szi_sheet import compute_szi_summary

            summary_rows = compute_szi_summary(rows)
            t3 = time.perf_counter()
            _SZI_VIEW_CACHE[cache_key] = (now, org, rows, summary_rows)  # type: ignore[name-defined]
            t_build = (t2 - t1) * 1000.0
            t_sum = (t3 - t2) * 1000.0
        try:
            items_count = sum(1 for rv in rows if getattr(getattr(rv, "row", None), "kind", "") == "item")
            groups_count = sum(1 for rv in rows if getattr(getattr(rv, "row", None), "kind", "") == "group")
        except Exception:
            items_count = 0
            groups_count = 0
        t4 = time.perf_counter()
        print(
            f"[perf] szi org_id={selected_org_id} build_ms={t_build:.1f} sum_ms={t_sum:.1f} counts_ms={(t4 - t1) * 1000.0:.1f} total_ms={(t4 - t0) * 1000.0:.1f}"
        )

    resp = templates.TemplateResponse(
        "auditor_index_kb_szi.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "org": org,
            "sheet_name": SZI_SHEET_NAME,
            "rows": rows,
            "summary_rows": summary_rows,
            "items_count": items_count,
            "groups_count": groups_count,
            "error": None,
            "base_prefix": "/auditor",
            "files_base": "/auditor/files",
            "artifacts_base": "/auditor/artifacts",
            "show_org_selector": True,
            "readonly": False,
            "org_picker_error": org_picker_error,
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "date_range_error": range_err or "",
            "chat_unread_total": chat_unread_total,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.post("/auditor/index-kb/szi/manual")
def auditor_index_kb_szi_save_manual(
    request: Request,
    org_id: int = Form(...),
    row_key: str = Form(...),
    value: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    _require_auditor_or_admin_for_org(db, user, org_id)
    try:
        v = float(value) if (value or "").strip() else 0.0
    except Exception:
        v = 0.0
    from app.index_kb.szi_sheet import upsert_manual_value

    upsert_manual_value(db, org_id=org_id, sheet_name=SZI_SHEET_NAME, row_key=row_key, value=v, actor=user)
    db.commit()
    ref = request.headers.get("referer") or f"/auditor/index-kb/szi?org_id={org_id}"
    return _redirect(ref)


@router.get("/auditor/index-kb/szi/export.xlsx")
def auditor_index_kb_szi_export_xlsx(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int = 0,
    date_from: str | None = None,
    date_to: str | None = None,
) -> Response:
    if not org_id:
        raise HTTPException(status_code=400, detail="org_id обязателен")
    _require_auditor_or_admin_for_org(db, user, org_id)
    if not get_szi_template_from_db(db):
        raise HTTPException(status_code=400, detail="Шаблон СЗИ не загружен в БД (нужна seed-миграция).")

    org_obj = db.get(Organization, org_id)
    df, dt, p_start, p_end, _range_err = _parse_date_range_bounds_utc(date_from, date_to)
    org, tpl, rows = build_szi_view(db, org_id=org_id, actor=user, range_start=p_start, range_end=p_end)
    from app.index_kb.szi_sheet import compute_szi_summary

    summary_rows = compute_szi_summary(rows)
    content = _build_uib_export_xlsx(org_name=org.name, sheet_title=SZI_SHEET_NAME, summary_rows=summary_rows, rows=rows)

    date_str = datetime.utcnow().date().isoformat()
    filename_utf8 = f"szi_{org.name}_{date_str}.xlsx"
    cd = _download_content_disposition(filename_utf8, fallback_prefix=f"szi_org{org_id}_{date_str}")

    write_audit_log(
        db,
        actor=user,
        org_id=int(org_id),
        action="index_kb_export_szi_xlsx",
        entity_type="index_kb",
        entity_id=f"szi:{int(org_id)}",
        after={
            "sheet": SZI_SHEET_NAME,
            "format": "xlsx",
            "filename": filename_utf8,
            "endpoint": "/auditor/index-kb/szi/export.xlsx",
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
        },
        request=request,
    )
    db.commit()
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": cd},
    )


# --- Chat (auditor/customer), MVP polling ---


@router.get("/auditor/chat", response_class=HTMLResponse)
def auditor_org_chat_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int = 0,
) -> HTMLResponse:
    if int(org_id or 0) <= 0:
        raise HTTPException(status_code=400, detail="org_id обязателен")
    _require_chat_access(db, user, int(org_id), allow_customer=False)
    org = db.get(Organization, int(org_id))
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")
    t = _get_or_create_chat_thread(db, org_id=int(org_id), org_artifact_id=None, actor=user)
    db.commit()
    return _render_chat_thread_page(
        request=request,
        db=db,
        user=user,
        thread=t,
        org=org,
        artifact_label=None,
        back_href=f"/auditor/artifacts?org_id={int(org_id)}",
        back_label="К аудиту",
    )


@router.get("/auditor/org_artifacts/{org_artifact_id}/chat", response_class=HTMLResponse)
def auditor_artifact_chat_page(
    org_artifact_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> HTMLResponse:
    oa = db.get(OrgArtifact, int(org_artifact_id))
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    _require_chat_access(db, user, int(oa.org_id), allow_customer=False)
    org = db.get(Organization, int(oa.org_id))
    a = db.get(Artifact, int(oa.artifact_id))
    label = (a.short_name if a else "") or ""
    if a and a.title:
        label = f"{label} · {a.title}" if label else a.title
    t = _get_or_create_chat_thread(db, org_id=int(oa.org_id), org_artifact_id=int(oa.id), actor=user)
    db.commit()
    return _render_chat_thread_page(
        request=request,
        db=db,
        user=user,
        thread=t,
        org=org,
        artifact_label=label,
        back_href=f"/auditor/artifacts?org_id={int(oa.org_id)}",
        back_label="К аудиту",
    )


@router.get("/my/chat", response_class=HTMLResponse)
def my_org_chat_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    org_id: int = 0,
) -> HTMLResponse:
    if int(org_id or 0) <= 0:
        raise HTTPException(status_code=400, detail="org_id обязателен")
    _require_chat_access(db, user, int(org_id), allow_customer=True)
    org = db.get(Organization, int(org_id))
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")
    t = _get_or_create_chat_thread(db, org_id=int(org_id), org_artifact_id=None, actor=user)
    db.commit()
    return _render_chat_thread_page(
        request=request,
        db=db,
        user=user,
        thread=t,
        org=org,
        artifact_label=None,
        back_href=f"/my/artifacts?org_id={int(org_id)}",
        back_label="К артефактам",
    )


@router.get("/my/org_artifacts/{org_artifact_id}/chat", response_class=HTMLResponse)
def my_artifact_chat_page(
    org_artifact_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> HTMLResponse:
    oa = db.get(OrgArtifact, int(org_artifact_id))
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    _require_chat_access(db, user, int(oa.org_id), allow_customer=True)
    org = db.get(Organization, int(oa.org_id))
    a = db.get(Artifact, int(oa.artifact_id))
    label = (a.short_name if a else "") or ""
    if a and a.title:
        label = f"{label} · {a.title}" if label else a.title
    t = _get_or_create_chat_thread(db, org_id=int(oa.org_id), org_artifact_id=int(oa.id), actor=user)
    db.commit()
    return _render_chat_thread_page(
        request=request,
        db=db,
        user=user,
        thread=t,
        org=org,
        artifact_label=label,
        back_href=f"/my/artifacts?org_id={int(oa.org_id)}",
        back_label="К артефактам",
    )


def _render_chat_thread_page(
    *,
    request: Request,
    db: Session,
    user: User,
    thread: ChatThread,
    org: Organization | None,
    artifact_label: str | None,
    back_href: str,
    back_label: str,
) -> HTMLResponse:
    # Sidebar thread list (org chat + artifact chats)
    all_threads = db.query(ChatThread).filter(ChatThread.org_id == int(thread.org_id)).all()
    total_unread, unread_by_thread_id, _unread_by_oa_id = _get_chat_unread_for_org(
        db, user=user, org_id=int(thread.org_id), only_org_artifact_ids=None
    )

    # Hide "empty" artifact threads in navigation by default:
    # a thread is considered visible if it has messages OR has unread OR is currently active.
    t_ids = [int(t.id) for t in all_threads]
    last_msg_rows = (
        db.query(ChatMessage.thread_id, func.max(ChatMessage.id).label("last_id"))
        .filter(ChatMessage.thread_id.in_(t_ids or [-1]))
        .group_by(ChatMessage.thread_id)
        .all()
    )
    has_messages_by_thread_id = {int(tid): (int(last_id or 0) > 0) for (tid, last_id) in last_msg_rows}
    oa_ids = sorted({int(t.org_artifact_id) for t in all_threads if t.org_artifact_id})
    oa_labels: dict[int, str] = {}
    if oa_ids:
        oa_rows = (
            db.query(OrgArtifact.id, Artifact.short_name, Artifact.title)
            .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
            .filter(OrgArtifact.id.in_(oa_ids))
            .all()
        )
        for oa_id, sn, title in oa_rows:
            label = (sn or "") or ""
            if title:
                label = f"{label} · {title}" if label else title
            oa_labels[int(oa_id)] = label or f"Артефакт #{int(oa_id)}"

    # Build navigation list: org chat first, then artifacts by short_name
    nav_items: list[dict] = []
    base_prefix = "/auditor" if (back_href or "").startswith("/auditor") else "/my"
    # Org chat link
    org_thread = next((t for t in all_threads if t.org_artifact_id is None), None)
    if org_thread:
        nav_items.append(
            {
                "kind": "org",
                "title": "Общий чат",
                "href": f"{base_prefix}/chat?org_id={int(thread.org_id)}",
                "thread_id": int(org_thread.id),
                "unread": int(unread_by_thread_id.get(int(org_thread.id), 0)),
                "active": int(thread.id) == int(org_thread.id),
            }
        )
    # Artifact chats
    art_threads_all = [t for t in all_threads if t.org_artifact_id is not None]
    art_threads_visible: list[ChatThread] = []
    for t in art_threads_all:
        tid = int(t.id)
        if int(thread.id) == tid:
            art_threads_visible.append(t)
            continue
        if unread_by_thread_id.get(tid, 0) > 0:
            art_threads_visible.append(t)
            continue
        if has_messages_by_thread_id.get(tid, False):
            art_threads_visible.append(t)
            continue

    # Sort: unread first, then by label.
    art_threads_visible.sort(
        key=lambda t: (
            0 if int(unread_by_thread_id.get(int(t.id), 0)) > 0 else 1,
            oa_labels.get(int(t.org_artifact_id or 0), ""),
            int(t.org_artifact_id or 0),
        )
    )
    for t in art_threads_visible:
        oa_id = int(t.org_artifact_id or 0)
        nav_items.append(
            {
                "kind": "artifact",
                "title": oa_labels.get(oa_id, f"Артефакт #{oa_id}"),
                "href": f"{base_prefix}/org_artifacts/{oa_id}/chat",
                "thread_id": int(t.id),
                "unread": int(unread_by_thread_id.get(int(t.id), 0)),
                "active": int(thread.id) == int(t.id),
            }
        )

    msgs = (
        db.query(ChatMessage)
        .filter(ChatMessage.thread_id == int(thread.id))
        .options(joinedload(ChatMessage.author))
        .order_by(ChatMessage.id.asc())
        .limit(200)
        .all()
    )
    last_id = int(msgs[-1].id) if msgs else 0
    if last_id:
        r = (
            db.query(ChatThreadRead)
            .filter(ChatThreadRead.thread_id == int(thread.id), ChatThreadRead.user_id == int(user.id))
            .one_or_none()
        )
        if r:
            r.last_read_message_id = last_id
            r.last_read_at = datetime.utcnow()
        else:
            db.add(ChatThreadRead(thread_id=int(thread.id), user_id=int(user.id), last_read_message_id=last_id))
        write_audit_log(
            db,
            actor=user,
            org_id=int(thread.org_id),
            action="chat_read_update",
            entity_type="chat_thread",
            entity_id=str(int(thread.id)),
            after={"thread_id": int(thread.id), "last_read_message_id": last_id},
            request=request,
        )
        db.commit()

    resp = templates.TemplateResponse(
        "chat_thread.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "title": "Чат",
            "thread_id": int(thread.id),
            "org_name": (org.name if org else ""),
            "artifact_label": artifact_label or "",
            "back_href": back_href,
            "back_label": back_label,
            "current_user_login": user.login,
            "last_message_id": last_id,
            "threads_nav": nav_items,
            "threads_unread_total": int(total_unread),
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/api/chat/threads/{thread_id}/messages")
def api_chat_messages(
    thread_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    after_id: int = 0,
) -> dict:
    t = db.get(ChatThread, int(thread_id))
    if not t:
        raise HTTPException(status_code=404, detail="Тред не найден")
    _require_chat_access(db, user, int(t.org_id), allow_customer=True)
    q = db.query(ChatMessage).options(joinedload(ChatMessage.author)).filter(ChatMessage.thread_id == int(t.id))
    if int(after_id or 0) > 0:
        q = q.filter(ChatMessage.id > int(after_id))
    msgs = q.order_by(ChatMessage.id.asc()).limit(200).all()
    return {"messages": [_chat_message_to_dict(m) for m in msgs]}


@router.post("/api/chat/threads/{thread_id}/messages")
async def api_chat_send_message(
    thread_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    t = db.get(ChatThread, int(thread_id))
    if not t:
        raise HTTPException(status_code=404, detail="Тред не найден")
    _require_chat_access(db, user, int(t.org_id), allow_customer=True)
    payload = await request.json()
    body = str((payload or {}).get("body") or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Пустое сообщение")
    if len(body) > 5000:
        raise HTTPException(status_code=400, detail="Слишком длинное сообщение")
    m = ChatMessage(thread_id=int(t.id), author_user_id=int(user.id), body=body)
    db.add(m)
    db.flush()
    write_audit_log(
        db,
        actor=user,
        org_id=int(t.org_id),
        action="chat_message_create",
        entity_type="chat_message",
        entity_id=str(int(m.id)),
        after={"thread_id": int(t.id), "author_user_id": int(user.id)},
        request=request,
    )
    db.commit()
    m2 = db.query(ChatMessage).options(joinedload(ChatMessage.author)).filter(ChatMessage.id == int(m.id)).one()
    return {"message": _chat_message_to_dict(m2)}


@router.post("/api/chat/threads/{thread_id}/read")
async def api_chat_mark_read(
    thread_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict:
    t = db.get(ChatThread, int(thread_id))
    if not t:
        raise HTTPException(status_code=404, detail="Тред не найден")
    _require_chat_access(db, user, int(t.org_id), allow_customer=True)
    payload = await request.json()
    last_id = int((payload or {}).get("last_read_message_id") or 0)
    if last_id <= 0:
        return {"ok": True}
    r = (
        db.query(ChatThreadRead)
        .filter(ChatThreadRead.thread_id == int(t.id), ChatThreadRead.user_id == int(user.id))
        .one_or_none()
    )
    if r:
        if (r.last_read_message_id or 0) < last_id:
            r.last_read_message_id = last_id
        r.last_read_at = datetime.utcnow()
    else:
        db.add(ChatThreadRead(thread_id=int(t.id), user_id=int(user.id), last_read_message_id=last_id))
    write_audit_log(
        db,
        actor=user,
        org_id=int(t.org_id),
        action="chat_read_update",
        entity_type="chat_thread",
        entity_id=str(int(t.id)),
        after={"thread_id": int(t.id), "last_read_message_id": last_id},
        request=request,
    )
    db.commit()
    return {"ok": True}

@router.post("/auditor/index-kb/uib/manual")
def auditor_index_kb_uib_save_manual(
    request: Request,
    org_id: int = Form(...),
    row_key: str = Form(...),
    value: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    _require_auditor_or_admin_for_org(db, user, org_id)
    try:
        v = float((value or "").strip().replace(",", "."))
    except Exception:
        v = 0.0
    upsert_manual_value(db, org_id=org_id, sheet_name=UIB_SHEET_NAME, row_key=row_key, value=v, actor=user)
    db.commit()
    ref = request.headers.get("referer") or f"/auditor/index-kb/uib?org_id={org_id}"
    return _redirect(ref)


@router.get("/auditor/org_artifacts/{org_artifact_id}/download")
def auditor_download_current_file(
    org_artifact_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    version: int | None = None,
    fv_id: int | None = None,
) -> Response:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role not in (Role.auditor, Role.admin):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")
    # История версий доступна аудитору и админу.
    if version is not None and role not in (Role.admin, Role.auditor):
        raise HTTPException(status_code=403, detail="История версий недоступна")

    qv = db.query(FileVersion).filter(FileVersion.org_artifact_id == oa.id)
    if fv_id is not None:
        fv = qv.filter(FileVersion.id == int(fv_id)).one_or_none()
    elif version is not None:
        fv = qv.filter(FileVersion.version_no == int(version)).one_or_none()
    else:
        fv = db.get(FileVersion, oa.current_file_version_id) if oa.current_file_version_id else qv.order_by(FileVersion.version_no.desc()).first()
    if not fv or not fv.blob:
        raise HTTPException(status_code=404, detail="Файл не найден")

    headers = {"Content-Disposition": _download_content_disposition(fv.original_filename, fallback_prefix="artifact")}
    return Response(content=fv.blob, media_type=fv.content_type, headers=headers)


@router.get("/auditor/org_artifacts/{org_artifact_id}/content")
def auditor_view_content(
    org_artifact_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    version: int | None = None,
    mode: str | None = None,
) -> Response:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role not in (Role.auditor, Role.admin):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")

    qv = db.query(FileVersion).filter(FileVersion.org_artifact_id == oa.id)
    if version is not None:
        fv = qv.filter(FileVersion.version_no == version).one_or_none()
    else:
        fv = db.get(FileVersion, oa.current_file_version_id) if oa.current_file_version_id else qv.order_by(FileVersion.version_no.desc()).first()
    if not fv or not fv.blob:
        raise HTTPException(status_code=404, detail="Файл не найден")

    # mode=preview -> generate PDF preview for MS Office formats
    if (mode or "").lower() == "preview":
        if not _is_office_file(fv.original_filename or "", fv.content_type or ""):
            # for non-office just return inline original
            headers = {
                "Content-Disposition": _inline_content_disposition(fv.original_filename, fallback_prefix="artifact"),
                "Cache-Control": "no-store, max-age=0",
                "Pragma": "no-cache",
            }
            return Response(content=fv.blob, media_type=fv.content_type, headers=headers)

        pdf_bytes, err = _get_or_build_pdf_preview(db, fv)
        db.commit()
        if err or not pdf_bytes:
            raise HTTPException(status_code=503, detail=f"Превью недоступно: {err or 'ошибка конвертации'}")
        headers = {
            "Content-Disposition": _inline_content_disposition((fv.original_filename or "preview") + ".pdf", fallback_prefix="preview"),
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        }
        return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)

    headers = {
        "Content-Disposition": _inline_content_disposition(fv.original_filename, fallback_prefix="artifact"),
        "Cache-Control": "no-store, max-age=0",
        "Pragma": "no-cache",
    }
    return Response(content=fv.blob, media_type=fv.content_type, headers=headers)


@router.get("/auditor/org_artifacts/{org_artifact_id}/view", response_class=HTMLResponse)
def auditor_org_artifact_view_page(
    org_artifact_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    back: str | None = None,
    version: int | None = None,
) -> HTMLResponse:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role not in (Role.admin, Role.auditor):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")
    a = db.get(Artifact, oa.artifact_id)
    org = db.get(Organization, oa.org_id)
    if not a or not org:
        raise HTTPException(status_code=404, detail="Данные не найдены")

    qv = db.query(FileVersion).filter(FileVersion.org_artifact_id == oa.id)
    if version is not None:
        fv = qv.filter(FileVersion.version_no == version).one_or_none()
    else:
        fv = db.get(FileVersion, oa.current_file_version_id) if oa.current_file_version_id else qv.order_by(FileVersion.version_no.desc()).first()

    content_url = None
    viewer_kind = "none"
    content_type = ""
    filename = ""
    preview_error = ""
    if fv and fv.blob:
        content_type = (fv.content_type or "").lower()
        filename = fv.original_filename or ""
        base_qs = (f"version={fv.version_no}" if version is not None else "")
        content_url = f"/auditor/org_artifacts/{oa.id}/content" + (f"?{base_qs}" if base_qs else "")
        if "application/pdf" in content_type:
            viewer_kind = "pdf"
        elif content_type.startswith("image/"):
            viewer_kind = "image"
        elif content_type.startswith("audio/"):
            viewer_kind = "audio"
        elif content_type.startswith("video/"):
            viewer_kind = "video"
        elif content_type.startswith("text/") or content_type in ("application/json", "application/xml"):
            viewer_kind = "text"
        elif _is_office_file(filename, content_type):
            # Try to build preview proactively, so the page shows a friendly message on failure
            _, err = _get_or_build_pdf_preview(db, fv)
            db.commit()
            if err:
                viewer_kind = "unknown"
                preview_error = err
                content_url = None
            else:
                viewer_kind = "pdf"
                qs = base_qs + ("&" if base_qs else "") + "mode=preview"
                content_url = f"/auditor/org_artifacts/{oa.id}/content?{qs}"
        else:
            viewer_kind = "unknown"

    back_url = back or (request.headers.get("referer") or f"/auditor/artifacts?org_id={oa.org_id}")
    if "://" in back_url:
        back_url = f"/auditor/artifacts?org_id={oa.org_id}"

    # Unread badge for "chat by artifact" button on this page
    _total_unread, _unread_by_thread_id, unread_by_oa_id = _get_chat_unread_for_org(
        db, user=user, org_id=int(oa.org_id), only_org_artifact_ids=[int(oa.id)]
    )
    chat_unread_for_oa = int(unread_by_oa_id.get(int(oa.id), 0) if unread_by_oa_id else 0)

    resp = templates.TemplateResponse(
        "auditor_org_artifact_view.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "oa": oa,
            "a": a,
            "org": org,
            "fv": fv,
            "viewer_kind": viewer_kind,
            "content_url": content_url,
            "content_type": content_type,
            "filename": filename,
            "preview_error": preview_error,
            "back_url": back_url,
            "current_url": request.url.path + (f"?{request.url.query}" if request.url.query else ""),
            "chat_unread_for_oa": chat_unread_for_oa,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/auditor/org_artifacts/{org_artifact_id}/history", response_class=HTMLResponse)
def auditor_org_artifact_history_page(
    org_artifact_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    back: str | None = None,
) -> HTMLResponse:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role not in (Role.admin, Role.auditor):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")
    a = db.get(Artifact, oa.artifact_id)
    org = db.get(Organization, oa.org_id)
    if not a or not org:
        raise HTTPException(status_code=404, detail="Данные не найдены")

    CreatedBy = aliased(User)
    versions = (
        db.query(FileVersion, CreatedBy)
        .outerjoin(CreatedBy, CreatedBy.id == FileVersion.created_by_user_id)
        .filter(FileVersion.org_artifact_id == oa.id)
        .order_by(FileVersion.version_no.desc())
        .all()
    )
    version_rows = [
        {
            "id": fv.id,
            "version_no": fv.version_no,
            "original_filename": fv.original_filename,
            "created_at": fv.created_at,
            "created_by_login": created_by.login if created_by else "",
        }
        for (fv, created_by) in versions
    ]

    # audit events for this org_artifact
    Actor = aliased(User)
    logs = (
        db.query(AuditLog, Actor.login)
        .outerjoin(Actor, Actor.id == AuditLog.actor_user_id)
        .filter(AuditLog.entity_type == "org_artifact", AuditLog.entity_id == str(oa.id))
        .order_by(AuditLog.at.desc(), AuditLog.id.desc())
        .limit(250)
        .all()
    )
    action_labels = {
        "upload": "Загрузка",
        "delete_file": "Удаление файла",
        "nextcloud_import": "Синхронизация Nextcloud",
        "nextcloud_import_v2": "Синхронизация Nextcloud",
        "migration_import": "Импорт (миграция)",
        "patch": "Правка",
        "audit": "Проверено (аудит)",
        "comment": "Комментарий",
    }

    fv_meta = {}
    for (fv, _) in versions:
        fv_meta[fv.id] = {"version_no": fv.version_no, "filename": fv.original_filename or ""}

    def _fmt_status(v: object) -> str:
        s = (str(v) if v is not None else "").strip().lower()
        if s == "uploaded":
            return "Загружен"
        if s == "missing":
            return "Не загружен"
        return s or "—"

    def _fmt_fv(fid: object) -> str:
        if not fid:
            return "—"
        try:
            fid_int = int(fid)
        except Exception:
            return str(fid)
        m = fv_meta.get(fid_int)
        if not m:
            return f"fv#{fid_int}"
        fn = m.get("filename") or ""
        return f"v{m.get('version_no')} · {fn}" if fn else f"v{m.get('version_no')}"

    def _audit_state(cur_id: object, aud_id: object) -> str:
        if not cur_id:
            return "—"
        if not aud_id:
            return "Требует аудита"
        try:
            if int(aud_id) == int(cur_id):
                return "Проаудировано"
        except Exception:
            pass
        return "Изменён"

    def _human_details(action: str, before: dict | None, after: dict | None) -> str:
        b = before if isinstance(before, dict) else {}
        a = after if isinstance(after, dict) else {}

        status = a.get("status", b.get("status"))
        cur = a.get("current_file_version_id", b.get("current_file_version_id"))
        aud = a.get("audited_file_version_id", b.get("audited_file_version_id"))
        audit_state = _audit_state(cur, aud)

        parts: list[str] = []

        if action == "comment":
            txt = (a.get("comment") or "").strip()
            if txt:
                one = " ".join(txt.split())
                if len(one) > 140:
                    one = one[:137] + "…"
                return f"Комментарий: {one}"
            return ""

        if action == "upload":
            parts.append(f"Файл: {_fmt_fv(cur)}")
            parts.append(f"Статус: {_fmt_status(status)}")
            parts.append(f"Аудит: {audit_state}")
            return " · ".join([p for p in parts if p and p != "None"])

        if action == "audit":
            parts.append(f"Проаудировано: {_fmt_fv(aud or cur)}")
            parts.append(f"Аудит: {audit_state}")
            return " · ".join([p for p in parts if p and p != "None"])

        if action == "delete_file":
            prev_cur = b.get("current_file_version_id")
            if prev_cur:
                parts.append(f"Удалено: {_fmt_fv(prev_cur)}")
            parts.append(f"Статус: {_fmt_status(status or 'missing')}")
            parts.append(f"Аудит: {audit_state}")
            return " · ".join([p for p in parts if p and p != "None"])

        if action in ("nextcloud_import", "nextcloud_import_v2"):
            rp = (a.get("remote_path") or "").strip()
            if rp:
                parts.append(f"Источник: Nextcloud · {rp}")
            if status or cur or aud:
                parts.append(f"Статус: {_fmt_status(status)}")
                if cur:
                    parts.append(f"Текущая: {_fmt_fv(cur)}")
                parts.append(f"Аудит: {audit_state}")
            return " · ".join([p for p in parts if p and p != "None"])

        if action == "patch":
            if status or cur or aud:
                parts.append(f"Статус: {_fmt_status(status)}")
                if cur:
                    parts.append(f"Текущая: {_fmt_fv(cur)}")
                parts.append(f"Аудит: {audit_state}")
            return " · ".join([p for p in parts if p and p != "None"])

        # fallback: show the key bits if present
        if status:
            parts.append(f"Статус: {_fmt_status(status)}")
        if cur:
            parts.append(f"Текущая: {_fmt_fv(cur)}")
        if aud:
            parts.append(f"Аудит: {_fmt_fv(aud)} ({audit_state})")
        if a.get("remote_path"):
            parts.append(f"remote: {a.get('remote_path')}")
        return " · ".join([p for p in parts if p and p != "None"])

    events = []
    for (log, actor_login) in logs:
        events.append(
            {
                "at": log.at,
                "actor_login": actor_login or "",
                "action_label": action_labels.get(log.action, log.action),
                "details": _human_details(
                    log.action,
                    log.before_json if isinstance(log.before_json, dict) else None,
                    log.after_json if isinstance(log.after_json, dict) else None,
                ),
            }
        )

    back_url = back or (request.headers.get("referer") or f"/auditor/artifacts?org_id={oa.org_id}")
    if "://" in back_url:
        back_url = f"/auditor/artifacts?org_id={oa.org_id}"

    resp = templates.TemplateResponse(
        "auditor_org_artifact_history.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "oa": oa,
            "a": a,
            "org": org,
            "versions": version_rows,
            "events": events,
            "back_url": back_url,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.post("/auditor/org_artifacts/{org_artifact_id}/delete")
def admin_delete_current_file_for_org_artifact(
    org_artifact_id: int,
    request: Request,
    back: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role != Role.admin:
        raise HTTPException(status_code=403, detail="Требуются права admin")

    before = {
        "status": oa.status.value,
        "current_file_version_id": oa.current_file_version_id,
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at.isoformat() if oa.audited_at else None,
        "audited_by_user_id": oa.audited_by_user_id,
    }

    fv = db.get(FileVersion, oa.current_file_version_id) if oa.current_file_version_id else None
    if fv and (fv.storage_key or "").startswith("nextcloud:"):
        remote_path = (fv.storage_key or "")[len("nextcloud:") :].strip()
        if remote_path:
            # allow re-import on next sync after manual delete
            db.query(NextcloudRemoteFileState).filter(
                NextcloudRemoteFileState.org_id == oa.org_id,
                NextcloudRemoteFileState.remote_path == remote_path,
            ).delete(synchronize_session=False)

    oa.current_file_version_id = None
    oa.status = OrgArtifactStatus.missing
    oa.updated_at = datetime.utcnow()
    oa.updated_by_user_id = user.id
    oa.audited_file_version_id = None
    oa.audited_at = None
    oa.audited_by_user_id = None
    after = {
        "status": oa.status.value,
        "current_file_version_id": oa.current_file_version_id,
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at,
        "audited_by_user_id": oa.audited_by_user_id,
    }

    write_audit_log(
        db,
        actor=user,
        org_id=oa.org_id,
        action="delete_file",
        entity_type="org_artifact",
        entity_id=str(oa.id),
        before=before,
        after=after,
        request=request,
    )
    db.commit()

    # Prefer explicit back (sent by form), fallback to referer.
    ref = (back or "").strip() or (request.headers.get("referer") or "")
    if not ref or "://" in ref or not ref.startswith("/"):
        ref = f"/auditor/files?org_id={oa.org_id}"
    return _redirect(ref)


@router.get("/admin/org_artifacts/{org_artifact_id}/versions", response_class=HTMLResponse)
def admin_org_artifact_versions_page(
    org_artifact_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    back: str | None = None,
) -> HTMLResponse:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role not in (Role.admin, Role.auditor):
        raise HTTPException(status_code=403, detail="Требуются права auditor/admin")
    a = db.get(Artifact, oa.artifact_id)
    org = db.get(Organization, oa.org_id)
    if not a or not org:
        raise HTTPException(status_code=404, detail="Данные не найдены")

    CreatedBy = aliased(User)
    versions = (
        db.query(FileVersion, CreatedBy)
        .outerjoin(CreatedBy, CreatedBy.id == FileVersion.created_by_user_id)
        .filter(FileVersion.org_artifact_id == oa.id)
        .order_by(FileVersion.version_no.desc())
        .all()
    )
    rows = [
        {
            "id": fv.id,
            "version_no": fv.version_no,
            "original_filename": fv.original_filename,
            "created_at": fv.created_at,
            "created_by_login": created_by.login if created_by else "",
        }
        for (fv, created_by) in versions
    ]

    # safe back url: allow only local paths
    back_url = back or (request.headers.get("referer") or f"/auditor/files?org_id={oa.org_id}")
    if "://" in back_url:
        back_url = f"/auditor/files?org_id={oa.org_id}"

    resp = templates.TemplateResponse(
        "admin/org_artifact_versions.html",
        {"request": request, "user": user, "oa": oa, "a": a, "org": org, "versions": rows, "back_url": back_url},
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.post("/my/artifacts/{org_artifact_id}/upload")
def my_artifacts_upload(
    org_artifact_id: int,
    request: Request,
    upload: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role != Role.customer:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    content = upload.file.read()
    size_bytes = len(content)
    if size_bytes > settings.max_upload_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"Файл слишком большой. Лимит {settings.max_upload_mb} МБ")
    sha256 = hashlib.sha256(content).hexdigest()

    current_max = db.query(func.max(FileVersion.version_no)).filter(FileVersion.org_artifact_id == oa.id).scalar() or 0
    fv = FileVersion(
        org_artifact_id=oa.id,
        version_no=int(current_max) + 1,
        original_filename=upload.filename or "file",
        content_type=upload.content_type or "application/octet-stream",
        size_bytes=size_bytes,
        sha256=sha256,
        storage_backend="postgres",
        storage_key=None,
        blob=content,
        created_at=datetime.utcnow(),
        created_by_user_id=user.id,
    )
    db.add(fv)
    db.flush()

    before = {
        "status": oa.status.value,
        "current_file_version_id": oa.current_file_version_id,
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at.isoformat() if oa.audited_at else None,
        "audited_by_user_id": oa.audited_by_user_id,
        "review_status": getattr(oa, "review_status", None).value if getattr(oa, "review_status", None) else None,
    }
    oa.status = OrgArtifactStatus.uploaded
    oa.current_file_version_id = fv.id
    oa.updated_at = datetime.utcnow()
    oa.updated_by_user_id = user.id
    # New version => audit reset
    oa.audited_file_version_id = None
    oa.audited_at = None
    oa.audited_by_user_id = None
    oa.review_status = OrgArtifactReviewStatus.pending
    after = {
        "status": oa.status.value,
        "current_file_version_id": oa.current_file_version_id,
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at,
        "audited_by_user_id": oa.audited_by_user_id,
        "review_status": oa.review_status.value,
    }

    write_audit_log(
        db,
        actor=user,
        org_id=oa.org_id,
        action="upload",
        entity_type="org_artifact",
        entity_id=str(oa.id),
        before=before,
        after=after,
        request=request,
    )
    db.commit()
    return _redirect(f"/my/artifacts?org_id={oa.org_id}")


@router.get("/my/artifacts/{org_artifact_id}/download")
def my_artifacts_download(
    org_artifact_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    version: int | None = None,
    fv_id: int | None = None,
) -> Response:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role != Role.customer:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    qv = db.query(FileVersion).filter(FileVersion.org_artifact_id == oa.id)
    # Customer can download a specific version when viewing past periods.
    if fv_id is not None:
        fv = qv.filter(FileVersion.id == int(fv_id)).one_or_none()
    elif version is not None:
        fv = qv.filter(FileVersion.version_no == int(version)).one_or_none()
    else:
        fv = db.get(FileVersion, oa.current_file_version_id) if oa.current_file_version_id else qv.order_by(FileVersion.version_no.desc()).first()
    if not fv or not fv.blob:
        raise HTTPException(status_code=404, detail="Файл не найден")

    headers = {"Content-Disposition": _download_content_disposition(fv.original_filename, fallback_prefix="artifact")}
    return Response(content=fv.blob, media_type=fv.content_type, headers=headers)


@router.post("/my/artifacts/{org_artifact_id}/delete")
def my_artifacts_delete(
    org_artifact_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    oa = db.get(OrgArtifact, org_artifact_id)
    if not oa:
        raise HTTPException(status_code=404, detail="Артефакт организации не найден")
    role = get_user_role_for_org(db, user, oa.org_id)
    if role != Role.customer:
        raise HTTPException(status_code=403, detail="Недостаточно прав")

    before = {
        "status": oa.status.value,
        "current_file_version_id": oa.current_file_version_id,
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at.isoformat() if oa.audited_at else None,
        "audited_by_user_id": oa.audited_by_user_id,
        "review_status": getattr(oa, "review_status", None).value if getattr(oa, "review_status", None) else None,
    }
    oa.current_file_version_id = None
    oa.status = OrgArtifactStatus.missing
    oa.updated_at = datetime.utcnow()
    oa.updated_by_user_id = user.id
    oa.audited_file_version_id = None
    oa.audited_at = None
    oa.audited_by_user_id = None
    oa.review_status = OrgArtifactReviewStatus.pending
    after = {
        "status": oa.status.value,
        "current_file_version_id": oa.current_file_version_id,
        "audited_file_version_id": oa.audited_file_version_id,
        "audited_at": oa.audited_at,
        "audited_by_user_id": oa.audited_by_user_id,
        "review_status": oa.review_status.value,
    }

    write_audit_log(
        db,
        actor=user,
        org_id=oa.org_id,
        action="delete_file",
        entity_type="org_artifact",
        entity_id=str(oa.id),
        before=before,
        after=after,
        request=request,
    )
    db.commit()
    return _redirect(f"/my/artifacts?org_id={oa.org_id}")

def _require_admin_or_global_auditor(db: Session, user: User) -> None:
    if user.is_admin:
        return
    is_global_auditor = (
        db.query(UserOrgMembership)
        .filter(UserOrgMembership.user_id == user.id, UserOrgMembership.role == Role.auditor)
        .first()
        is not None
    )
    if not is_global_auditor:
        raise HTTPException(status_code=403, detail="Требуются права admin или auditor")


@router.get("/artifacts", response_class=HTMLResponse)
def artifacts_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    topic: str | None = None,
    domain: str | None = None,
    kb_level: str | None = None,
    short_name: str | None = None,
    q: str | None = None,
) -> HTMLResponse:
    # Справочник артефактов: только admin/auditor. Здесь ничего не загружаем.
    _require_admin_or_global_auditor(db, user)

    query = (
        db.query(Artifact)
        .order_by(Artifact.topic.asc(), Artifact.domain.asc(), Artifact.short_name.asc(), Artifact.achievement_item_no.asc().nullsfirst())
    )
    if topic:
        query = query.filter(Artifact.topic == topic)
    if domain:
        query = query.filter(Artifact.domain == domain)
    if kb_level:
        query = query.filter(Artifact.kb_level == kb_level)
    if short_name:
        query = query.filter(Artifact.short_name == short_name)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            (Artifact.indicator_name.ilike(like))
            | (Artifact.title.ilike(like))
            | (Artifact.achievement_text.ilike(like))
            | (Artifact.achievement_item_text.ilike(like))
        )

    artifacts = query.limit(2000).all()
    rows = artifacts

    # Сколько файлов "требуется" по short_name: если есть пункты 1./2., это количество строк.
    counts = dict(db.query(Artifact.short_name, func.count(Artifact.id)).group_by(Artifact.short_name).all())

    topics = [t for (t,) in db.query(Artifact.topic).filter(Artifact.topic != "").distinct().order_by(Artifact.topic.asc()).all()]
    domains = [d for (d,) in db.query(Artifact.domain).filter(Artifact.domain != "").distinct().order_by(Artifact.domain.asc()).all()]
    kb_levels = [k for (k,) in db.query(Artifact.kb_level).filter(Artifact.kb_level != "").distinct().order_by(Artifact.kb_level.asc()).all()]

    resp = templates.TemplateResponse(
        "artifacts.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "rows": rows,
            "topic": topic,
            "domain": domain,
            "kb_level": kb_level,
            "short_name": short_name,
            "q": q,
            "topics": topics,
            "domains": domains,
            "kb_levels": kb_levels,
            "required_counts": counts,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp

@router.get("/artifacts/{artifact_id}/edit", response_class=HTMLResponse)
def artifact_edit_page(
    artifact_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> HTMLResponse:
    _require_admin_or_global_auditor(db, user)
    a = db.get(Artifact, artifact_id)
    if not a:
        raise HTTPException(status_code=404, detail="Артефакт не найден")
    resp = templates.TemplateResponse(
        "artifact_edit.html",
        {"request": request, "user": user, "a": a, "error": None, "container_class": "container-wide"},
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.post("/artifacts/{artifact_id}/edit")
def artifact_edit_save(
    artifact_id: int,
    request: Request,
    topic: str = Form(""),
    domain: str = Form(""),
    kb_level: str = Form(""),
    indicator_name: str = Form(""),
    title: str = Form(""),
    achievement_text: str = Form(""),
    achievement_item_text: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    _require_admin_or_global_auditor(db, user)
    a = db.get(Artifact, artifact_id)
    if not a:
        raise HTTPException(status_code=404, detail="Артефакт не найден")

    before = {
        "topic": a.topic,
        "domain": a.domain,
        "kb_level": a.kb_level,
        "indicator_name": a.indicator_name,
        "title": a.title,
        "achievement_text": a.achievement_text,
        "achievement_item_text": a.achievement_item_text,
    }
    a.topic = (topic or "").strip()
    a.domain = (domain or "").strip()
    a.kb_level = (kb_level or "").strip()
    a.indicator_name = (indicator_name or "").strip()
    a.title = (title or "").strip()
    a.achievement_text = (achievement_text or "").strip()
    a.achievement_item_text = (achievement_item_text or "").strip()

    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="update",
        entity_type="artifact",
        entity_id=str(a.id),
        before=before,
        after={
            "topic": a.topic,
            "domain": a.domain,
            "kb_level": a.kb_level,
            "indicator_name": a.indicator_name,
            "title": a.title,
            "achievement_text": a.achievement_text,
            "achievement_item_text": a.achievement_item_text,
        },
        request=request,
    )
    db.commit()
    return _redirect("/artifacts")


@router.post("/files/upload")
def upload_file(
    org_id: int = Form(...),
    note: str = Form(""),
    upload: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    role = get_user_role_for_org(db, user, org_id)
    if not role:
        raise HTTPException(status_code=403, detail="Нет доступа к организации")

    content = upload.file.read()
    size_bytes = len(content)
    if size_bytes > settings.max_upload_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"Файл слишком большой. Лимит {settings.max_upload_mb} МБ")

    sha256 = hashlib.sha256(content).hexdigest()
    stored = StoredFile(
        org_id=org_id,
        original_filename=upload.filename or "file",
        content_type=upload.content_type or "application/octet-stream",
        size_bytes=size_bytes,
        sha256=sha256,
        blob=content,
        created_at=datetime.utcnow(),
        created_by_user_id=user.id,
        note=note or "",
    )
    db.add(stored)
    db.commit()

    return _redirect(f"/?org_id={org_id}")


@router.get("/files/{file_id}/download")
def download_file(
    file_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
) -> Response:
    stored = db.get(StoredFile, file_id)
    if not stored:
        raise HTTPException(status_code=404, detail="Файл не найден")
    role = get_user_role_for_org(db, user, stored.org_id)
    if not role:
        raise HTTPException(status_code=403, detail="Нет доступа к организации")

    headers = {"Content-Disposition": _download_content_disposition(stored.original_filename, fallback_prefix="file")}
    return Response(content=stored.blob, media_type=stored.content_type, headers=headers)


@router.get("/admin", response_class=HTMLResponse)
def admin_index(request: Request, user: User = Depends(require_admin)) -> HTMLResponse:
    return templates.TemplateResponse("admin/index.html", {"request": request, "user": user})


@router.get("/admin/artifact-levels", response_class=HTMLResponse)
def admin_artifact_levels_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
    err: str | None = None,
) -> HTMLResponse:
    levels = db.query(ArtifactLevel).order_by(ArtifactLevel.sort_order.asc(), ArtifactLevel.id.asc()).all()
    err_text = (err or "").strip() or None
    return templates.TemplateResponse(
        "admin/artifact_levels.html",
        {"request": request, "user": user, "container_class": "container-wide", "levels": levels, "error": err_text},
    )


@router.post("/admin/artifact-levels")
def admin_artifact_levels_create(
    request: Request,
    code: str = Form(...),
    name: str = Form(...),
    sort_order: int = Form(0),
    color: str = Form("#64748b"),
    is_active: str = Form("1"),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    code = (code or "").strip().upper()
    name = (name or "").strip()
    color = (color or "").strip() or "#64748b"
    if not code or not name:
        return _redirect("/admin/artifact-levels?err=" + quote("Код и название обязательны"))
    if len(code) > 32:
        return _redirect("/admin/artifact-levels?err=" + quote("Код слишком длинный (макс 32)"))
    active = (is_active or "").strip() not in ("0", "false", "False", "off")
    exists = db.query(ArtifactLevel).filter(ArtifactLevel.code == code).one_or_none()
    if exists:
        return _redirect("/admin/artifact-levels?err=" + quote("Уровень с таким кодом уже существует"))
    lvl = ArtifactLevel(code=code, name=name, sort_order=int(sort_order or 0), color=color[:32], is_active=active)
    db.add(lvl)
    db.flush()
    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="create",
        entity_type="artifact_level",
        entity_id=str(lvl.id),
        after={"code": lvl.code, "name": lvl.name, "sort_order": lvl.sort_order, "color": lvl.color, "is_active": lvl.is_active},
        request=request,
    )
    db.commit()
    return _redirect("/admin/artifact-levels")


@router.post("/admin/artifact-levels/{level_id}")
def admin_artifact_levels_update(
    level_id: int,
    request: Request,
    name: str = Form(...),
    sort_order: int = Form(0),
    color: str = Form("#64748b"),
    is_active: str = Form("1"),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    lvl = db.get(ArtifactLevel, level_id)
    if not lvl:
        raise HTTPException(status_code=404, detail="Уровень не найден")
    nm = (name or "").strip()
    if not nm:
        return _redirect("/admin/artifact-levels?err=" + quote("Название обязательно"))
    active = (is_active or "").strip() not in ("0", "false", "False", "off")
    before = {"name": lvl.name, "sort_order": lvl.sort_order, "color": lvl.color, "is_active": lvl.is_active}
    lvl.name = nm
    lvl.sort_order = int(sort_order or 0)
    lvl.color = (color or "").strip()[:32] or "#64748b"
    lvl.is_active = active
    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="update",
        entity_type="artifact_level",
        entity_id=str(lvl.id),
        before=before,
        after={"code": lvl.code, "name": lvl.name, "sort_order": lvl.sort_order, "color": lvl.color, "is_active": lvl.is_active},
        request=request,
    )
    db.commit()
    return _redirect("/admin/artifact-levels")


@router.get("/admin/artifact-levels/{level_id}", response_class=HTMLResponse)
def admin_artifact_level_edit_page(
    level_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
    q: str | None = None,
    topic: str | None = None,
    domain: str | None = None,
    kb_level: str | None = None,
    short_name: str | None = None,
) -> HTMLResponse:
    lvl = db.get(ArtifactLevel, level_id)
    if not lvl:
        raise HTTPException(status_code=404, detail="Уровень не найден")

    items = (
        db.query(ArtifactLevelItem, Artifact)
        .join(Artifact, Artifact.id == ArtifactLevelItem.artifact_id)
        .filter(ArtifactLevelItem.level_id == lvl.id)
        .order_by(Artifact.topic.asc(), Artifact.domain.asc(), Artifact.short_name.asc(), Artifact.achievement_item_no.asc().nullsfirst())
        .all()
    )
    current = [{"artifact": a, "item_id": it.id} for (it, a) in items]
    current_ids = {a.id for (_, a) in items}

    aq = db.query(Artifact).order_by(Artifact.topic.asc(), Artifact.domain.asc(), Artifact.short_name.asc(), Artifact.achievement_item_no.asc().nullsfirst())
    if topic:
        aq = aq.filter(Artifact.topic == topic)
    if domain:
        aq = aq.filter(Artifact.domain == domain)
    if kb_level:
        aq = aq.filter(Artifact.kb_level == kb_level)
    if short_name:
        aq = aq.filter(Artifact.short_name == short_name)
    if q:
        like = f"%{q.strip()}%"
        aq = aq.filter(
            (Artifact.indicator_name.ilike(like))
            | (Artifact.title.ilike(like))
            | (Artifact.achievement_text.ilike(like))
            | (Artifact.achievement_item_text.ilike(like))
        )
    candidates = aq.limit(250).all()

    topics = [t for (t,) in db.query(Artifact.topic).filter(Artifact.topic != "").distinct().order_by(Artifact.topic.asc()).all()]
    domains = [d for (d,) in db.query(Artifact.domain).filter(Artifact.domain != "").distinct().order_by(Artifact.domain.asc()).all()]
    kb_levels = [k for (k,) in db.query(Artifact.kb_level).filter(Artifact.kb_level != "").distinct().order_by(Artifact.kb_level.asc()).all()]

    return templates.TemplateResponse(
        "admin/artifact_level_edit.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "level": lvl,
            "current": current,
            "current_ids": current_ids,
            "candidates": candidates,
            "filters": {"q": q or "", "topic": topic or "", "domain": domain or "", "kb_level": kb_level or "", "short_name": short_name or ""},
            "topics": topics,
            "domains": domains,
            "kb_levels": kb_levels,
        },
    )


@router.post("/admin/artifact-levels/{level_id}/items/add")
def admin_artifact_level_items_add(
    level_id: int,
    request: Request,
    artifact_ids: list[int] = Form([]),
    back: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    lvl = db.get(ArtifactLevel, level_id)
    if not lvl:
        raise HTTPException(status_code=404, detail="Уровень не найден")
    ids = sorted({int(x) for x in (artifact_ids or []) if int(x) > 0})
    if not ids:
        return _redirect(back or f"/admin/artifact-levels/{lvl.id}")

    existing = {
        int(aid)
        for (aid,) in db.query(ArtifactLevelItem.artifact_id)
        .filter(ArtifactLevelItem.level_id == lvl.id, ArtifactLevelItem.artifact_id.in_(ids))
        .all()
    }
    created = 0
    for aid in ids:
        if aid in existing:
            continue
        db.add(ArtifactLevelItem(level_id=lvl.id, artifact_id=aid))
        created += 1
    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="add_items",
        entity_type="artifact_level",
        entity_id=str(lvl.id),
        after={"added_count": created, "artifact_ids": ids},
        request=request,
    )
    db.commit()
    ref = (back or "").strip() or f"/admin/artifact-levels/{lvl.id}"
    if not ref.startswith("/") or "://" in ref:
        ref = f"/admin/artifact-levels/{lvl.id}"
    return _redirect(ref)


@router.post("/admin/artifact-levels/{level_id}/items/{artifact_id}/delete")
def admin_artifact_level_items_delete(
    level_id: int,
    artifact_id: int,
    request: Request,
    back: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    lvl = db.get(ArtifactLevel, level_id)
    if not lvl:
        raise HTTPException(status_code=404, detail="Уровень не найден")
    it = db.query(ArtifactLevelItem).filter(ArtifactLevelItem.level_id == lvl.id, ArtifactLevelItem.artifact_id == artifact_id).one_or_none()
    if it:
        db.delete(it)
        write_audit_log(
            db,
            actor=user,
            org_id=None,
            action="remove_item",
            entity_type="artifact_level",
            entity_id=str(lvl.id),
            before={"artifact_id": int(artifact_id)},
            after=None,
            request=request,
        )
        db.commit()
    ref = (back or "").strip() or f"/admin/artifact-levels/{lvl.id}"
    if not ref.startswith("/") or "://" in ref:
        ref = f"/admin/artifact-levels/{lvl.id}"
    return _redirect(ref)


def _get_nextcloud_settings(db: Session) -> NextcloudIntegrationSettings:
    s = db.query(NextcloudIntegrationSettings).order_by(NextcloudIntegrationSettings.id.asc()).first()
    if not s:
        s = NextcloudIntegrationSettings()
        db.add(s)
        db.commit()
        db.refresh(s)
    return s


@router.get("/admin/integrations/nextcloud", response_class=HTMLResponse)
def admin_nextcloud_page(request: Request, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> HTMLResponse:
    s = _get_nextcloud_settings(db)
    ok = "Настройки сохранены." if request.query_params.get("saved") == "1" else None
    return templates.TemplateResponse(
        "admin/nextcloud.html",
        {"request": request, "user": user, "s": s, "error": None, "ok": ok, "discovered_orgs": None, "stats": None},
    )


@router.post("/admin/integrations/nextcloud/save")
def admin_nextcloud_save(
    request: Request,
    base_url: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    root_folder: str = Form(""),
    create_orgs: str = Form("true"),
    is_enabled: str = Form("false"),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    s = _get_nextcloud_settings(db)
    before = {"base_url": s.base_url, "username": s.username, "root_folder": s.root_folder, "create_orgs": s.create_orgs, "is_enabled": s.is_enabled}
    s.base_url = (base_url or "").strip()
    s.username = (username or "").strip()
    s.password = password or ""
    s.root_folder = (root_folder or "").strip().strip("/")
    s.create_orgs = str(create_orgs).lower() == "true"
    s.is_enabled = str(is_enabled).lower() == "true"
    s.last_error = ""
    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="update",
        entity_type="nextcloud_settings",
        entity_id=str(s.id or "1"),
        before=before,
        after={"base_url": s.base_url, "username": s.username, "root_folder": s.root_folder, "create_orgs": s.create_orgs, "is_enabled": s.is_enabled},
        request=request,
    )
    db.commit()
    return _redirect("/admin/integrations/nextcloud?saved=1")


@router.get("/admin/audit", response_class=HTMLResponse)
def admin_audit_log_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
    actor_user_id: int | None = None,
    org_id: int | None = None,
    action: str | None = None,
    entity_type: str | None = None,
    date_from: str | None = None,  # YYYY-MM-DD
    date_to: str | None = None,  # YYYY-MM-DD
    page: int = 1,
    page_size: int = 50,
) -> HTMLResponse:
    page = max(int(page or 1), 1)
    page_size = int(page_size or 50)
    if page_size < 10:
        page_size = 10
    if page_size > 200:
        page_size = 200

    Actor = aliased(User)
    Org = aliased(Organization)
    q = db.query(AuditLog, Actor.login, Org.name).outerjoin(Actor, Actor.id == AuditLog.actor_user_id).outerjoin(Org, Org.id == AuditLog.org_id)

    if actor_user_id:
        q = q.filter(AuditLog.actor_user_id == actor_user_id)
    if org_id:
        q = q.filter(AuditLog.org_id == org_id)
    if action:
        q = q.filter(AuditLog.action == action)
    if entity_type:
        q = q.filter(AuditLog.entity_type == entity_type)

    def parse_date(s: str | None) -> datetime | None:
        if not s:
            return None
        try:
            return datetime.strptime(s.strip(), "%Y-%m-%d")
        except Exception:
            return None

    d_from = parse_date(date_from)
    d_to = parse_date(date_to)
    if d_from:
        q = q.filter(AuditLog.at >= d_from)
    if d_to:
        # inclusive end date
        q = q.filter(AuditLog.at < (d_to + timedelta(days=1)))

    total = q.count()
    total_pages = max((total + page_size - 1) // page_size, 1)
    if page > total_pages:
        page = total_pages
    offset = (page - 1) * page_size

    rows = q.order_by(AuditLog.at.desc(), AuditLog.id.desc()).offset(offset).limit(page_size).all()

    def jdump(v: dict | None) -> str:
        if not v:
            return ""
        try:
            return json.dumps(v, ensure_ascii=False, indent=2, sort_keys=True)
        except Exception:
            return str(v)

    def _fmt_val(v: object) -> str:
        if v is None:
            return "—"
        try:
            if isinstance(v, (dict, list)):
                s = json.dumps(v, ensure_ascii=False, sort_keys=True)
            else:
                s = str(v)
        except Exception:
            s = str(v)
        s = s.replace("\r\n", "\n")
        if len(s) > 260:
            s = s[:257] + "…"
        return s

    def _diff_top_level(before: dict | None, after: dict | None) -> list[dict]:
        b = before or {}
        a = after or {}
        keys = sorted(set(b.keys()) | set(a.keys()))
        out: list[dict] = []
        for k in keys:
            in_b = k in b
            in_a = k in a
            if in_b and not in_a:
                out.append({"key": k, "kind": "removed", "before": _fmt_val(b.get(k)), "after": "—"})
                continue
            if in_a and not in_b:
                out.append({"key": k, "kind": "added", "before": "—", "after": _fmt_val(a.get(k))})
                continue
            vb = b.get(k)
            va = a.get(k)
            if vb != va:
                out.append({"key": k, "kind": "changed", "before": _fmt_val(vb), "after": _fmt_val(va)})
        return out

    items = []
    def _audit_status_from_state(state: dict | None) -> str:
        s = state or {}
        cur = s.get("current_file_version_id")
        aud = s.get("audited_file_version_id")
        review = (s.get("review_status") or "").strip().lower()
        if not cur:
            return "—"
        if review == "needs_correction":
            return "Требует корректировки"
        if not aud:
            return "Требует аудита"
        if review in ("", "approved") and aud == cur:
            return "Проаудировано"
        return "Изменён"

    for (log, actor_login, org_name) in rows:
        changes = _diff_top_level(log.before_json if isinstance(log.before_json, dict) else None, log.after_json if isinstance(log.after_json, dict) else None)
        audit_status = ""
        if log.entity_type == "org_artifact":
            state = log.after_json if isinstance(log.after_json, dict) else (log.before_json if isinstance(log.before_json, dict) else None)
            audit_status = _audit_status_from_state(state)
        items.append(
            {
                "id": log.id,
                "at": log.at,
                "actor_login": actor_login or "",
                "org_name": org_name or "",
                "action": log.action,
                "entity_type": log.entity_type,
                "entity_id": log.entity_id,
                "ip": log.ip,
                "user_agent": log.user_agent,
                "changes": changes,
                "before": jdump(log.before_json),
                "after": jdump(log.after_json),
                "audit_status": audit_status,
            }
        )

    # Russian labels for UI
    action_labels = {
        "create": "Создание",
        "update": "Изменение",
        "delete": "Удаление",
        "upload": "Загрузка файла",
        "delete_file": "Удаление файла",
        "comment": "Комментарий",
        "audit": "Проверено",
        "audit_needs_correction": "Требует корректировки",
        "add_items": "Добавить артефакты",
        "remove_item": "Убрать артефакт",
        "import_apply": "Импорт (применить)",
        "export_xlsx": "Экспорт XLSX",
        "org_artifacts_export_xlsx": "Экспорт Excel (артефакты)",
        "index_kb_export_uib_xlsx": "Экспорт Excel (Управление ИБ)",
        "index_kb_export_szi_xlsx": "Экспорт Excel (СЗИ)",
        "nextcloud_import": "Синхронизация Nextcloud",
        "nextcloud_import_v2": "Синхронизация Nextcloud (03 Артефакты)",
        "nextcloud_import_v1": "Синхронизация Nextcloud (v1)",
        "patch": "Правка",
    }
    entity_type_labels = {
        "org_artifact": "Артефакт организации",
        "artifact": "Артефакт (справочник)",
        "artifacts": "Справочник артефактов",
        "organization": "Организация",
        "user": "Пользователь",
        "membership": "Роль/доступ",
        "nextcloud_settings": "Настройки Nextcloud",
        "audit_period": "Период аудита",
        "artifact_level": "Уровень",
        "index_kb": "Индекс КБ",
        "org_artifacts": "Артефакты (выгрузка)",
        "org": "Организация",
    }

    # filter options
    users = db.query(User).order_by(User.login.asc()).all()
    orgs = _filter_out_default_orgs(db.query(Organization).order_by(Organization.name.asc()).all())
    actions = [a for (a,) in db.query(AuditLog.action).distinct().order_by(AuditLog.action.asc()).limit(300).all()]
    entity_types = [t for (t,) in db.query(AuditLog.entity_type).distinct().order_by(AuditLog.entity_type.asc()).limit(300).all()]

    base_qs = {
        "actor_user_id": str(actor_user_id or ""),
        "org_id": str(org_id or ""),
        "action": action or "",
        "entity_type": entity_type or "",
        "date_from": date_from or "",
        "date_to": date_to or "",
        "page_size": str(page_size),
    }
    base_query = urlencode({k: v for k, v in base_qs.items() if v})

    window = 3
    start = max(1, page - window)
    end = min(total_pages, page + window)
    page_links = list(range(start, end + 1))

    resp = templates.TemplateResponse(
        "admin/audit_log.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "items": items,
            "users": users,
            "orgs": orgs,
            "actions": actions,
            "entity_types": entity_types,
            "action_labels": action_labels,
            "entity_type_labels": entity_type_labels,
            "filters": {
                "actor_user_id": actor_user_id,
                "org_id": org_id,
                "action": action or "",
                "entity_type": entity_type or "",
                "date_from": date_from or "",
                "date_to": date_to or "",
            },
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "page_links": page_links,
            "base_query": base_query,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/admin/dashboards", response_class=HTMLResponse)
def admin_dashboards_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
    date_from: str | None = None,
    date_to: str | None = None,
    details_org_id: str | None = None,
) -> HTMLResponse:
    # admin wrapper around shared dashboards renderer
    return _dashboards_page_impl(
        request=request,
        db=db,
        user=user,
        date_from=date_from,
        date_to=date_to,
        details_org_id=details_org_id,
        action_path="/admin/dashboards",
        back_href="/admin",
        back_label="В админку",
    )


@router.get("/auditor/dashboards", response_class=HTMLResponse)
def auditor_dashboards_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
    date_from: str | None = None,
    date_to: str | None = None,
    details_org_id: str | None = None,
) -> HTMLResponse:
    _require_admin_or_global_auditor(db, user)
    return _dashboards_page_impl(
        request=request,
        db=db,
        user=user,
        date_from=date_from,
        date_to=date_to,
        details_org_id=details_org_id,
        action_path="/auditor/dashboards",
        back_href="/auditor/artifacts",
        back_label="К аудиту",
    )


def _dashboards_page_impl(
    *,
    request: Request,
    db: Session,
    user: User,
    date_from: str | None,
    date_to: str | None,
    details_org_id: str | None,
    action_path: str,
    back_href: str,
    back_label: str,
) -> HTMLResponse:
    # org selection: support multi-select via repeated query param org_ids=1&org_ids=2...
    raw_org_ids = list(request.query_params.getlist("org_ids"))
    if not raw_org_ids:
        # fallback: allow comma-separated org_ids=1,2,3
        raw = (request.query_params.get("org_ids") or "").strip()
        if raw:
            raw_org_ids = [p.strip() for p in raw.split(",") if p.strip()]

    sel: list[int] = []
    for s in raw_org_ids:
        try:
            v = int(str(s).strip())
        except Exception:
            continue
        if v > 0 and v not in sel:
            sel.append(v)

    # Cap to avoid huge server-side work
    if len(sel) > 50:
        sel = sel[:50]

    df, dt, p_start, p_end, range_err = _parse_date_range_bounds_utc(date_from, date_to)

    orgs = _filter_out_default_orgs(db.query(Organization).order_by(Organization.name.asc()).all())
    org_by_id = {int(o.id): o for o in orgs if o.id}
    selected_org_ids = [oid for oid in sel if oid in org_by_id]

    # selected chart (single)
    chart_key = (request.query_params.get("chart") or "").strip()
    allowed_chart_keys = {
        "overview_uib",
        "overview_szi",
        "uib_total",
        "szi_total",
        "uib_overall",
        "szi_overall",
        "uib_radar",
        "szi_radar",
        "uploads",
        "uploads_repeat",
        "statuses",
        "statuses_uib",
        "statuses_szi",
        "statuses_breakdown",
        "statuses_breakdown_uib",
        "statuses_breakdown_szi",
        "levels_statuses",
        "levels_statuses_uib",
        "levels_statuses_szi",
        "uib_section_statuses",
        "szi_section_statuses",
        "backlog_age",
        "backlog_age_uib",
        "backlog_age_szi",
        "backlog_sla_uib",
        "backlog_sla_szi",
    }
    if chart_key not in allowed_chart_keys:
        # UX: по умолчанию показываем обычный “Индекс по организациям”.
        # “Обзор” должен включаться только явным выбором кнопки.
        chart_key = "uib_overall"
    needs_details = chart_key in {
        "uib_radar",
        "szi_radar",
        "statuses_breakdown",
        "statuses_breakdown_uib",
        "statuses_breakdown_szi",
        "levels_statuses",
        "levels_statuses_uib",
        "levels_statuses_szi",
        "uib_section_statuses",
        "szi_section_statuses",
        "backlog_age",
        "backlog_age_uib",
        "backlog_age_szi",
        "backlog_sla_uib",
        "backlog_sla_szi",
    }

    # details org: parse safely from string (FastAPI иначе падает на details_org_id="")
    det_id: int | None = None
    if details_org_id is not None:
        s = str(details_org_id).strip()
        if s:
            try:
                det_id = int(s)
            except Exception:
                det_id = None
    if needs_details and selected_org_ids:
        if det_id is None:
            det_id = int(selected_org_ids[0])
        if det_id not in selected_org_ids:
            det_id = int(selected_org_ids[0])
    else:
        det_id = None
    details_org_name = org_by_id.get(int(det_id)).name if (det_id and int(det_id) in org_by_id) else ""

    # Audit period overlay selector (used on Files -> by days chart).
    audit_org_id: int | None = None
    s_audit = (request.query_params.get("audit_org_id") or "").strip()
    if s_audit:
        try:
            audit_org_id = int(s_audit)
        except Exception:
            audit_org_id = None
    if audit_org_id is not None and audit_org_id not in selected_org_ids:
        audit_org_id = None

    audit_periods_by_org: dict[int, dict[str, object]] = {}
    for oid in selected_org_ids:
        o = org_by_id.get(int(oid))
        if not o:
            continue
        start = getattr(o, "audit_period_start", None)
        weeks = getattr(o, "audit_period_weeks", None)
        if start and weeks:
            try:
                w = int(weeks)
            except Exception:
                continue
            if w <= 0:
                continue
            end = start + timedelta(days=w * 7)
            audit_periods_by_org[int(oid)] = {
                "org_id": int(oid),
                "org_name": o.name,
                "start": start.isoformat(),
                "end": end.isoformat(),
                "weeks": int(w),
            }

    if audit_org_id is None:
        # Prefer first org that has an audit period configured.
        if audit_periods_by_org:
            audit_org_id = sorted(audit_periods_by_org.keys())[0]
        elif selected_org_ids:
            audit_org_id = int(selected_org_ids[0])

    # Filters for breakdowns
    dim_topic = (request.query_params.get("dim_topic") or "").strip() in ("1", "true", "on", "yes")
    dim_domain = (request.query_params.get("dim_domain") or "").strip() in ("1", "true", "on", "yes")
    if not dim_topic and not dim_domain:
        dim_topic = True  # default

    filter_topics = [t.strip() for t in request.query_params.getlist("topic") if str(t).strip()]
    filter_domains = [d.strip() for d in request.query_params.getlist("domain") if str(d).strip()]
    # UX: "only_effective" убрали (лишняя сложность для дашбордов).
    try:
        top_n = int((request.query_params.get("top_n") or "12").strip() or "12")
    except Exception:
        top_n = 12
    if top_n < 5:
        top_n = 5
    if top_n > 30:
        top_n = 30

    # Files: repeated uploads threshold (versions count)
    try:
        repeat_min = int((request.query_params.get("repeat_min") or "4").strip() or "4")
    except Exception:
        repeat_min = 4
    if repeat_min < 2:
        repeat_min = 2
    if repeat_min > 50:
        repeat_min = 50

    def _mean(vals: list[float]) -> float | None:
        xs = [float(v) for v in vals if v is not None]  # type: ignore[comparison-overlap]
        if not xs:
            # Для дашборда лучше показывать "0", чем пустой график.
            return 0.0
        return float(sum(xs) / len(xs))

    def _overall_from_rows(rows: list[object]) -> dict[str, float | None]:
        kb1: list[float] = []
        kb2: list[float] = []
        kb3: list[float] = []
        for rv in rows:
            row = getattr(rv, "row", None)
            if not row or getattr(row, "kind", "") == "group":
                continue
            v1 = getattr(rv, "kb1", None)
            v2 = getattr(rv, "kb2", None)
            v3 = getattr(rv, "kb3", None)
            if v1 is not None:
                kb1.append(float(v1))
            if v2 is not None:
                kb2.append(float(v2))
            if v3 is not None:
                kb3.append(float(v3))
        return {"kb1": _mean(kb1), "kb2": _mean(kb2), "kb3": _mean(kb3)}

    error = range_err or ""

    chart_title = ""
    chart_subtitle = ""
    dash: dict = {"type": chart_key, "data": {}}

    # Charts require choosing at least one organization.
    if selected_org_ids and not error:
        def _sheet_name_from_chart_key(k: str) -> str:
            if k.startswith("szi_") or k.endswith("_szi"):
                return SZI_SHEET_NAME
            # default: UIB
            return UIB_SHEET_NAME

        def _artifact_ids_for_sheet(sheet_name: str):
            return (
                db.query(Artifact.id)
                .join(IndexKbTemplateRow, func.upper(IndexKbTemplateRow.short_name) == func.upper(Artifact.short_name))
                .filter(IndexKbTemplateRow.sheet_name == sheet_name, IndexKbTemplateRow.kind == "item", IndexKbTemplateRow.short_name != "")
                .distinct()
                .subquery()
            )

        def _level_id_by_code(code: str) -> int | None:
            lvl = db.query(ArtifactLevel).filter(ArtifactLevel.code == (code or "").strip().upper()).one_or_none()
            return int(lvl.id) if lvl else None

        def _effective_ids_for_level(code: str) -> list[int]:
            lvl_id = _level_id_by_code(code)
            return _get_effective_artifact_ids_for_level(db, level_id=lvl_id) if lvl_id else []

        def _audited_flag_expr(*, fv_alias) -> object:
            # Returns SQL expression 1/0 for "audited & approved".
            if p_start or p_end:
                # latest file version id in range per org_artifact
                fv_conds = []
                if p_start:
                    fv_conds.append(FileVersion.created_at >= p_start)
                if p_end:
                    fv_conds.append(FileVersion.created_at < p_end)
                sub = (
                    db.query(
                        FileVersion.org_artifact_id.label("oa_id"),
                        func.max(FileVersion.id).label("fv_id"),
                    )
                    .filter(*fv_conds)
                    .group_by(FileVersion.org_artifact_id)
                    .subquery()
                )
                aud_conds = []
                if p_start:
                    aud_conds.append(OrgArtifact.audited_at >= p_start)
                if p_end:
                    aud_conds.append(OrgArtifact.audited_at < p_end)
                return case(
                    (
                        and_(
                            sub.c.fv_id.isnot(None),
                            OrgArtifact.audited_file_version_id.isnot(None),
                            OrgArtifact.audited_file_version_id == sub.c.fv_id,
                            OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                            OrgArtifact.audited_at.isnot(None),
                            *aud_conds,
                        ),
                        1,
                    ),
                    else_=0,
                ), sub
            # current snapshot
            return (
                case(
                    (
                        and_(
                            OrgArtifact.current_file_version_id.isnot(None),
                            OrgArtifact.audited_file_version_id.isnot(None),
                            OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                            OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                        ),
                        1,
                    ),
                    else_=0,
                ),
                None,
            )

        def _section_means_for_allowed_ids(*, sheet_name: str, org_id: int, allowed_ids: list[int]) -> dict[str, float]:
            """
            Возвращает средний скор (0..5) по каждому group_code для указанного набора artifact_ids.
            Учитывает фильтры тематика/домен и диапазон дат.
            """
            q = (
                db.query(
                    IndexKbTemplateRow.group_code.label("gc"),
                    func.avg(
                        case(
                            (OrgArtifact.current_file_version_id.is_(None), 0.0),
                            else_=(case((sa.literal(1) == sa.literal(1), 0.0), else_=0.0)),
                        )
                    ).label("avg0"),
                )
                .join(Artifact, func.upper(Artifact.short_name) == func.upper(IndexKbTemplateRow.short_name))
                .join(OrgArtifact, and_(OrgArtifact.artifact_id == Artifact.id, OrgArtifact.org_id == int(org_id)))
                .filter(IndexKbTemplateRow.sheet_name == sheet_name, IndexKbTemplateRow.kind == "item")
            )
            if allowed_ids:
                q = q.filter(Artifact.id.in_(allowed_ids))
            else:
                # no allowed ids => no rows
                return {}
            if filter_topics:
                q = q.filter(Artifact.topic.in_(filter_topics))
            if filter_domains:
                q = q.filter(Artifact.domain.in_(filter_domains))

            aud_expr, sub = _audited_flag_expr(fv_alias=None)
            if sub is not None:
                q = q.outerjoin(sub, sub.c.oa_id == OrgArtifact.id)
            # score: 5 if audited else 0 (missing file already handled by current_file_version_id is null -> 0)
            score = case((aud_expr == 1, 5.0), else_=0.0)
            q = q.with_entities(IndexKbTemplateRow.group_code.label("gc"), func.avg(score).label("avg_score"))
            rows = q.group_by(IndexKbTemplateRow.group_code).all()
            out: dict[str, float] = {}
            for gc, avg_score in rows:
                if not gc:
                    continue
                out[str(gc).strip()] = float(avg_score or 0.0)
            return out

        if chart_key in ("uib_radar", "szi_radar") and det_id:
            sheet_name = "Управление ИБ" if chart_key == "uib_radar" else "СЗИ"
            chart_title = f"{sheet_name} · Радар (как в шаблоне) · {details_org_name}"
            chart_subtitle = "Оси — разделы шаблона. Линии: КБ3/КБ2/КБ1 + расчётный 2025 (и текущий показатель для УИБ)."

            # order + titles from template groups
            grp_rows = (
                db.query(IndexKbTemplateRow.short_name, IndexKbTemplateRow.title, IndexKbTemplateRow.sort_order, IndexKbTemplateRow.id)
                .filter(IndexKbTemplateRow.sheet_name == sheet_name, IndexKbTemplateRow.kind == "group")
                .order_by(IndexKbTemplateRow.sort_order.asc(), IndexKbTemplateRow.id.asc())
                .all()
            )
            group_order = [(str(sn or "").strip(), str(title or "").strip()) for (sn, title, _so, _id) in grp_rows if (sn or "").strip()]

            ids_l1 = _effective_ids_for_level("L1")
            ids_l2 = _effective_ids_for_level("L2")
            ids_l3 = _effective_ids_for_level("L3")
            org_obj = db.get(Organization, int(det_id))

            m1 = _section_means_for_allowed_ids(sheet_name=sheet_name, org_id=int(det_id), allowed_ids=ids_l1)
            m2 = _section_means_for_allowed_ids(sheet_name=sheet_name, org_id=int(det_id), allowed_ids=ids_l2)
            m3 = _section_means_for_allowed_ids(sheet_name=sheet_name, org_id=int(det_id), allowed_ids=ids_l3)

            labels = [code for (code, _title) in group_order]
            kb1 = [float(m1.get(code, 0.0)) for code in labels]
            kb2 = [float(m2.get(code, 0.0)) for code in labels]
            kb3 = [float(m3.get(code, 0.0)) for code in labels]
            calc_2025 = [float(m3.get(code, 0.0)) for code in labels]

            cur_series: list[float] = []
            lvl_code = ""
            if org_obj and getattr(org_obj, "artifact_level_id", None):
                lvl = db.get(ArtifactLevel, int(org_obj.artifact_level_id))
                lvl_code = (lvl.code or "").strip().upper() if lvl else ""
            if lvl_code == "L1":
                cur_series = kb1[:]
            elif lvl_code == "L2":
                cur_series = kb2[:]
            else:
                cur_series = kb3[:]

            datasets = [
                {"label": "КБ3", "data": kb3},
                {"label": "КБ2", "data": kb2},
                {"label": "КБ1", "data": kb1},
                {"label": "Расчетный показатель 2025", "data": calc_2025},
            ]
            if sheet_name == "Управление ИБ":
                datasets.append({"label": "Текущий показатель", "data": cur_series})

            dash["data"] = {
                "chart": "radar",
                "labels": labels,
                "datasets": datasets,
                "label_titles": {code: title for (code, title) in group_order},
            }

        if chart_key in ("overview_uib", "overview_szi"):
            if len(selected_org_ids) != 1:
                chart_key = "szi_overall" if chart_key == "overview_szi" else "uib_overall"
            else:
                oid = int(selected_org_ids[0])
                org_name = org_by_id[int(oid)].name
                sheet_name = SZI_SHEET_NAME if chart_key == "overview_szi" else UIB_SHEET_NAME

                chart_title = f"Обзор · {sheet_name} ({org_name})"
                chart_subtitle = "Все ключевые графики на одной странице. Кликните по карточке, чтобы открыть полную версию."

                # ---- Index overall (bar) ----
                if sheet_name == UIB_SHEET_NAME:
                    _org_uib, _tpl_uib, uib_rows = build_uib_view(db, org_id=int(oid), actor=user, range_start=p_start, range_end=p_end)
                    overall = _overall_from_rows(uib_rows)
                else:
                    _org_szi, _tpl_szi, szi_rows = build_szi_view(db, org_id=int(oid), actor=user, range_start=p_start, range_end=p_end)
                    overall = _overall_from_rows(szi_rows)

                tgt = (
                    db.query(OrgIndexKbTarget.target_value)
                    .filter(OrgIndexKbTarget.org_id == int(oid), OrgIndexKbTarget.sheet_name == sheet_name)
                    .scalar()
                )
                index_rows = [
                    {
                        "org_id": int(oid),
                        "org_name": org_name,
                        "target": float(tgt) if tgt is not None else None,
                        **overall,
                    }
                ]

                # ---- Statuses ----
                sheet_ids = _artifact_ids_for_sheet(sheet_name)
                st_row = (
                    db.query(
                        func.sum(case((OrgArtifact.current_file_version_id.is_(None), 1), else_=0)).label("missing"),
                        func.sum(case((OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, 1), else_=0)).label("needs_correction"),
                        func.sum(
                            case(
                                (
                                    and_(
                                        OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                                        OrgArtifact.audited_file_version_id.isnot(None),
                                        OrgArtifact.current_file_version_id.isnot(None),
                                        OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                                    ),
                                    1,
                                ),
                                else_=0,
                            )
                        ).label("approved"),
                        func.sum(
                            case(
                                (
                                    and_(
                                        OrgArtifact.current_file_version_id.isnot(None),
                                        OrgArtifact.audited_file_version_id.is_(None),
                                        OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                                    ),
                                    1,
                                ),
                                else_=0,
                            )
                        ).label("needs_audit"),
                        func.sum(
                            case(
                                (
                                    and_(
                                        OrgArtifact.current_file_version_id.isnot(None),
                                        OrgArtifact.audited_file_version_id.isnot(None),
                                        OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id,
                                    ),
                                    1,
                                ),
                                else_=0,
                            )
                        ).label("changed"),
                    )
                    .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
                    .filter(OrgArtifact.org_id == int(oid))
                    .filter(Artifact.id.in_(select(sheet_ids.c.id)))
                    .one()
                )
                statuses_payload = {
                    "labels": [org_name],
                    "datasets": [
                        {"label": "Проаудировано", "data": [int(getattr(st_row, "approved", 0) or 0)]},
                        {"label": "Требует аудита", "data": [int(getattr(st_row, "needs_audit", 0) or 0)]},
                        {"label": "Требует корректировки", "data": [int(getattr(st_row, "needs_correction", 0) or 0)]},
                        {"label": "Изменён (есть новая версия)", "data": [int(getattr(st_row, "changed", 0) or 0)]},
                        {"label": "Нет файла", "data": [int(getattr(st_row, "missing", 0) or 0)]},
                    ],
                    "indexAxis": "y",
                }

                # ---- Backlog (actionable only) ----
                now_dt = datetime.utcnow()
                fv = aliased(FileVersion)
                q = (
                    db.query(
                        case(
                            (OrgArtifact.current_file_version_id.is_(None), "нет файла"),
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    func.date_part("day", func.cast(now_dt, sa.DateTime()) - fv.created_at) <= 7,
                                ),
                                "0-7",
                            ),
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    func.date_part("day", func.cast(now_dt, sa.DateTime()) - fv.created_at) <= 14,
                                ),
                                "8-14",
                            ),
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    func.date_part("day", func.cast(now_dt, sa.DateTime()) - fv.created_at) <= 30,
                                ),
                                "15-30",
                            ),
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    func.date_part("day", func.cast(now_dt, sa.DateTime()) - fv.created_at) <= 60,
                                ),
                                "31-60",
                            ),
                            else_="61+",
                        ).label("bucket"),
                        case(
                            (OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, "требует корректировки"),
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.is_(None),
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                                ),
                                "требует аудита",
                            ),
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id,
                                ),
                                "изменён",
                            ),
                            (OrgArtifact.current_file_version_id.is_(None), "нет файла"),
                            else_="прочее",
                        ).label("st"),
                        func.count(OrgArtifact.id).label("cnt"),
                    )
                    .outerjoin(fv, fv.id == OrgArtifact.current_file_version_id)
                    .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
                    .filter(OrgArtifact.org_id == int(oid))
                    .filter(Artifact.id.in_(select(sheet_ids.c.id)))
                )
                if filter_topics:
                    q = q.filter(Artifact.topic.in_(filter_topics))
                if filter_domains:
                    q = q.filter(Artifact.domain.in_(filter_domains))
                q = q.filter(
                    or_(
                        OrgArtifact.current_file_version_id.is_(None),
                        OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction,
                        and_(OrgArtifact.current_file_version_id.isnot(None), OrgArtifact.audited_file_version_id.is_(None), OrgArtifact.review_status == OrgArtifactReviewStatus.pending),
                        and_(OrgArtifact.current_file_version_id.isnot(None), OrgArtifact.audited_file_version_id.isnot(None), OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id),
                    )
                )
                rows = q.group_by("bucket", "st").all()
                buckets = ["нет файла", "0-7", "8-14", "15-30", "31-60", "61+"]
                m: dict[tuple[str, str], int] = {}
                for b, st, cnt in rows:
                    m[(str(b), str(st))] = int(cnt or 0)
                backlog_payload = {
                    "labels": buckets,
                    "datasets": [
                        {"label": "Требует аудита", "data": [m.get((b, "требует аудита"), 0) for b in buckets]},
                        {"label": "Требует корректировки", "data": [m.get((b, "требует корректировки"), 0) for b in buckets]},
                        {"label": "Изменён", "data": [m.get((b, "изменён"), 0) for b in buckets]},
                        {"label": "Нет файла", "data": [m.get((b, "нет файла"), 0) for b in buckets]},
                    ],
                    "indexAxis": "y",
                }

                # ---- Uploads (single org) ----
                r_start = p_start or (datetime.utcnow() - timedelta(days=90))
                r_end = p_end or (datetime.utcnow() + timedelta(days=1))
                up_rows = (
                    db.query(func.date(FileVersion.created_at).label("d"), func.count(FileVersion.id))
                    .join(OrgArtifact, OrgArtifact.id == FileVersion.org_artifact_id)
                    .filter(OrgArtifact.org_id == int(oid))
                    .filter(FileVersion.created_at >= r_start, FileVersion.created_at < r_end)
                    .group_by(func.date(FileVersion.created_at))
                    .all()
                )
                start_day = r_start.date()
                end_day = (r_end - timedelta(days=1)).date()
                labels: list[str] = []
                cur = start_day
                while cur <= end_day:
                    labels.append(cur.isoformat())
                    cur = cur + timedelta(days=1)
                by_day = {str(d): int(c or 0) for d, c in up_rows if d is not None}
                uploads_payload = {
                    "labels": labels,
                    "series": [{"org_id": int(oid), "org_name": org_name, "data": [by_day.get(lbl, 0) for lbl in labels]}],
                }

                dash["data"] = {
                    "org_id": int(oid),
                    "org_name": org_name,
                    "sheet_name": sheet_name,
                    "index": {"rows": index_rows},
                    "statuses": statuses_payload,
                    "backlog": backlog_payload,
                    "uploads": uploads_payload,
                }

        elif chart_key == "uib_total":
            chart_title = "Управление ИБ · Суммарный индекс"
            chart_subtitle = "Агрегация по организациям: среднее значение КБ1/КБ2/КБ3. Если организации не выбраны — берём все."
            org_ids_for_calc = list(selected_org_ids)
            if not org_ids_for_calc:
                org_ids_for_calc = [int(o.id) for o in orgs if getattr(o, "id", None)]

            tgt_rows = (
                db.query(OrgIndexKbTarget.org_id, OrgIndexKbTarget.target_value)
                .filter(OrgIndexKbTarget.org_id.in_(org_ids_for_calc), OrgIndexKbTarget.sheet_name == UIB_SHEET_NAME)
                .all()
            )
            targets = [float(v) for (_oid, v) in tgt_rows if v is not None]
            target_avg = _mean(targets) if targets else None

            kb1_vals: list[float] = []
            kb2_vals: list[float] = []
            kb3_vals: list[float] = []
            for oid in org_ids_for_calc:
                _org_uib, _tpl_uib, uib_rows = build_uib_view(db, org_id=int(oid), actor=user, range_start=p_start, range_end=p_end)
                overall = _overall_from_rows(uib_rows)
                kb1_vals.append(float(overall.get("kb1") or 0.0))
                kb2_vals.append(float(overall.get("kb2") or 0.0))
                kb3_vals.append(float(overall.get("kb3") or 0.0))

            dash["data"] = {
                "org_count": int(len(org_ids_for_calc)),
                "kb1": _mean(kb1_vals),
                "kb2": _mean(kb2_vals),
                "kb3": _mean(kb3_vals),
                "dist": {"kb1": kb1_vals, "kb2": kb2_vals, "kb3": kb3_vals},
                "target": target_avg,
            }

        elif chart_key == "szi_total":
            chart_title = "СЗИ · Суммарный индекс"
            chart_subtitle = "Агрегация по организациям: среднее значение КБ1/КБ2/КБ3. Если организации не выбраны — берём все."
            org_ids_for_calc = list(selected_org_ids)
            if not org_ids_for_calc:
                org_ids_for_calc = [int(o.id) for o in orgs if getattr(o, "id", None)]

            tgt_rows = (
                db.query(OrgIndexKbTarget.org_id, OrgIndexKbTarget.target_value)
                .filter(OrgIndexKbTarget.org_id.in_(org_ids_for_calc), OrgIndexKbTarget.sheet_name == SZI_SHEET_NAME)
                .all()
            )
            targets = [float(v) for (_oid, v) in tgt_rows if v is not None]
            target_avg = _mean(targets) if targets else None

            kb1_vals: list[float] = []
            kb2_vals: list[float] = []
            kb3_vals: list[float] = []
            for oid in org_ids_for_calc:
                _org_szi, _tpl_szi, szi_rows = build_szi_view(db, org_id=int(oid), actor=user, range_start=p_start, range_end=p_end)
                overall = _overall_from_rows(szi_rows)
                kb1_vals.append(float(overall.get("kb1") or 0.0))
                kb2_vals.append(float(overall.get("kb2") or 0.0))
                kb3_vals.append(float(overall.get("kb3") or 0.0))

            dash["data"] = {
                "org_count": int(len(org_ids_for_calc)),
                "kb1": _mean(kb1_vals),
                "kb2": _mean(kb2_vals),
                "kb3": _mean(kb3_vals),
                "dist": {"kb1": kb1_vals, "kb2": kb2_vals, "kb3": kb3_vals},
                "target": target_avg,
            }

        elif chart_key == "uib_overall":
            chart_title = "Управление ИБ · Индекс по организациям"
            chart_subtitle = "Средние значения по требованиям (КБ1/КБ2/КБ3 = уровни L1/L2/L3)."
            tgt_rows = (
                db.query(OrgIndexKbTarget.org_id, OrgIndexKbTarget.target_value)
                .filter(OrgIndexKbTarget.org_id.in_(selected_org_ids), OrgIndexKbTarget.sheet_name == UIB_SHEET_NAME)
                .all()
            )
            targets = {int(oid): float(v) for (oid, v) in tgt_rows if oid is not None and v is not None}
            rows_out = []
            for oid in selected_org_ids:
                _org_uib, _tpl_uib, uib_rows = build_uib_view(db, org_id=int(oid), actor=user, range_start=p_start, range_end=p_end)
                rows_out.append(
                    {
                        "org_id": int(oid),
                        "org_name": org_by_id[int(oid)].name,
                        "target": targets.get(int(oid)),
                        **_overall_from_rows(uib_rows),
                    }
                )
            dash["data"] = {"rows": rows_out}

        elif chart_key == "szi_overall":
            chart_title = "СЗИ · Индекс по организациям"
            chart_subtitle = "Средние значения по требованиям (КБ1/КБ2/КБ3 = уровни L1/L2/L3)."
            tgt_rows = (
                db.query(OrgIndexKbTarget.org_id, OrgIndexKbTarget.target_value)
                .filter(OrgIndexKbTarget.org_id.in_(selected_org_ids), OrgIndexKbTarget.sheet_name == SZI_SHEET_NAME)
                .all()
            )
            targets = {int(oid): float(v) for (oid, v) in tgt_rows if oid is not None and v is not None}
            rows_out = []
            for oid in selected_org_ids:
                _org_szi, _tpl_szi, szi_rows = build_szi_view(db, org_id=int(oid), actor=user, range_start=p_start, range_end=p_end)
                rows_out.append(
                    {
                        "org_id": int(oid),
                        "org_name": org_by_id[int(oid)].name,
                        "target": targets.get(int(oid)),
                        **_overall_from_rows(szi_rows),
                    }
                )
            dash["data"] = {"rows": rows_out}

        elif chart_key == "uib_categories" and det_id:
            from app.index_kb.uib_sheet import compute_uib_summary

            chart_title = f"Управление ИБ · По категориям ({details_org_name})"
            chart_subtitle = "Средние значения по категориям. Пустые (—) не учитываются."
            _org_uib, _tpl_uib, uib_rows = build_uib_view(db, org_id=int(det_id), actor=user, range_start=p_start, range_end=p_end)
            uib_summary = compute_uib_summary(uib_rows)
            dash["data"] = {
                "categories": [
                    {"title": getattr(s, "title", ""), "short_name": getattr(s, "short_name", ""), "kb1": getattr(s, "kb1", None), "kb2": getattr(s, "kb2", None), "kb3": getattr(s, "kb3", None)}
                    for s in (uib_summary or [])
                ]
            }

        elif chart_key == "szi_categories" and det_id:
            from app.index_kb.szi_sheet import compute_szi_summary

            chart_title = f"СЗИ · По категориям ({details_org_name})"
            chart_subtitle = "Средние значения по категориям. Пустые (—) не учитываются."
            _org_szi, _tpl_szi, szi_rows = build_szi_view(db, org_id=int(det_id), actor=user, range_start=p_start, range_end=p_end)
            szi_summary = compute_szi_summary(szi_rows)
            dash["data"] = {
                "categories": [
                    {"title": getattr(s, "title", ""), "short_name": getattr(s, "short_name", ""), "kb1": getattr(s, "kb1", None), "kb2": getattr(s, "kb2", None), "kb3": getattr(s, "kb3", None)}
                    for s in (szi_summary or [])
                ]
            }

        elif chart_key == "uploads":
            chart_title = "Загруженные файлы · По дням"
            chart_subtitle = "Если диапазон дат пустой — показываем последние 90 дней."
            # if range not set: last 90 days for a useful chart
            r_start = p_start or (datetime.utcnow() - timedelta(days=90))
            r_end = p_end or (datetime.utcnow() + timedelta(days=1))
            # group by day
            rows = (
                db.query(OrgArtifact.org_id, func.date(FileVersion.created_at).label("d"), func.count(FileVersion.id))
                .join(FileVersion, FileVersion.org_artifact_id == OrgArtifact.id)
                .filter(OrgArtifact.org_id.in_(selected_org_ids))
                .filter(FileVersion.created_at >= r_start, FileVersion.created_at < r_end)
                .group_by(OrgArtifact.org_id, func.date(FileVersion.created_at))
                .all()
            )
            # build date labels
            start_day = r_start.date()
            end_day = (r_end - timedelta(days=1)).date()
            labels: list[str] = []
            cur = start_day
            while cur <= end_day:
                labels.append(cur.isoformat())
                cur = cur + timedelta(days=1)
            # map counts
            by_org_day: dict[int, dict[str, int]] = {int(oid): {} for oid in selected_org_ids}
            for oid, d, c in rows:
                if oid is None or d is None:
                    continue
                by_org_day[int(oid)][str(d)] = int(c or 0)
            series = []
            for oid in selected_org_ids:
                series.append({"org_id": int(oid), "org_name": org_by_id[int(oid)].name, "data": [by_org_day[int(oid)].get(lbl, 0) for lbl in labels]})
            ap = audit_periods_by_org.get(int(audit_org_id)) if audit_org_id else None
            dash["data"] = {"labels": labels, "series": series, "audit_period": ap}

        elif chart_key == "uploads_repeat":
            chart_title = "Файлы · Количество загрузок"
            if p_start or p_end:
                chart_subtitle = f"Артефакты, у которых за выбранный период было ≥ {repeat_min} загрузок (версий). Клик по столбцу — история версий."
            else:
                chart_subtitle = f"Артефакты, у которых за весь период было ≥ {repeat_min} загрузок (версий). Клик по столбцу — история версий."

            q = (
                db.query(
                    OrgArtifact.id.label("oa_id"),
                    OrgArtifact.org_id.label("org_id"),
                    Organization.name.label("org_name"),
                    Artifact.short_name.label("short_name"),
                    Artifact.title.label("title"),
                    func.count(FileVersion.id).label("cnt"),
                )
                .join(FileVersion, FileVersion.org_artifact_id == OrgArtifact.id)
                .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
                .join(Organization, Organization.id == OrgArtifact.org_id)
                .filter(OrgArtifact.org_id.in_(selected_org_ids))
            )
            if p_start:
                q = q.filter(FileVersion.created_at >= p_start)
            if p_end:
                q = q.filter(FileVersion.created_at < p_end)
            q = (
                q.group_by(OrgArtifact.id, OrgArtifact.org_id, Organization.name, Artifact.short_name, Artifact.title)
                .having(func.count(FileVersion.id) >= int(repeat_min))
                .order_by(func.count(FileVersion.id).desc(), OrgArtifact.id.asc())
                .limit(20)
            )
            rows = q.all()
            items = []
            for r in rows:
                oa_id = int(getattr(r, "oa_id", 0) or 0)
                org_name = str(getattr(r, "org_name", "") or "")
                short_name = str(getattr(r, "short_name", "") or "")
                title = str(getattr(r, "title", "") or "")
                cnt = int(getattr(r, "cnt", 0) or 0)
                items.append(
                    {
                        "oa_id": oa_id,
                        "org_id": int(getattr(r, "org_id", 0) or 0),
                        "org_name": org_name,
                        "artifact": short_name or title or f"OA {oa_id}",
                        "count": cnt,
                    }
                )
            dash["data"] = {"items": items, "threshold": int(repeat_min)}

        elif chart_key in ("statuses", "statuses_uib", "statuses_szi"):
            sheet_name = _sheet_name_from_chart_key(chart_key)
            chart_title = "Статусы артефактов · По организациям"
            chart_subtitle = f"Только требования листа «{sheet_name}». Включая «Нет файла». «Изменён» = есть новая версия после аудита."
            sheet_ids = _artifact_ids_for_sheet(sheet_name)
            sub = (
                db.query(
                    OrgArtifact.org_id.label("org_id"),
                    func.sum(case((OrgArtifact.current_file_version_id.is_(None), 1), else_=0)).label("missing"),
                    func.sum(case((OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, 1), else_=0)).label("needs_correction"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("approved"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.is_(None),
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("needs_audit"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("changed"),
                )
                .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
                .filter(OrgArtifact.org_id.in_(selected_org_ids))
                .filter(Artifact.id.in_(select(sheet_ids.c.id)))
                .group_by(OrgArtifact.org_id)
                .all()
            )
            labels = [org_by_id[int(oid)].name for oid in selected_org_ids]
            idx = {int(oid): i for i, oid in enumerate(selected_org_ids)}
            needs_audit = [0] * len(selected_org_ids)
            approved = [0] * len(selected_org_ids)
            needs_corr = [0] * len(selected_org_ids)
            changed = [0] * len(selected_org_ids)
            missing = [0] * len(selected_org_ids)
            for r in sub:
                oid = int(getattr(r, "org_id", 0) or 0)
                if oid not in idx:
                    continue
                i = idx[oid]
                needs_corr[i] = int(getattr(r, "needs_correction", 0) or 0)
                approved[i] = int(getattr(r, "approved", 0) or 0)
                needs_audit[i] = int(getattr(r, "needs_audit", 0) or 0)
                changed[i] = int(getattr(r, "changed", 0) or 0)
                missing[i] = int(getattr(r, "missing", 0) or 0)
            dash["statuses"] = {
                "labels": labels,
                "datasets": [
                    {"label": "Проаудировано", "data": approved},
                    {"label": "Требует аудита", "data": needs_audit},
                    {"label": "Требует корректировки", "data": needs_corr},
                    {"label": "Изменён (есть новая версия)", "data": changed},
                    {"label": "Нет файла", "data": missing},
                ],
            }
            dash["data"] = dash["statuses"]

        elif chart_key in ("statuses_breakdown", "statuses_breakdown_uib", "statuses_breakdown_szi") and det_id:
            sheet_name = _sheet_name_from_chart_key(chart_key)
            sheet_ids = _artifact_ids_for_sheet(sheet_name)
            # Group by topic/domain or both, with optional filters by topic/domain values.
            if dim_topic and dim_domain:
                chart_title = f"Статусы по тематикам+доменам ({details_org_name})"
                chart_subtitle = f"Лист «{sheet_name}». Топ‑{top_n} комбинаций по количеству артефактов."
                group_cols = [Artifact.topic.label("topic"), Artifact.domain.label("domain")]
                group_by_cols = [Artifact.topic, Artifact.domain]
            elif dim_domain:
                chart_title = f"Статусы по доменам ({details_org_name})"
                chart_subtitle = f"Лист «{sheet_name}». Топ‑{top_n} доменов по количеству артефактов."
                group_cols = [Artifact.domain.label("domain")]
                group_by_cols = [Artifact.domain]
            else:
                chart_title = f"Статусы по тематикам ({details_org_name})"
                chart_subtitle = f"Лист «{sheet_name}». Топ‑{top_n} тематик по количеству артефактов."
                group_cols = [Artifact.topic.label("topic")]
                group_by_cols = [Artifact.topic]

            base = (
                db.query(
                    *group_cols,
                    func.sum(case((OrgArtifact.current_file_version_id.is_(None), 1), else_=0)).label("missing"),
                    func.sum(case((OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, 1), else_=0)).label("needs_correction"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("approved"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.is_(None),
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("needs_audit"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("changed"),
                    func.count(OrgArtifact.id).label("total"),
                )
                .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
                .filter(OrgArtifact.org_id == int(det_id))
                .filter(Artifact.id.in_(select(sheet_ids.c.id)))
            )
            if filter_topics:
                base = base.filter(Artifact.topic.in_(filter_topics))
            if filter_domains:
                base = base.filter(Artifact.domain.in_(filter_domains))
            base = base.group_by(*group_by_cols).order_by(func.count(OrgArtifact.id).desc()).limit(int(top_n))
            rows = base.all()

            labels = []
            for r in rows:
                if dim_topic and dim_domain:
                    labels.append(f"{getattr(r, 'topic', '') or '—'} / {getattr(r, 'domain', '') or '—'}")
                elif dim_domain:
                    labels.append(str(getattr(r, "domain", "") or "—"))
                else:
                    labels.append(str(getattr(r, "topic", "") or "—"))

            dash["data"] = {
                "labels": labels,
                "datasets": [
                    {"label": "Проаудировано", "data": [int(getattr(r, "approved", 0) or 0) for r in rows]},
                    {"label": "Требует аудита", "data": [int(getattr(r, "needs_audit", 0) or 0) for r in rows]},
                    {"label": "Требует корректировки", "data": [int(getattr(r, "needs_correction", 0) or 0) for r in rows]},
                    {"label": "Изменён", "data": [int(getattr(r, "changed", 0) or 0) for r in rows]},
                    {"label": "Нет файла", "data": [int(getattr(r, "missing", 0) or 0) for r in rows]},
                ],
            }

        elif chart_key in ("levels_statuses", "levels_statuses_uib", "levels_statuses_szi") and det_id:
            sheet_name = _sheet_name_from_chart_key(chart_key)
            sheet_ids = _artifact_ids_for_sheet(sheet_name)
            chart_title = f"Статусы по уровням (КБ1/КБ2/КБ3) · {details_org_name}"
            chart_subtitle = f"Лист «{sheet_name}». Срез по полю артефакта kb_level (КБ1..КБ3)."
            q = (
                db.query(
                    Artifact.kb_level.label("kb_level"),
                    func.sum(case((OrgArtifact.current_file_version_id.is_(None), 1), else_=0)).label("missing"),
                    func.sum(case((OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, 1), else_=0)).label("needs_correction"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("approved"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.is_(None),
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("needs_audit"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("changed"),
                    func.count(OrgArtifact.id).label("total"),
                )
                .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
                .filter(OrgArtifact.org_id == int(det_id))
                .filter(Artifact.id.in_(select(sheet_ids.c.id)))
            )
            if filter_topics:
                q = q.filter(Artifact.topic.in_(filter_topics))
            if filter_domains:
                q = q.filter(Artifact.domain.in_(filter_domains))
            rows = q.group_by(Artifact.kb_level).order_by(func.count(OrgArtifact.id).desc()).all()
            def _kb_label(v: str) -> str:
                s = (v or "").strip().upper()
                if s == "КБ1":
                    return "КБ1 (L1)"
                if s == "КБ2":
                    return "КБ2 (L2)"
                if s == "КБ3":
                    return "КБ3 (L3)"
                return s or "—"
            labels = [_kb_label(getattr(r, "kb_level", "") or "") for r in rows]
            dash["data"] = {
                "labels": labels,
                "datasets": [
                    {"label": "Проаудировано", "data": [int(getattr(r, "approved", 0) or 0) for r in rows]},
                    {"label": "Требует аудита", "data": [int(getattr(r, "needs_audit", 0) or 0) for r in rows]},
                    {"label": "Требует корректировки", "data": [int(getattr(r, "needs_correction", 0) or 0) for r in rows]},
                    {"label": "Изменён", "data": [int(getattr(r, "changed", 0) or 0) for r in rows]},
                    {"label": "Нет файла", "data": [int(getattr(r, "missing", 0) or 0) for r in rows]},
                ],
            }

        elif chart_key in ("uib_section_statuses", "szi_section_statuses") and det_id:
            sheet_name = "Управление ИБ" if chart_key == "uib_section_statuses" else "СЗИ"
            chart_title = f"{sheet_name} · Разделы ({details_org_name})"
            chart_subtitle = "Статусы по разделам из шаблона (group_code). Включая «Нет файла»."

            # group meta (order + title)
            grp_rows = db.query(IndexKbTemplateRow.group_code, IndexKbTemplateRow.short_name, IndexKbTemplateRow.title, IndexKbTemplateRow.sort_order).filter(
                IndexKbTemplateRow.sheet_name == sheet_name, IndexKbTemplateRow.kind == "group"
            ).order_by(IndexKbTemplateRow.sort_order.asc(), IndexKbTemplateRow.id.asc()).all()
            meta: dict[str, dict] = {}
            for gc, sn, title, so in grp_rows:
                code = (sn or gc or "").strip()
                if not code:
                    continue
                meta[code] = {"title": str(title or code), "sort": int(so or 0)}

            # effective scope disabled for dashboards
            allowed_ids: list[int] | None = None

            q = (
                db.query(
                    IndexKbTemplateRow.group_code.label("gc"),
                    func.sum(case((OrgArtifact.current_file_version_id.is_(None), 1), else_=0)).label("missing"),
                    func.sum(case((OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, 1), else_=0)).label("needs_correction"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.approved,
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id == OrgArtifact.current_file_version_id,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("approved"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.is_(None),
                                    OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("needs_audit"),
                    func.sum(
                        case(
                            (
                                and_(
                                    OrgArtifact.current_file_version_id.isnot(None),
                                    OrgArtifact.audited_file_version_id.isnot(None),
                                    OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id,
                                ),
                                1,
                            ),
                            else_=0,
                        )
                    ).label("changed"),
                    func.count(OrgArtifact.id).label("total"),
                )
                .join(Artifact, func.upper(Artifact.short_name) == func.upper(IndexKbTemplateRow.short_name))
                .join(OrgArtifact, OrgArtifact.artifact_id == Artifact.id)
                .filter(IndexKbTemplateRow.sheet_name == sheet_name, IndexKbTemplateRow.kind == "item")
                .filter(OrgArtifact.org_id == int(det_id))
            )
            if filter_topics:
                q = q.filter(Artifact.topic.in_(filter_topics))
            if filter_domains:
                q = q.filter(Artifact.domain.in_(filter_domains))
            # (disabled) q = q.filter(Artifact.id.in_(allowed_ids))
            rows = q.group_by(IndexKbTemplateRow.group_code).all()

            def _label_for_gc(gc: str) -> tuple[int, str]:
                c = (gc or "").strip()
                if c in meta:
                    return int(meta[c]["sort"]), f"{meta[c]['title']} ({c})"
                return 9999, (c or "—")

            rows_sorted = sorted(rows, key=lambda r: _label_for_gc(str(getattr(r, "gc", "") or ""))[0])
            labels = [_label_for_gc(str(getattr(r, "gc", "") or ""))[1] for r in rows_sorted]
            dash["data"] = {
                "indexAxis": "y",
                "labels": labels,
                "datasets": [
                    {"label": "Проаудировано", "data": [int(getattr(r, "approved", 0) or 0) for r in rows_sorted]},
                    {"label": "Требует аудита", "data": [int(getattr(r, "needs_audit", 0) or 0) for r in rows_sorted]},
                    {"label": "Требует корректировки", "data": [int(getattr(r, "needs_correction", 0) or 0) for r in rows_sorted]},
                    {"label": "Изменён", "data": [int(getattr(r, "changed", 0) or 0) for r in rows_sorted]},
                    {"label": "Нет файла", "data": [int(getattr(r, "missing", 0) or 0) for r in rows_sorted]},
                ],
            }

        elif chart_key in ("backlog_age", "backlog_age_uib", "backlog_age_szi") and det_id:
            sheet_name = _sheet_name_from_chart_key(chart_key)
            sheet_ids = _artifact_ids_for_sheet(sheet_name)
            chart_title = f"Бэклог · Сколько дней без результата ({details_org_name})"
            chart_subtitle = f"Лист «{sheet_name}». Распределение по давности текущей версии файла для статусов, требующих действий."

            org_obj = db.get(Organization, int(det_id))
            allowed_ids: list[int] | None = None

            # bucketed counts
            now_dt = datetime.utcnow()
            fv = aliased(FileVersion)
            q = (
                db.query(
                    case(
                        (OrgArtifact.current_file_version_id.is_(None), "нет файла"),
                        (
                            and_(
                                OrgArtifact.current_file_version_id.isnot(None),
                                func.date_part("day", func.cast(now_dt, sa.DateTime()) - fv.created_at) <= 7,
                            ),
                            "0-7",
                        ),
                        (
                            and_(
                                OrgArtifact.current_file_version_id.isnot(None),
                                func.date_part("day", func.cast(now_dt, sa.DateTime()) - fv.created_at) <= 14,
                            ),
                            "8-14",
                        ),
                        (
                            and_(
                                OrgArtifact.current_file_version_id.isnot(None),
                                func.date_part("day", func.cast(now_dt, sa.DateTime()) - fv.created_at) <= 30,
                            ),
                            "15-30",
                        ),
                        (
                            and_(
                                OrgArtifact.current_file_version_id.isnot(None),
                                func.date_part("day", func.cast(now_dt, sa.DateTime()) - fv.created_at) <= 60,
                            ),
                            "31-60",
                        ),
                        else_="61+",
                    ).label("bucket"),
                    case(
                        (OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, "требует корректировки"),
                        (
                            and_(
                                OrgArtifact.current_file_version_id.isnot(None),
                                OrgArtifact.audited_file_version_id.is_(None),
                                OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                            ),
                            "требует аудита",
                        ),
                        (
                            and_(
                                OrgArtifact.current_file_version_id.isnot(None),
                                OrgArtifact.audited_file_version_id.isnot(None),
                                OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id,
                            ),
                            "изменён",
                        ),
                        (OrgArtifact.current_file_version_id.is_(None), "нет файла"),
                        else_="прочее",
                    ).label("st"),
                    func.count(OrgArtifact.id).label("cnt"),
                )
                .outerjoin(fv, fv.id == OrgArtifact.current_file_version_id)
                .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
                .filter(OrgArtifact.org_id == int(det_id))
                .filter(Artifact.id.in_(select(sheet_ids.c.id)))
            )
            if filter_topics:
                q = q.filter(Artifact.topic.in_(filter_topics))
            if filter_domains:
                q = q.filter(Artifact.domain.in_(filter_domains))
            # (disabled) q = q.filter(Artifact.id.in_(allowed_ids))
            # only actionable
            q = q.filter(
                or_(
                    OrgArtifact.current_file_version_id.is_(None),
                    OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction,
                    and_(OrgArtifact.current_file_version_id.isnot(None), OrgArtifact.audited_file_version_id.is_(None), OrgArtifact.review_status == OrgArtifactReviewStatus.pending),
                    and_(OrgArtifact.current_file_version_id.isnot(None), OrgArtifact.audited_file_version_id.isnot(None), OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id),
                )
            )
            rows = q.group_by("bucket", "st").all()
            buckets = ["нет файла", "0-7", "8-14", "15-30", "31-60", "61+"]
            statuses = ["требует аудита", "требует корректировки", "изменён", "нет файла"]
            # map
            m: dict[tuple[str, str], int] = {}
            for b, st, cnt in rows:
                m[(str(b), str(st))] = int(cnt or 0)
            dash["data"] = {
                "labels": buckets,
                "datasets": [
                    {"label": "Требует аудита", "data": [m.get((b, "требует аудита"), 0) for b in buckets]},
                    {"label": "Требует корректировки", "data": [m.get((b, "требует корректировки"), 0) for b in buckets]},
                    {"label": "Изменён", "data": [m.get((b, "изменён"), 0) for b in buckets]},
                    {"label": "Нет файла", "data": [m.get((b, "нет файла"), 0) for b in buckets]},
                ],
            }

        elif chart_key in ("backlog_sla_uib", "backlog_sla_szi") and det_id:
            sheet_name = _sheet_name_from_chart_key(chart_key)
            sheet_ids = _artifact_ids_for_sheet(sheet_name)
            chart_title = f"Бэклог · Сейчас в работе ({details_org_name})"
            chart_subtitle = (
                f"Лист «{sheet_name}». Только не финальные статусы (требует аудита/корректировки/изменён). "
                "Периоды: более 3 дней / более 2 недель / более месяца + «никогда не предоставлялся»."
            )

            fv = aliased(FileVersion)
            # Age meaning:
            # - needs_audit / changed: time since file was uploaded (fv.created_at)
            # - needs_correction: time since audit requested corrections (oa.audited_at)
            age_base_dt = case(
                (OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, func.coalesce(OrgArtifact.audited_at, fv.created_at)),
                else_=fv.created_at,
            )
            # Use DB time (now()) to avoid TZ/type issues.
            age_days = func.date_part("day", func.now() - age_base_dt)

            q = (
                db.query(
                    case(
                        (OrgArtifact.current_file_version_id.is_(None), "никогда не предоставлялся"),
                        (and_(OrgArtifact.current_file_version_id.isnot(None), age_days > 30), "более месяца"),
                        (and_(OrgArtifact.current_file_version_id.isnot(None), age_days > 14), "более 2 недель"),
                        else_="более 3 дней",
                    ).label("bucket"),
                    case(
                        (OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction, "требует корректировки"),
                        (
                            and_(
                                OrgArtifact.current_file_version_id.isnot(None),
                                OrgArtifact.audited_file_version_id.is_(None),
                                OrgArtifact.review_status == OrgArtifactReviewStatus.pending,
                            ),
                            "требует аудита",
                        ),
                        (
                            and_(
                                OrgArtifact.current_file_version_id.isnot(None),
                                OrgArtifact.audited_file_version_id.isnot(None),
                                OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id,
                            ),
                            "изменён",
                        ),
                        (OrgArtifact.current_file_version_id.is_(None), "никогда не предоставлялся"),
                        else_="прочее",
                    ).label("st"),
                    func.count(OrgArtifact.id).label("cnt"),
                )
                .outerjoin(fv, fv.id == OrgArtifact.current_file_version_id)
                .join(Artifact, Artifact.id == OrgArtifact.artifact_id)
                .filter(OrgArtifact.org_id == int(det_id))
                .filter(Artifact.id.in_(select(sheet_ids.c.id)))
            )
            if filter_topics:
                q = q.filter(Artifact.topic.in_(filter_topics))
            if filter_domains:
                q = q.filter(Artifact.domain.in_(filter_domains))

            # Only actionable and only older than 3 days (or never provided).
            q = q.filter(
                or_(
                    OrgArtifact.current_file_version_id.is_(None),
                    and_(OrgArtifact.current_file_version_id.isnot(None), age_days > 3),
                )
            ).filter(
                or_(
                    OrgArtifact.current_file_version_id.is_(None),
                    OrgArtifact.review_status == OrgArtifactReviewStatus.needs_correction,
                    and_(OrgArtifact.current_file_version_id.isnot(None), OrgArtifact.audited_file_version_id.is_(None), OrgArtifact.review_status == OrgArtifactReviewStatus.pending),
                    and_(OrgArtifact.current_file_version_id.isnot(None), OrgArtifact.audited_file_version_id.isnot(None), OrgArtifact.current_file_version_id != OrgArtifact.audited_file_version_id),
                )
            )

            rows = q.group_by("bucket", "st").all()
            buckets = ["никогда не предоставлялся", "более 3 дней", "более 2 недель", "более месяца"]
            m: dict[tuple[str, str], int] = {}
            for b, st, cnt in rows:
                m[(str(b), str(st))] = int(cnt or 0)
            dash["data"] = {
                "indexAxis": "y",
                "labels": buckets,
                "datasets": [
                    {"label": "Требует аудита", "data": [m.get((b, "требует аудита"), 0) for b in buckets]},
                    {"label": "Требует корректировки", "data": [m.get((b, "требует корректировки"), 0) for b in buckets]},
                    {"label": "Изменён", "data": [m.get((b, "изменён"), 0) for b in buckets]},
                    {"label": "Никогда не предоставлялся", "data": [m.get((b, "никогда не предоставлялся"), 0) for b in buckets]},
                ],
            }

    resp = templates.TemplateResponse(
        "admin/dashboards.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "dashboards_action": action_path,
            "dashboards_back_href": back_href,
            "dashboards_back_label": back_label,
            "orgs": orgs,
            "selected_org_ids": selected_org_ids,
            "details_org_id": int(det_id) if det_id else None,
            "details_org_name": details_org_name or "",
            "date_from": (df.isoformat() if df else ""),
            "date_to": (dt.isoformat() if dt else ""),
            "error": error,
            "dash": dash,
            "selected_chart": chart_key,
            "chart_title": chart_title,
            "chart_subtitle": chart_subtitle,
            "dim_topic": dim_topic,
            "dim_domain": dim_domain,
            "topics": [t for (t,) in db.query(Artifact.topic).filter(Artifact.topic != "").distinct().order_by(Artifact.topic.asc()).limit(200).all() if t],
            "domains": [d for (d,) in db.query(Artifact.domain).filter(Artifact.domain != "").distinct().order_by(Artifact.domain.asc()).limit(200).all() if d],
            "filter_topics": filter_topics,
            "filter_domains": filter_domains,
            "top_n": int(top_n),
            "repeat_min": int(repeat_min),
            "audit_org_id": int(audit_org_id) if audit_org_id else None,
            "audit_periods_by_org": audit_periods_by_org,
            # "only_effective" removed
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


def _dav_from_settings(s: NextcloudIntegrationSettings) -> NextcloudDavClient:
    webdav_base = build_webdav_base_url(s.base_url, s.username)
    return NextcloudDavClient(base_webdav_url=webdav_base, username=s.username, password=s.password)


@router.post("/admin/integrations/nextcloud/test", response_class=HTMLResponse)
def admin_nextcloud_test(request: Request, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> HTMLResponse:
    s = _get_nextcloud_settings(db)
    try:
        dav = _dav_from_settings(s)
        dav.propfind(s.root_folder, depth=1)
        s.last_error = ""
        db.commit()
        ok = "Подключение успешно. WebDAV доступен."
        err = None
    except Exception as e:
        s.last_error = str(e)
        db.commit()
        ok = None
        err = f"Ошибка подключения: {e}"
    return templates.TemplateResponse(
        "admin/nextcloud.html",
        {"request": request, "user": user, "s": s, "error": err, "ok": ok, "discovered_orgs": None, "stats": None},
    )


@router.post("/admin/integrations/nextcloud/discover", response_class=HTMLResponse)
def admin_nextcloud_discover(request: Request, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> HTMLResponse:
    s = _get_nextcloud_settings(db)
    try:
        dav = _dav_from_settings(s)
        items = dav.propfind(s.root_folder, depth=1)
        orgs = sorted({x.name for x in items if x.is_dir and x.name})
        s.last_error = ""
        db.commit()
        return templates.TemplateResponse(
            "admin/nextcloud.html",
            {"request": request, "user": user, "s": s, "error": None, "ok": f"Найдено папок организаций: {len(orgs)}", "discovered_orgs": orgs, "stats": None},
        )
    except Exception as e:
        s.last_error = str(e)
        db.commit()
        return templates.TemplateResponse(
            "admin/nextcloud.html",
            {"request": request, "user": user, "s": s, "error": f"Ошибка: {e}", "ok": None, "discovered_orgs": None, "stats": None},
        )


@router.post("/admin/integrations/nextcloud/sync", response_class=HTMLResponse)
def admin_nextcloud_sync(request: Request, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> HTMLResponse:
    s = _get_nextcloud_settings(db)
    if not s.is_enabled:
        return templates.TemplateResponse(
            "admin/nextcloud.html",
            {"request": request, "user": user, "s": s, "error": "Интеграция выключена (включите и сохраните настройки).", "ok": None, "discovered_orgs": None, "stats": None},
            status_code=400,
        )
    try:
        dav = _dav_from_settings(s)
        stats = sync_from_nextcloud(
            db=db,
            actor=user,
            dav=dav,
            root_folder=s.root_folder,
            create_orgs=s.create_orgs,
            request=request,
        )
        s.last_sync_at = datetime.utcnow()
        s.last_error = ""
        db.commit()
        return templates.TemplateResponse(
            "admin/nextcloud.html",
            {"request": request, "user": user, "s": s, "error": None, "ok": "Синхронизация завершена.", "discovered_orgs": None, "stats": stats},
        )
    except Exception as e:
        s.last_error = str(e)
        db.commit()
        return templates.TemplateResponse(
            "admin/nextcloud.html",
            {"request": request, "user": user, "s": s, "error": f"Ошибка синхронизации: {e}", "ok": None, "discovered_orgs": None, "stats": None},
            status_code=500,
        )


@router.post("/admin/integrations/nextcloud/sync-v2", response_class=HTMLResponse)
def admin_nextcloud_sync_v2(request: Request, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> HTMLResponse:
    """
    New sync (V2): ROOT/Org/03 Артефакты/... (as in docs/03 Артефакты.zip)
    Old sync remains available.
    """
    s = _get_nextcloud_settings(db)
    if not s.is_enabled:
        return templates.TemplateResponse(
            "admin/nextcloud.html",
            {"request": request, "user": user, "s": s, "error": "Интеграция выключена (включите и сохраните настройки).", "ok": None, "discovered_orgs": None, "stats": None},
            status_code=400,
        )
    try:
        dav = _dav_from_settings(s)
        from app.integrations.nextcloud_sync import sync_from_nextcloud_v2

        stats = sync_from_nextcloud_v2(
            db=db,
            actor=user,
            dav=dav,
            root_folder=s.root_folder,
            create_orgs=s.create_orgs,
            request=request,
        )
        s.last_sync_at = datetime.utcnow()
        s.last_error = ""
        db.commit()
        return templates.TemplateResponse(
            "admin/nextcloud.html",
            {"request": request, "user": user, "s": s, "error": None, "ok": "Синхронизация (V2) завершена.", "discovered_orgs": None, "stats": stats},
        )
    except Exception as e:
        s.last_error = str(e)
        db.commit()
        return templates.TemplateResponse(
            "admin/nextcloud.html",
            {"request": request, "user": user, "s": s, "error": f"Ошибка синхронизации (V2): {e}", "ok": None, "discovered_orgs": None, "stats": None},
            status_code=500,
        )


@router.get("/admin/artifacts", response_class=HTMLResponse)
def admin_artifacts(request: Request, user: User = Depends(require_admin)) -> HTMLResponse:
    return templates.TemplateResponse("admin/artifacts.html", {"request": request, "user": user, "result": None, "error": None})


def _artifact_key(short_name: str, item_no: int | None) -> str:
    return f"{short_name}#{item_no}" if item_no is not None else short_name


def _resolve_artifact_key_and_segment(
    db: Session,
    *,
    seen_base_keys: Dict[str, int],
    short_name: str,
    item_no: int | None,
    indicator_name: str,
    achievement_item_text: str,
    topic: str,
    domain: str,
    kb_level: str,
) -> Tuple[str, str | None]:
    """
    Возвращает (artifact_key, extra_segment_for_node_path).
    - Если item_no задан (есть перечисление 1./2.) — ключ стабилен short_name#item_no и segment=item_no.
    - Если item_no нет — ключ по умолчанию short_name. Но если в БД уже есть другой артефакт с тем же short_name,
      создаём стабильный ключ short_name~<hash8> и кладём hash8 как последний сегмент пути, чтобы не конфликтовать по node_id.
    """
    short_name = (short_name or "").strip()
    base_key = _artifact_key(short_name, item_no)
    if item_no is not None:
        return base_key, str(item_no)

    # В одном Excel иногда встречаются дубли short_name без нумерации 1./2.
    # Тогда делаем стабильный под-ключ (short_name~hash8) и добавляем hash8 как сегмент пути.
    seen_base_keys[base_key] = seen_base_keys.get(base_key, 0) + 1
    if seen_base_keys[base_key] > 1:
        stable = f"{topic}|{domain}|{short_name}|{kb_level}|{indicator_name}|{achievement_item_text}"
        h = hashlib.sha256(stable.encode("utf-8")).hexdigest()[:8]
        return f"{short_name}~{h}", h

    existing = db.query(Artifact).filter(Artifact.artifact_key == base_key).one_or_none()
    if not existing:
        return base_key, None

    # Если это тот же самый артефакт (идемпотентный импорт) — используем базовый ключ.
    if (existing.indicator_name or "") == (indicator_name or "") and (existing.achievement_item_text or "") == (achievement_item_text or ""):
        return base_key, None

    stable = f"{topic}|{domain}|{short_name}|{kb_level}|{indicator_name}|{achievement_item_text}"
    h = hashlib.sha256(stable.encode("utf-8")).hexdigest()[:8]
    key = f"{short_name}~{h}"
    return key, h


def _ensure_node_path(db: Session, segments: list[str]) -> ArtifactNode:
    parent_id: int | None = None
    full_path_parts: list[str] = []
    node: ArtifactNode | None = None
    for seg in segments:
        seg = (seg or "").strip()
        if not seg:
            continue
        full_path_parts.append(seg)
        full_path = ".".join(full_path_parts)
        node = db.query(ArtifactNode).filter(ArtifactNode.full_path == full_path).one_or_none()
        if not node:
            node = ArtifactNode(parent_id=parent_id, segment=seg, full_path=full_path, sort_order=0, created_at=datetime.utcnow())
            db.add(node)
            db.flush()
        parent_id = node.id
    if not node:
        raise ValueError("Пустой путь узлов")
    return node


@router.post("/admin/artifacts/import", response_class=HTMLResponse)
def admin_artifacts_import_apply(
    request: Request,
    upload: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> HTMLResponse:
    content = upload.file.read()
    sha = hashlib.sha256(content).hexdigest()
    try:
        rows = parse_program_xlsx(content)
    except Exception as e:
        return templates.TemplateResponse("admin/artifacts.html", {"request": request, "user": user, "result": None, "error": str(e)}, status_code=400)

    created = 0
    updated = 0
    seen_base_keys: Dict[str, int] = {}
    try:
        for r in rows:
            key, extra_segment = _resolve_artifact_key_and_segment(
                db,
                seen_base_keys=seen_base_keys,
                short_name=r.short_name,
                item_no=r.achievement_item_no,
                indicator_name=r.indicator_name,
                achievement_item_text=r.achievement_item_text,
                topic=r.topic,
                domain=r.domain,
                kb_level=r.kb_level,
            )
            short_parts = [p.strip() for p in (r.short_name or "").split(".") if p.strip()]
            segments = [r.topic or "", r.domain or "", *short_parts]
            if extra_segment:
                segments.append(extra_segment)
            leaf = _ensure_node_path(db, segments)

            a = db.query(Artifact).filter(Artifact.artifact_key == key).one_or_none()
            payload = {
                "node_id": leaf.id,
                "artifact_key": key,
                "topic": r.topic,
                "domain": r.domain,
                "indicator_name": r.indicator_name,
                "short_name": r.short_name,
                "kb_level": r.kb_level,
                "achievement_text": r.achievement_text,
                "achievement_item_no": r.achievement_item_no,
                "achievement_item_text": r.achievement_item_text,
                "title": r.achievement_item_text,
            }
            if not a:
                a = Artifact(**payload, description="", created_at=datetime.utcnow())
                db.add(a)
                db.flush()
                created += 1
            else:
                before = {
                    "topic": a.topic,
                    "domain": a.domain,
                    "indicator_name": a.indicator_name,
                    "kb_level": a.kb_level,
                    "achievement_item_text": a.achievement_item_text,
                }
                for k, v in payload.items():
                    setattr(a, k, v)
                db.flush()
                updated += 1
                write_audit_log(
                    db,
                    actor=user,
                    org_id=None,
                    action="update",
                    entity_type="artifact",
                    entity_id=str(a.id),
                    before=before,
                    after={k: payload[k] for k in before.keys()},
                    request=request,
                )
    except IntegrityError as e:
        db.rollback()
        msg = str(e.orig) if getattr(e, "orig", None) else str(e)
        return templates.TemplateResponse(
            "admin/artifacts.html",
            {"request": request, "user": user, "result": None, "error": f"Ошибка целостности БД при импорте: {msg}"},
            status_code=400,
        )

    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="import_apply",
        entity_type="artifacts",
        entity_id=sha,
        after={"created": created, "updated": updated, "filename": upload.filename or ""},
        request=request,
    )
    db.commit()

    result = {"sha256": sha, "created": created, "updated": updated}
    return templates.TemplateResponse("admin/artifacts.html", {"request": request, "user": user, "result": result, "error": None})


@router.get("/admin/orgs", response_class=HTMLResponse)
def admin_orgs(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
    q: str | None = None,
    err: str | None = None,
    page: int = 1,
    page_size: int = 20,
    sort: str = "created_at",
    dir: str = "desc",
) -> HTMLResponse:
    page = max(int(page or 1), 1)
    page_size = int(page_size or 20)
    if page_size < 10:
        page_size = 10
    if page_size > 200:
        page_size = 200

    sort_key = (sort or "created_at").strip().lower()
    sort_dir = (dir or "desc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"

    # For "created_by" sorting we need the creator login; use an outer join to avoid N+1.
    creator_login = func.coalesce(User.login, Organization.created_via)
    members_cnt_sub = (
        db.query(UserOrgMembership.org_id.label("org_id"), func.count(UserOrgMembership.id).label("members_cnt"))
        .group_by(UserOrgMembership.org_id)
        .subquery()
    )
    base_q = (
        db.query(
            Organization,
            creator_login.label("creator_login"),
            func.coalesce(members_cnt_sub.c.members_cnt, 0).label("members_cnt"),
        )
        .outerjoin(User, User.id == Organization.created_by_user_id)
        .outerjoin(members_cnt_sub, members_cnt_sub.c.org_id == Organization.id)
    )

    q_text = (q or "").strip()
    if q_text:
        base_q = base_q.filter(Organization.name.ilike(f"%{q_text}%"))

    if sort_key == "name":
        order_expr = Organization.name.asc() if sort_dir == "asc" else Organization.name.desc()
        base_q = base_q.order_by(order_expr, Organization.id.desc())
    elif sort_key == "created_by":
        order_expr = creator_login.asc() if sort_dir == "asc" else creator_login.desc()
        base_q = base_q.order_by(order_expr, Organization.created_at.desc(), Organization.id.desc())
    else:
        # created_at
        order_expr = Organization.created_at.asc() if sort_dir == "asc" else Organization.created_at.desc()
        base_q = base_q.order_by(order_expr, Organization.id.desc())

    total = base_q.count()

    total_pages = max((total + page_size - 1) // page_size, 1)
    if page > total_pages:
        page = total_pages
    offset = (page - 1) * page_size

    rows = base_q.offset(offset).limit(page_size).all()

    org_rows = []
    for (o, creator_login_val, members_cnt_val) in rows:
        if o.created_by_user_id and creator_login_val and str(creator_login_val) not in ("system", "nextcloud", "manual"):
            created_by_label = str(creator_login_val)
        else:
            via = getattr(o, "created_via", "") or ""
            if via == "nextcloud":
                created_by_label = "Синхронизация (Nextcloud)"
            elif via == "system":
                created_by_label = "Система"
            else:
                created_by_label = "—"
        org_rows.append({"org": o, "created_by_label": created_by_label, "members_cnt": int(members_cnt_val or 0)})

    base_qd: dict[str, str] = {"page_size": str(page_size), "sort": sort_key, "dir": sort_dir}
    if q_text:
        base_qd["q"] = q_text
    base_query = urlencode(base_qd)
    window = 3
    start = max(1, page - window)
    end = min(total_pages, page + window)
    page_links = list(range(start, end + 1))

    err_text = (err or "").strip()
    if err_text:
        # минимальная санитация: убираем переносы/ограничиваем размер
        err_text = " ".join(err_text.split())
        err_text = err_text[:240]
    else:
        err_text = None

    resp = templates.TemplateResponse(
        "admin/orgs.html",
        {
            "request": request,
            "user": user,
            "orgs": org_rows,
            "error": err_text,
            "filters": {"q": q_text},
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "has_prev": page > 1,
            "has_next": offset + page_size < total,
            "page_links": page_links,
            "base_query": base_query,
            "sort": sort_key,
            "dir": sort_dir,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/admin/orgs/new", response_class=HTMLResponse)
def admin_orgs_create_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> HTMLResponse:
    return templates.TemplateResponse(
        "admin/org_create.html",
        {"request": request, "user": user, "error": None, "form": {"name": ""}},
    )


@router.post("/admin/orgs")
def admin_orgs_create(
    request: Request,
    name: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    name = name.strip()
    if not name:
        return templates.TemplateResponse(
            "admin/org_create.html",
            {"request": request, "user": user, "error": "Имя организации обязательно", "form": {"name": ""}},
            status_code=400,
        )
    exists = db.query(Organization).filter(Organization.name == name).one_or_none()
    if exists:
        return templates.TemplateResponse(
            "admin/org_create.html",
            {"request": request, "user": user, "error": "Организация уже существует", "form": {"name": name}},
            status_code=400,
        )
    org = Organization(
        name=name,
        created_by_user_id=user.id,
        created_via="manual",
        artifact_level_id=_get_default_artifact_level_id(db),
    )
    db.add(org)
    db.flush()
    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="create",
        entity_type="organization",
        entity_id=str(org.id),
        after={"name": org.name, "created_via": getattr(org, "created_via", "")},
        request=request,
    )
    db.commit()
    return _redirect("/admin/orgs")


@router.get("/admin/orgs/{org_id}/edit", response_class=HTMLResponse)
def admin_orgs_edit_page(
    org_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> HTMLResponse:
    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")
    levels = _get_active_artifact_levels(db)
    t_uib = (
        db.query(OrgIndexKbTarget)
        .filter(OrgIndexKbTarget.org_id == int(org.id), OrgIndexKbTarget.sheet_name == UIB_SHEET_NAME)
        .one_or_none()
    )
    t_szi = (
        db.query(OrgIndexKbTarget)
        .filter(OrgIndexKbTarget.org_id == int(org.id), OrgIndexKbTarget.sheet_name == SZI_SHEET_NAME)
        .one_or_none()
    )
    return templates.TemplateResponse(
        "admin/org_edit.html",
        {
            "request": request,
            "user": user,
            "org": org,
            "error": None,
            "levels": levels,
            "target_uib": (float(t_uib.target_value) if t_uib else None),
            "target_szi": (float(t_szi.target_value) if t_szi else None),
            "audit_period_start": (org.audit_period_start.isoformat() if getattr(org, "audit_period_start", None) else ""),
            "audit_period_weeks": (int(org.audit_period_weeks) if getattr(org, "audit_period_weeks", None) else None),
        },
    )


@router.post("/admin/orgs/{org_id}/edit")
def admin_orgs_edit_save(
    org_id: int,
    request: Request,
    name: str = Form(...),
    artifact_level_id: str = Form(""),
    target_uib: str = Form(""),
    target_szi: str = Form(""),
    audit_period_start: str = Form(""),
    audit_period_weeks: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")
    name = name.strip()
    if not name:
        levels = _get_active_artifact_levels(db)
        return templates.TemplateResponse(
            "admin/org_edit.html",
            {
                "request": request,
                "user": user,
                "org": org,
                "error": "Имя организации обязательно",
                "levels": levels,
                "target_uib": (target_uib or "").strip(),
                "target_szi": (target_szi or "").strip(),
                "audit_period_start": (audit_period_start or "").strip(),
                "audit_period_weeks": (int(audit_period_weeks) if (audit_period_weeks or "").strip().isdigit() else None),
            },
            status_code=400,
        )
    exists = db.query(Organization).filter(Organization.name == name, Organization.id != org.id).one_or_none()
    if exists:
        levels = _get_active_artifact_levels(db)
        return templates.TemplateResponse(
            "admin/org_edit.html",
            {
                "request": request,
                "user": user,
                "org": org,
                "error": "Организация с таким именем уже существует",
                "levels": levels,
                "target_uib": (target_uib or "").strip(),
                "target_szi": (target_szi or "").strip(),
                "audit_period_start": (audit_period_start or "").strip(),
                "audit_period_weeks": (int(audit_period_weeks) if (audit_period_weeks or "").strip().isdigit() else None),
            },
            status_code=400,
        )

    def _parse_int_or_none(v: str) -> int | None:
        s = (v or "").strip()
        return int(s) if s.isdigit() else None

    def _parse_float_or_none(v: str) -> float | None:
        s = (v or "").strip()
        if not s:
            return None
        try:
            return float(s.replace(",", "."))
        except Exception:
            return None

    def _parse_date_or_none(v: str) -> date | None:
        s = (v or "").strip()
        if not s:
            return None
        try:
            return date.fromisoformat(s)
        except Exception:
            return None

    artifact_level_id_val = _parse_int_or_none(artifact_level_id)
    target_uib_val = _parse_float_or_none(target_uib)
    target_szi_val = _parse_float_or_none(target_szi)
    audit_start_val = _parse_date_or_none(audit_period_start)
    audit_weeks_val = _parse_int_or_none(audit_period_weeks)

    if artifact_level_id_val is not None and not db.get(ArtifactLevel, artifact_level_id_val):
        levels = _get_active_artifact_levels(db)
        return templates.TemplateResponse(
            "admin/org_edit.html",
            {
                "request": request,
                "user": user,
                "org": org,
                "error": "Некорректный уровень",
                "levels": levels,
                "target_uib": (target_uib or "").strip(),
                "target_szi": (target_szi or "").strip(),
                "audit_period_start": (audit_period_start or "").strip(),
                "audit_period_weeks": _parse_int_or_none(audit_period_weeks),
            },
            status_code=400,
        )

    # Validate targets if provided
    def _validate_target(v: float | None) -> bool:
        if v is None:
            return True
        return 0.0 <= float(v) <= 5.0

    if not _validate_target(target_uib_val) or not _validate_target(target_szi_val):
        levels = _get_active_artifact_levels(db)
        return templates.TemplateResponse(
            "admin/org_edit.html",
            {
                "request": request,
                "user": user,
                "org": org,
                "error": "Целевой показатель должен быть числом от 0 до 5",
                "levels": levels,
                "target_uib": (target_uib or "").strip(),
                "target_szi": (target_szi or "").strip(),
                "audit_period_start": (audit_period_start or "").strip(),
                "audit_period_weeks": _parse_int_or_none(audit_period_weeks),
            },
            status_code=400,
        )

    # Validate audit period: both fields must be set together, weeks must be 1/2/4.
    allowed_weeks = {1, 2, 4}
    if (audit_start_val is None) != (audit_weeks_val is None):
        levels = _get_active_artifact_levels(db)
        return templates.TemplateResponse(
            "admin/org_edit.html",
            {
                "request": request,
                "user": user,
                "org": org,
                "error": "Период аудита должен содержать и дату начала, и длительность",
                "levels": levels,
                "target_uib": (target_uib or "").strip(),
                "target_szi": (target_szi or "").strip(),
                "audit_period_start": (audit_period_start or "").strip(),
                "audit_period_weeks": audit_weeks_val,
            },
            status_code=400,
        )
    if audit_weeks_val is not None and audit_weeks_val not in allowed_weeks:
        levels = _get_active_artifact_levels(db)
        return templates.TemplateResponse(
            "admin/org_edit.html",
            {
                "request": request,
                "user": user,
                "org": org,
                "error": "Длительность периода аудита должна быть 1, 2 или 4 недели",
                "levels": levels,
                "target_uib": (target_uib or "").strip(),
                "target_szi": (target_szi or "").strip(),
                "audit_period_start": (audit_period_start or "").strip(),
                "audit_period_weeks": audit_weeks_val,
            },
            status_code=400,
        )

    def _get_target(sheet_name: str) -> float | None:
        row = (
            db.query(OrgIndexKbTarget)
            .filter(OrgIndexKbTarget.org_id == int(org.id), OrgIndexKbTarget.sheet_name == sheet_name)
            .one_or_none()
        )
        return float(row.target_value) if row else None

    def _set_target(sheet_name: str, v: float | None) -> None:
        row = (
            db.query(OrgIndexKbTarget)
            .filter(OrgIndexKbTarget.org_id == int(org.id), OrgIndexKbTarget.sheet_name == sheet_name)
            .one_or_none()
        )
        if v is None:
            if row:
                db.delete(row)
            return
        if row:
            row.target_value = float(v)
            return
        db.add(OrgIndexKbTarget(org_id=int(org.id), sheet_name=sheet_name, target_value=float(v)))

    before = {
        "name": org.name,
        "artifact_level_id": getattr(org, "artifact_level_id", None),
        "target_uib": _get_target(UIB_SHEET_NAME),
        "target_szi": _get_target(SZI_SHEET_NAME),
        "audit_period_start": (org.audit_period_start.isoformat() if getattr(org, "audit_period_start", None) else None),
        "audit_period_weeks": (int(org.audit_period_weeks) if getattr(org, "audit_period_weeks", None) else None),
    }
    org.name = name
    org.artifact_level_id = artifact_level_id_val
    org.audit_period_start = audit_start_val
    org.audit_period_weeks = audit_weeks_val
    _set_target(UIB_SHEET_NAME, target_uib_val)
    _set_target(SZI_SHEET_NAME, target_szi_val)
    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="update",
        entity_type="organization",
        entity_id=str(org.id),
        before=before,
        after={
            "name": org.name,
            "artifact_level_id": getattr(org, "artifact_level_id", None),
            "target_uib": _get_target(UIB_SHEET_NAME) if target_uib_val is None else float(target_uib_val),
            "target_szi": _get_target(SZI_SHEET_NAME) if target_szi_val is None else float(target_szi_val),
            "audit_period_start": (audit_start_val.isoformat() if audit_start_val else None),
            "audit_period_weeks": (int(audit_weeks_val) if audit_weeks_val is not None else None),
        },
        request=request,
    )
    db.commit()
    return _redirect("/admin/orgs")


@router.post("/admin/orgs/{org_id}/delete")
def admin_orgs_delete(
    org_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")
    try:
        before = {"name": org.name}
        db.delete(org)
        write_audit_log(
            db,
            actor=user,
            org_id=None,
            action="delete",
            entity_type="organization",
            entity_id=str(org_id),
            before=before,
            after=None,
            request=request,
        )
        db.commit()
        return _redirect("/admin/orgs")
    except IntegrityError:
        db.rollback()
        # Редирект на список с сообщением (не рендерим orgs.html напрямую, т.к. нужен сложный контекст)
        return _redirect(
            "/admin/orgs?"
            + urlencode(
                {
                    "err": "Нельзя удалить организацию: есть связанные данные (артефакты/пользователи/файлы).",
                }
            )
        )


@router.get("/admin/users", response_class=HTMLResponse)
def admin_users(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
    org_id: str | None = None,
    login: str | None = None,
    full_name: str | None = None,
    page: int = 1,
    page_size: int = 50,
    sort: str = "login",
    dir: str = "asc",
) -> HTMLResponse:
    page = max(int(page or 1), 1)
    page_size = int(page_size or 50)
    if page_size < 10:
        page_size = 10
    if page_size > 200:
        page_size = 200

    orgs = _filter_out_default_orgs(db.query(Organization).order_by(Organization.name.asc()).all())
    selected_org_id = int(org_id) if (org_id and str(org_id).isdigit()) else None
    if selected_org_id and selected_org_id not in {o.id for o in orgs}:
        selected_org_id = None
    selected_org_name = ""
    if selected_org_id:
        for o in orgs:
            if o.id == selected_org_id:
                selected_org_name = o.name
                break

    sort_key = (sort or "login").strip().lower()
    sort_dir = (dir or "asc").strip().lower()
    if sort_dir not in ("asc", "desc"):
        sort_dir = "asc"

    q = db.query(User)
    if selected_org_id:
        q = q.join(UserOrgMembership, UserOrgMembership.user_id == User.id).filter(UserOrgMembership.org_id == selected_org_id)

    login_q = (login or "").strip()
    full_name_q = (full_name or "").strip()
    if login_q:
        q = q.filter(User.login.ilike(f"%{login_q}%"))
    if full_name_q:
        q = q.filter(User.full_name.ilike(f"%{full_name_q}%"))

    if sort_key == "created_at":
        order_expr = User.created_at.asc() if sort_dir == "asc" else User.created_at.desc()
        q = q.order_by(order_expr, User.login.asc())
    elif sort_key == "is_admin":
        order_expr = User.is_admin.asc() if sort_dir == "asc" else User.is_admin.desc()
        q = q.order_by(order_expr, User.login.asc())
    elif sort_key == "is_active":
        order_expr = User.is_active.asc() if sort_dir == "asc" else User.is_active.desc()
        q = q.order_by(order_expr, User.login.asc())
    else:
        # login
        order_expr = User.login.asc() if sort_dir == "asc" else User.login.desc()
        q = q.order_by(order_expr)

    total = q.count()
    total_pages = max((total + page_size - 1) // page_size, 1)
    if page > total_pages:
        page = total_pages
    offset = (page - 1) * page_size
    users = q.offset(offset).limit(page_size).all()

    base_qd: dict[str, str] = {"page_size": str(page_size), "sort": sort_key, "dir": sort_dir}
    if selected_org_id:
        base_qd["org_id"] = str(selected_org_id)
    if login_q:
        base_qd["login"] = login_q
    if full_name_q:
        base_qd["full_name"] = full_name_q
    base_query = urlencode(base_qd)
    window = 3
    start = max(1, page - window)
    end = min(total_pages, page + window)
    page_links = list(range(start, end + 1))
    current_url = request.url.path + (f"?{request.url.query}" if request.url.query else "")

    resp = templates.TemplateResponse(
        "admin/users.html",
        {
            "request": request,
            "user": user,
            "container_class": "container-wide",
            "users": users,
            "orgs": orgs,
            "selected_org_id": selected_org_id,
            "selected_org_name": selected_org_name,
            "filters": {"login": login_q, "full_name": full_name_q},
            "error": None,
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
            "has_prev": page > 1,
            "has_next": offset + page_size < total,
            "page_links": page_links,
            "base_query": base_query,
            "current_url": current_url,
            "sort": sort_key,
            "dir": sort_dir,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@router.get("/admin/users/new", response_class=HTMLResponse)
def admin_users_create_page(
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> HTMLResponse:
    return templates.TemplateResponse(
        "admin/user_create.html",
        {"request": request, "user": user, "error": None, "form": {"login": "", "full_name": "", "is_admin": False}},
    )


@router.post("/admin/users")
def admin_users_create(
    request: Request,
    login: str = Form(...),
    password: str = Form(...),
    full_name: str = Form(""),
    is_admin: bool = Form(False),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    login = login.strip()
    if not login:
        return templates.TemplateResponse(
            "admin/user_create.html",
            {"request": request, "user": user, "error": "Логин обязателен", "form": {"login": "", "full_name": full_name, "is_admin": bool(is_admin)}},
            status_code=400,
        )
    exists = db.query(User).filter(User.login == login).one_or_none()
    if exists:
        return templates.TemplateResponse(
            "admin/user_create.html",
            {"request": request, "user": user, "error": "Логин уже используется", "form": {"login": login, "full_name": full_name, "is_admin": bool(is_admin)}},
            status_code=400,
        )

    pwd_err = _validate_password(password)
    if pwd_err:
        return templates.TemplateResponse(
            "admin/user_create.html",
            {"request": request, "user": user, "error": pwd_err, "form": {"login": login, "full_name": full_name, "is_admin": bool(is_admin)}},
            status_code=400,
        )
    new_user = User(
        login=login,
        password_hash="",
        full_name=full_name.strip(),
        is_active=True,
        is_admin=bool(is_admin),
    )
    new_user.password_hash = hash_password(password)
    db.add(new_user)
    db.flush()
    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="create",
        entity_type="user",
        entity_id=str(new_user.id),
        after={"login": new_user.login, "is_admin": bool(new_user.is_admin), "is_active": bool(new_user.is_active)},
        request=request,
    )
    db.commit()
    return _redirect("/admin/users")


@router.post("/admin/users/{user_id}/toggle_active")
def admin_users_toggle_active(
    user_id: int,
    request: Request,
    back: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if u.id == user.id:
        # Не даём заблокировать себя, чтобы не потерять доступ.
        return _redirect(back or "/admin/users")
    before = {"is_active": bool(u.is_active)}
    u.is_active = not bool(u.is_active)
    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="update",
        entity_type="user",
        entity_id=str(u.id),
        before=before,
        after={"is_active": bool(u.is_active)},
        request=request,
    )
    db.commit()
    return _redirect(back or "/admin/users")


@router.post("/admin/users/{user_id}/delete")
def admin_users_delete(
    user_id: int,
    request: Request,
    back: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if u.id == user.id:
        # Не даём удалить себя.
        return _redirect(back or "/admin/users")
    try:
        before = {"login": u.login, "full_name": u.full_name, "is_active": bool(u.is_active), "is_admin": bool(u.is_admin)}
        # Сначала удаляем роли/привязки к организациям (иначе ORM пытается проставить NULL в user_id).
        db.query(UserOrgMembership).filter(UserOrgMembership.user_id == u.id).delete(synchronize_session=False)
        db.delete(u)
        write_audit_log(
            db,
            actor=user,
            org_id=None,
            action="delete",
            entity_type="user",
            entity_id=str(user_id),
            before=before,
            after=None,
            request=request,
        )
        db.commit()
        return _redirect(back or "/admin/users")
    except IntegrityError:
        db.rollback()
        # Фоллбек: не падаем 500, а возвращаемся назад.
        return _redirect(back or "/admin/users")


@router.get("/admin/users/{user_id}/edit", response_class=HTMLResponse)
def admin_users_edit_page(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> HTMLResponse:
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    return templates.TemplateResponse("admin/user_edit.html", {"request": request, "user": user, "u": u, "error": None})


@router.post("/admin/users/{user_id}/edit")
def admin_users_edit_save(
    user_id: int,
    request: Request,
    full_name: str = Form(""),
    is_active: str = Form("true"),
    is_admin: str = Form("false"),
    new_password: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    before = {"full_name": u.full_name, "is_active": bool(u.is_active), "is_admin": bool(u.is_admin)}
    u.full_name = (full_name or "").strip()
    u.is_active = str(is_active).lower() == "true"
    u.is_admin = str(is_admin).lower() == "true"
    changed_password = False

    if new_password and new_password.strip():
        pwd_err = _validate_password(new_password)
        if pwd_err:
            return templates.TemplateResponse("admin/user_edit.html", {"request": request, "user": user, "u": u, "error": pwd_err}, status_code=400)
        u.password_hash = hash_password(new_password)
        changed_password = True

    write_audit_log(
        db,
        actor=user,
        org_id=None,
        action="update",
        entity_type="user",
        entity_id=str(u.id),
        before=before,
        after={"full_name": u.full_name, "is_active": bool(u.is_active), "is_admin": bool(u.is_admin), "password_changed": changed_password},
        request=request,
    )
    db.commit()
    return _redirect("/admin/users")


@router.get("/admin/memberships", response_class=HTMLResponse)
def admin_memberships(request: Request, db: Session = Depends(get_db), user: User = Depends(require_admin)) -> HTMLResponse:
    all_users = db.query(User).order_by(User.login.asc()).all()
    # "Системные" аккаунты (показываем отдельно): admin (is_admin) и служебный Auditor.
    system_users = [u for u in all_users if u.is_admin or (u.login or "").strip().lower() in ("admin", "auditor")]
    users = [u for u in all_users if u not in system_users]

    # Default организация — служебная (нужна системе), но в UI не показываем.
    orgs = db.query(Organization).filter(Organization.name != "Default").order_by(Organization.name.asc()).all()
    all_ms = (
        db.query(UserOrgMembership)
        .join(Organization, Organization.id == UserOrgMembership.org_id)
        .join(User, User.id == UserOrgMembership.user_id)
        .order_by(UserOrgMembership.created_at.desc())
        .all()
    )
    system_user_ids = {u.id for u in system_users}
    system_memberships = [m for m in all_ms if (m.org and m.org.name == "Default") or (m.user_id in system_user_ids)]
    org_memberships = [m for m in all_ms if m not in system_memberships and (m.org and m.org.name != "Default")]
    role_labels = {"admin": "Администратор", "auditor": "Аудитор", "customer": "Заказчик"}
    return templates.TemplateResponse(
        "admin/memberships.html",
        {
            "request": request,
            "user": user,
            "users": users,
            "system_users": system_users,
            "orgs": orgs,
            "system_memberships": system_memberships,
            "memberships": org_memberships,
            "roles": [r.value for r in Role],
            "role_labels": role_labels,
            "error": None,
        },
    )


@router.post("/admin/memberships")
def admin_memberships_create(
    request: Request,
    user_id: int | None = Form(None),
    org_id: int | None = Form(None),
    role: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> Response:
    def _render_error(msg: str, status_code: int = 400) -> HTMLResponse:
        all_users = db.query(User).order_by(User.login.asc()).all()
        system_users = [u for u in all_users if u.is_admin or (u.login or "").strip().lower() in ("admin", "auditor")]
        users = [u for u in all_users if u not in system_users]
        orgs = db.query(Organization).filter(Organization.name != "Default").order_by(Organization.name.asc()).all()
        all_ms = (
            db.query(UserOrgMembership)
            .join(Organization, Organization.id == UserOrgMembership.org_id)
            .join(User, User.id == UserOrgMembership.user_id)
            .order_by(UserOrgMembership.created_at.desc())
            .all()
        )
        system_user_ids = {u.id for u in system_users}
        system_memberships = [m for m in all_ms if (m.org and m.org.name == "Default") or (m.user_id in system_user_ids)]
        org_memberships = [m for m in all_ms if m not in system_memberships and (m.org and m.org.name != "Default")]
        role_labels = {"admin": "Администратор", "auditor": "Аудитор", "customer": "Заказчик"}
        return templates.TemplateResponse(
            "admin/memberships.html",
            {
                "request": request,
                "user": user,
                "users": users,
                "system_users": system_users,
                "orgs": orgs,
                "system_memberships": system_memberships,
                "memberships": org_memberships,
                "roles": [r.value for r in Role],
                "role_labels": role_labels,
                "error": msg,
            },
            status_code=status_code,
        )

    if not user_id:
        return _render_error("Выберите пользователя")
    if not org_id:
        return _render_error("Выберите организацию")

    # Не даём назначать роли на служебную Default организацию.
    org = db.get(Organization, org_id)
    if org and org.name == "Default":
        return _redirect("/admin/memberships")
    try:
        role_enum = Role(role)
    except ValueError:
        return _render_error("Некорректная роль")
    exists = (
        db.query(UserOrgMembership)
        .filter(UserOrgMembership.user_id == user_id, UserOrgMembership.org_id == org_id)
        .one_or_none()
    )
    if exists:
        before = {"role": exists.role.value}
        exists.role = role_enum
        write_audit_log(
            db,
            actor=user,
            org_id=org_id,
            action="update",
            entity_type="membership",
            entity_id=str(exists.id),
            before=before,
            after={"role": exists.role.value},
            request=request,
        )
    else:
        m = UserOrgMembership(user_id=user_id, org_id=org_id, role=role_enum)
        db.add(m)
        db.flush()
        write_audit_log(
            db,
            actor=user,
            org_id=org_id,
            action="create",
            entity_type="membership",
            entity_id=str(m.id),
            after={"user_id": user_id, "org_id": org_id, "role": m.role.value},
            request=request,
        )
    db.commit()
    return _redirect("/admin/memberships")


@router.get("/admin/orgs/{org_id}/users", response_class=HTMLResponse)
def admin_org_users(
    org_id: int,
    request: Request,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
) -> HTMLResponse:
    org = db.get(Organization, org_id)
    if not org:
        raise HTTPException(status_code=404, detail="Организация не найдена")

    memberships = (
        db.query(UserOrgMembership)
        .join(User, User.id == UserOrgMembership.user_id)
        .filter(UserOrgMembership.org_id == org_id)
        .order_by(User.login.asc())
        .all()
    )
    role_labels = {"admin": "Администратор", "auditor": "Аудитор", "customer": "Заказчик"}
    resp = templates.TemplateResponse(
        "admin/org_users.html",
        {
            "request": request,
            "user": user,
            "org": org,
            "memberships": memberships,
            "role_labels": role_labels,
        },
    )
    resp.headers["Cache-Control"] = "no-store, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp
